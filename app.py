"""
Employee Ticket System — Single File Architecture
Flask + PostgreSQL + Bootstrap 5
Phase 1: Models + DB + Templates + SLA Scheduler
"""

import os
import gzip
import logging
import threading
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta, timezone
from typing import Optional

# ── Timezone support (zoneinfo — stdlib since Python 3.9) ────────────────────
try:
    from zoneinfo import ZoneInfo
except ImportError:
    try:
        from backports.zoneinfo import ZoneInfo   # pip install backports.zoneinfo
    except ImportError:
        ZoneInfo = None   # fallback: UTC-only mode, no conversion
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature

from flask import Flask, render_template_string, redirect, url_for, flash, request, abort, send_file
from flask_mail import Mail, Message as MailMessage
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_wtf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from markupsafe import escape as _html_escape
import mimetypes
import re as _re  # used by @mention detection

# ── MIME detection: prefer python-magic (reads actual bytes), fallback to mimetypes ──
try:
    import magic as _magic
    _MAGIC_AVAILABLE = True
except ImportError:
    _MAGIC_AVAILABLE = False

# ── openpyxl (Excel export) — optional, checked at startup ──────────────────
try:
    import openpyxl as _openpyxl_check  # noqa: F401
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    import warnings
    warnings.warn("openpyxl is not installed — Excel export will be unavailable. Run: pip install openpyxl", RuntimeWarning)
import uuid as uuid_module
from sqlalchemy import extract
from sqlalchemy.orm import joinedload, selectinload
from apscheduler.schedulers.background import BackgroundScheduler

# ─────────────────────────────────────────────
# APP FACTORY
# ─────────────────────────────────────────────

app = Flask(__name__)

# Fix: Railway (and most PaaS platforms) sit behind a reverse proxy that terminates
# HTTPS and forwards requests over plain HTTP internally.  Without ProxyFix,
# url_for(_external=True) generates http:// links instead of https://, which breaks
# password-reset links, OAuth callbacks, and any absolute URL the app produces.
# ProxyFix reads the X-Forwarded-Proto / X-Forwarded-For headers that Railway injects
# and corrects the scheme and remote address Flask sees.
# x_for=1 / x_proto=1 / x_host=1 match Railway's single-proxy setup.
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# ─────────────────────────────────────────────
# CONFIGURATION CLASSES
# ─────────────────────────────────────────────

class BaseConfig:
    """Shared settings for all environments."""
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    WTF_CSRF_ENABLED               = True
    PERMANENT_SESSION_LIFETIME     = timedelta(hours=8)
    SESSION_COOKIE_HTTPONLY        = True
    SESSION_COOKIE_SAMESITE        = "Lax"
    SESSION_COOKIE_SECURE          = False   # overridden to True in ProductionConfig

    # Fix: Flask reads the full request body into memory before any route code runs.
    # Without this limit, a large upload can exhaust container RAM on Railway.
    # Set to 10 MB — matches MAX_FILE_BYTES defined in the attachment helpers below.
    MAX_CONTENT_LENGTH = 10 * 1024 * 1024   # 10 MB

    # Fix: Neon auto-pauses after inactivity. Connections kept alive in the pool
    # become stale and raise OperationalError on the next use.
    # pool_pre_ping=True issues a cheap "SELECT 1" before lending a connection,
    # transparently recycling any dead connections without surfacing errors to users.
    SQLALCHEMY_ENGINE_OPTIONS = {"pool_pre_ping": True}

    # ── Flask-Mail ───────────────────────────────────────────────────────────
    # Set these env vars in production.  When MAIL_SERVER is absent, email
    # sending is silently skipped so the app never crashes due to missing config.
    MAIL_SERVER      = os.environ.get("MAIL_SERVER", "")           # e.g. smtp.gmail.com
    MAIL_PORT        = int(os.environ.get("MAIL_PORT", 587))
    MAIL_USE_TLS     = os.environ.get("MAIL_USE_TLS", "true").lower() == "true"
    MAIL_USE_SSL     = os.environ.get("MAIL_USE_SSL", "false").lower() == "true"
    MAIL_USERNAME    = os.environ.get("MAIL_USERNAME", "")
    MAIL_PASSWORD    = os.environ.get("MAIL_PASSWORD", "")
    MAIL_TIMEOUT     = 5  # seconds — prevents long hangs when SMTP is blocked
    MAIL_DEFAULT_SENDER = os.environ.get(
        "MAIL_DEFAULT_SENDER",
        os.environ.get("MAIL_USERNAME", "noreply@ticketsystem.local"),
    )

    # Used by send_notification() to build absolute ticket URLs inside emails
    # that originate from background jobs (APScheduler — no active HTTP request).
    # Set to your Railway app URL in production, e.g.:
    #   APP_BASE_URL=https://ticket-system-production.up.railway.app
    # Without this, SLA-breach and assignment emails contain http://localhost links.
    APP_BASE_URL = os.environ.get("APP_BASE_URL", "")

    # ── Timezone for display & business-hours SLA calculation ────────────────
    # Railway servers always run in UTC.  Set APP_TIMEZONE to the IANA timezone
    # name matching the client's location so dates/times display correctly and
    # SLA business-hours windows are evaluated in local time rather than UTC.
    # Example values: "Africa/Cairo"  "Asia/Riyadh"  "Europe/London"
    # Default: "Africa/Cairo" (UTC+2/+3 — Egypt standard / daylight time).
    APP_TIMEZONE = os.environ.get("APP_TIMEZONE", "Africa/Cairo")


class DevelopmentConfig(BaseConfig):
    """Local development — Windows or Linux dev machine."""
    DEBUG                    = True
    SECRET_KEY               = os.environ.get("SECRET_KEY", "dev-secret-change-in-production")
    SQLALCHEMY_DATABASE_URI  = os.environ.get(
        "DATABASE_URL",
        "postgresql://postgres:admin123@localhost/Ticketing-system"
    )
    RATELIMIT_STORAGE_URI    = "memory://"   # single-process dev server — memory is fine
    # Dev upload & log paths (cross-platform)
    UPLOAD_FOLDER = os.path.join(os.path.expanduser("~"), "tickets_uploads")
    LOG_FOLDER    = ("C:\\tickets_logs" if os.name == "nt" else
                     os.path.join(os.path.expanduser("~"), "tickets_logs"))


class ProductionConfig(BaseConfig):
    """Production server (Linux + Gunicorn + Nginx)."""
    DEBUG                  = False
    SESSION_COOKIE_SECURE  = True    # requires HTTPS

    @property
    def SECRET_KEY(self):            # noqa: N802
        key = os.environ.get("SECRET_KEY")
        if not key:
            raise RuntimeError(
                "SECRET_KEY environment variable is not set. "
                "Generate one with: python3 -c \"import os,base64; "
                "print(base64.urlsafe_b64encode(os.urandom(32)).decode())\""
            )
        return key

    @property
    def SQLALCHEMY_DATABASE_URI(self):   # noqa: N802
        uri = os.environ.get("DATABASE_URL")
        if not uri:
            raise RuntimeError(
                "DATABASE_URL environment variable is not set."
            )
        # SQLAlchemy 1.4+ dropped the legacy "postgres://" scheme name.
        # Neon and some Railway setups deliver this exact form — normalize it
        # before SQLAlchemy sees it to avoid a NoSuchModuleError crash.
        if uri.startswith("postgres://"):
            uri = "postgresql://" + uri[len("postgres://"):]
        # psycopg2 rejects "sslmode=req" — Railway sometimes delivers this
        # truncated form. Expand it to the correct "sslmode=require".
        uri = uri.replace("sslmode=req&", "sslmode=require&").replace("sslmode=req", "sslmode=require")
        return uri

    @property
    def RATELIMIT_STORAGE_URI(self):     # noqa: N802
        """
        Gunicorn multi-workers CANNOT share in-memory state.
        A shared Redis backend is REQUIRED in production.
        pip install flask-limiter[redis]
        Set env var: REDIS_URL=redis://127.0.0.1:6379/0
        """
        uri = os.environ.get("REDIS_URL")
        if not uri:
            import warnings
            warnings.warn(
                "⚠ REDIS_URL is not set in production. "
                "Rate limiting will use in-memory storage — "
                "limits are NOT shared across Gunicorn workers. "
                "Set REDIS_URL=redis://127.0.0.1:6379/0 to fix this.",
                RuntimeWarning,
                stacklevel=3,
            )
            return "memory://"
        return uri

    UPLOAD_FOLDER = "/var/uploads/tickets"
    LOG_FOLDER    = "/tmp/tickets_logs"   # /tmp دايماً writable في أي container


# ── Select config based on FLASK_ENV ─────────
_env = os.environ.get("FLASK_ENV", "development").lower()
if _env == "production":
    _config = ProductionConfig()
else:
    _config = DevelopmentConfig()

app.config.from_object(_config)

# ── Dev safety warning ────────────────────────
if _env != "production":
    if app.config.get("SECRET_KEY") == "dev-secret-change-in-production":
        import warnings
        warnings.warn(
            "⚠ Using default fallback SECRET_KEY. "
            "Set the SECRET_KEY environment variable before any real usage.",
            stacklevel=1,
        )

# ── Rate Limiter ──────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],
    storage_uri=app.config.get("RATELIMIT_STORAGE_URI", "memory://"),
)

# ── Password Reset Token Helper ───────────────────────────────
def get_reset_serializer():
    return URLSafeTimedSerializer(app.config["SECRET_KEY"])

def generate_reset_token(email):
    s = get_reset_serializer()
    # Embed issued_at so we can compare against password_changed_at on verification
    payload = {"email": email, "iat": utc_now().isoformat()}
    return s.dumps(payload, salt="password-reset")

def verify_reset_token(token, expiration=3600):
    """
    Returns email string on success, None on failure.
    One-time use: if the user changed their password after this token was issued,
    the token is rejected — prevents reuse of intercepted/leaked tokens.
    """
    s = get_reset_serializer()
    try:
        payload = s.loads(token, salt="password-reset", max_age=expiration)
    except (SignatureExpired, BadSignature):
        return None
    email   = payload.get("email")
    iat_str = payload.get("iat")
    if not email or not iat_str:
        return None
    # One-time use check: reject if password was changed after token was issued
    user = User.query.filter_by(email=email).first()
    if user and user.password_changed_at:
        try:
            iat = datetime.fromisoformat(iat_str)
        except ValueError:
            return None
        if user.password_changed_at > iat:
            return None   # token already used — password was changed after issuance
    return email

# ── Username Policy ──────────────────────────────────────────────
_USERNAME_RE = _re.compile(r'^[a-zA-Z0-9_]{3,60}$')

def validate_username(username):
    """
    Returns error string or None.
    Allowed: 3–60 chars, letters / digits / underscore only.
    Prevents injection, spaces, and Arabic characters in login handle.
    """
    if not _USERNAME_RE.match(username):
        return "Username must be 3–60 characters and contain only letters, numbers, or underscores (_)"
    return None


# ── Password Policy ───────────────────────────────────────────
def validate_password(password):
    """Returns list of error strings; empty list = valid."""
    errors = []
    if len(password) < 10:
        errors.append(t("err_pw_min_len"))
    if not any(c.isupper() for c in password):
        errors.append(t("err_pw_upper"))
    if not any(c.isdigit() for c in password):
        errors.append(t("err_pw_digit"))
    if not any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password):
        errors.append(t("err_pw_special"))
    return errors


# ── Extensions ────────────────────────────────
db = SQLAlchemy(app)
migrate = Migrate(app, db)
csrf = CSRFProtect(app)

@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-Content-Type-Options"] = "nosniff"
    if not app.debug:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

login_manager = LoginManager(app)
login_manager.login_view = "auth.login"
login_manager.login_message = "Please log in first"
login_manager.login_message_category = "warning"
mail = Mail(app)

# ─────────────────────────────────────────────
# LANGUAGE / i18n  (EN default, AR optional)
# ─────────────────────────────────────────────

from flask import session as flask_session

TRANSLATIONS = {
    "en": {
        # nav / general
        "app_title":         "Ticket System",
        "dashboard":         "Dashboard",
        "all_tickets":       "All Tickets",
        "users":             "Users",
        "notifications":     "Notifications",
        "logout":            "Logout",
        "login":             "Login",
        "please_login":      "Please log in first",
        # dashboard employee
        "welcome":           "Welcome,",
        "total_tickets":     "Total Tickets",
        "open":              "Open",
        "in_progress":       "In Progress",
        "closed":            "Closed",
        "my_tickets":        "My Tickets",
        "assigned_to_me":    "Assigned to Me",
        "new_ticket":        "New Ticket",
        "ticket_number":     "Ticket #",
        "title":             "Title",
        "type":              "Type",
        "priority":          "Priority",
        "status":            "Status",
        "date":              "Date",
        "view":              "View",
        "no_tickets":        "No tickets yet",
        # new ticket
        "open_ticket":       "Open New Ticket",
        "section":           "Department",
        "choose_dept":       "-- Choose Department --",
        "assign_to":         "Assign To",
        "choose_dept_first": "-- Choose Department First --",
        "choose_assignee":   "-- Select Assignee --",
        "no_agents_in_dept": "-- No agents in this department --",
        "description":       "Description",
        "cancel":            "Cancel",
        "send_ticket":       "Submit Ticket",
        # ticket detail
        "comments":          "Comments",
        "no_comments":       "No comments yet",
        "add_comment":       "Add a comment…",
        "send":              "Send",
        "ticket_details":    "Ticket Details",
        "dept":              "Department",
        "assignee":          "Assignee",
        "unassigned":        "Unassigned",
        "sla_deadline":      "SLA Deadline",
        "actions":           "Actions",
        "change_status":     "Change Status",
        "save_changes":      "Save Changes",
        # admin dashboard
        "control_panel":     "Control Panel",
        "breached":          "SLA Breached",
        "resolved":          "Resolved",
        "critical_open":     "Critical (Open)",
        "unassigned_lbl":    "Unassigned",
        "urgent_tickets":    "Tickets Needing Immediate Action",
        "sla_breached_lbl":  "SLA Breached Tickets",
        "no_sla_breach":     "No SLA breaches",
        "no_urgent":         "No urgent tickets",
        "tickets_by_dept":   "Tickets by Department",
        "no_data":           "No data",
        "all_tickets_btn":   "All Tickets",
        # tickets list
        "all_tickets_title": "All Tickets",
        "filter":            "Filter",
        "clear":             "Clear",
        "all_statuses":      "All Statuses",
        "all_priorities":    "All Priorities",
        "all_depts":         "All Departments",
        "requester":         "Requester",
        "no_tickets_found":  "No tickets found",
        # users
        "manage_users":      "Users",
        "new_user":          "New User",
        "name":              "Name",
        "email":             "Email",
        "role":              "Role",
        "active":            "Active",
        "disabled":          "Disabled",
        "edit":              "Edit",
        # setup
        "setup_title":       "First-Run Setup",
        "setup_subtitle":    "Create the main admin account",
        "full_name":         "Full Name",
        "password":          "Password",
        "confirm_password":  "Confirm Password",
        "create_account":    "Create Account & Start",
        "one_time_page":     "This page appears only once",
        # login
        "login_title":       "Internal Ticket System",
        "login_subtitle":    "Sign in to continue",
        "sign_in":           "Sign In",
        # flash messages
        "flash_login_error": "Incorrect email, username or password",
        "flash_logged_out":  "Logged out successfully",
        "flash_ticket_ok":   "Ticket {num} opened successfully",
        "flash_update_ok":   "Ticket updated successfully",
        "flash_bulk_done":   "Bulk action applied to {count} ticket(s).",
        "flash_comment_ok":  "Comment added",
        "flash_user_ok":     "User {name} created",
        "flash_user_upd":    "User updated",
        # errors
        "err_403_title":     "403 — Forbidden",
        "err_403_msg":       "You don't have permission to access this page.",
        "err_404_title":     "404 — Page Not Found",
        "err_500_title":     "500 — Server Error",
        "err_500_body":      "Something went wrong on our end. Please try again or contact support.",
        "back_home":         "Back to Home",
        # notifications page
        "no_notifications":  "No notifications",
        # misc
        "na":                "—",
        # Password policy
        "pw_policy_title":   "Password Requirements",
        "pw_min_chars":      "At least 10 characters",
        "pw_uppercase":      "At least one uppercase letter (A-Z)",
        "pw_digit":          "At least one number (0-9)",
        "pw_special":        "At least one special character (!@#$%^&*...)",
        # Password reset
        "forgot_password":   "Forgot Password?",
        "reset_password":    "Reset Password",
        "reset_email_label": "Enter your account email",
        "send_reset_link":   "Send Reset Link",
        "new_password":      "New Password",
        "confirm_new_pw":    "Confirm New Password",
        "update_password":   "Update Password",
        "back_to_login":     "Back to Login",
        "reset_link_note":   "A reset link will be shown here (email delivery coming in v2)",
        # Username
        "username_lbl":      "Username",
        "email_or_username": "Email or username",
        "username_field":    "Username (optional — used for login)",
        "username_placeholder": "e.g. ahmed_it",
        "leave_blank_pw":    "leave blank to keep current",
        # Attachments
        "attachments":       "Attachments",
        "upload_file":       "Upload File",
        "no_attachments":    "No attachments",
        "download":          "Download",
        "file_too_large":    "File too large (max 10 MB)",
        "file_type_not_allowed": "File type not allowed (PDF, JPG, PNG, DOCX only)",
        "upload_ok":         "File uploaded successfully",
        "upload_error":      "Upload failed — please try again",
        # Reopen
        # 429 / Rate limit
        "err_429_title":     "429 — Too Many Requests",
        "err_429_msg":       "You have exceeded the allowed number of attempts. Please wait a few minutes and try again.",
        # Inline file-type hint (used in upload form)
        "file_types_hint":   "PDF, JPG, PNG, DOCX — max 10 MB",
        # SLA badge in ticket detail
        "sla_breached_badge":"SLA Breached",
        "reopen_ticket":     "Reopen Ticket",
        "flash_reopen_ok":   "Ticket reopened successfully",
        "reopen_not_allowed":"Only resolved or closed tickets can be reopened",
        # Audit log / History
        "audit_trail":       "Activity Log",
        "no_history":        "No activity recorded",
        "hist_actor":        "By",
        "hist_action":       "Action",
        "hist_old":          "Before",
        "hist_new":          "After",
        "hist_time":         "Time",
        "hist_sla":          "SLA",
        "hist_breached":     "Breached",
        "hist_ok":           "OK",
        "recent_activity":   "Recent Activity",
        "action_created":    "Ticket Created",
        "action_status_change": "Status Changed",
        "action_reassign":   "Reassigned",
        "action_comment_added": "Comment Added",
        "action_attachment_uploaded": "Attachment Uploaded",
        # Phase 4 — Reports & Search
        "reports":               "Reports",
        "reports_title":         "System Reports",
        "search":                "Search",
        "search_placeholder":    "Search tickets by title, description, or number…",
        "search_results":        "Search Results",
        "no_search_results":     "No tickets match your search",
        "search_query_label":    "Search Query",
        "rpt_total_by_type":     "Tickets by Type",
        "rpt_avg_resolution":    "Avg. Resolution Time (hours)",
        "rpt_per_agent":         "Tickets per Agent",
        "rpt_sla_compliance":    "SLA Compliance Rate",
        "rpt_overdue":           "Overdue Tickets",
        "rpt_dept":              "Department",
        "rpt_count":             "Count",
        "rpt_agent":             "Agent",
        "rpt_open":              "Open",
        "rpt_resolved":          "Resolved",
        "rpt_closed":            "Closed",
        "rpt_avg_hrs":           "Avg. Hours",
        "rpt_compliance_pct":    "Compliance %",
        "rpt_no_data":           "No data available",
        "rpt_ticket_type":       "Ticket Type",
        "rpt_resolved_within":   "Resolved Within SLA",
        "rpt_total_resolved":    "Total Resolved",
        "rpt_overdue_count":     "Overdue Count",
        "rpt_priority":          "Priority",
        "search_tip":            "Tip: searches title, description, and ticket number",
        # Department management
        "departments":           "Departments",
        "new_department":        "New Department",
        "edit_department":       "Edit Department",
        "dept_name_lbl":         "Department Name",
        "dept_manager_lbl":      "Department Manager (optional)",
        "no_manager":            "— No Manager —",
        "flash_dept_created":    "Department '{name}' created",
        "flash_dept_updated":    "Department updated",
        "flash_dept_deleted":    "Department deleted",
        "flash_dept_has_tickets":"Cannot delete — department has active tickets. Reassign them first.",
        "dept_already_exists":   "A department with that name already exists",
        # Delete ticket
        "delete_ticket":         "Delete Ticket",
        "flash_ticket_deleted":  "Ticket [{num}] moved to trash",
        "confirm_delete":        "Are you sure you want to delete this ticket?",
        # Admin filter assignee
        "all_assignees":         "All Assignees",
        # per-agent avg hours
        "rpt_avg_agent_hrs":     "Avg. Resolution (hrs)",
        # Availability
        "available":             "Available",
        "unavailable":           "Unavailable",
        "on_leave":              "On Leave",
        "not_on_leave":          "Not on Leave",
        "availability_on":       "You are available — click to set unavailable",
        "availability_off":      "You are unavailable — click to set available",
        "on_leave_note":         "On leave — contact Admin to clear",
        "auto_assigned":         "Auto-assigned",
        # ── Backup & Restore ─────────────────────────────────────────
        "backups":               "Backups",
        "backups_title":         "Backup & Restore",
        "backup_create":         "Create Backup Now",
        "backup_id":             "ID",
        "backup_date":           "Date",
        "backup_size":           "Size (KB)",
        "backup_source":         "Source",
        "backup_email":          "Email",
        "backup_drive":          "Drive",
        "backup_actions":        "Actions",
        "backup_source_auto":    "Auto",
        "backup_source_manual":  "Manual",
        "backup_email_sent":     "Sent ✓",
        "backup_email_none":     "Not sent",
        "backup_drive_saved":    "Saved ✓",
        "backup_drive_none":     "Not saved",
        "backup_download":       "Download",
        "backup_restore":        "Restore",
        "backup_confirm_restore":"Are you sure you want to restore this backup? ALL current data will be replaced.",
        "backup_no_records":     "No backups yet — click 'Create Backup Now' to start.",
        "flash_backup_created":  "Backup created successfully",
        "flash_backup_restored": "Backup restored successfully",
        "flash_backup_error":    "Backup failed — check server logs",
        "backup_delete":         "Delete",
        "backup_confirm_delete": "Are you sure you want to delete this backup? This action cannot be undone.",
        "flash_backup_deleted":  "Backup deleted successfully",
        "flash_delete_error":    "Delete failed — check server logs",
        "flash_restore_error":   "Restore failed — check server logs",
        # Ticket types (displayed in UI)
        "ttype_it_support":      "IT Support",
        "ttype_hr_request":      "HR Request",
        "ttype_complaint":       "Complaint",
        "ttype_general":         "General",
        # Ticket form validation
        "err_type_required":     "Please select a ticket type.",
        "err_dept_required":     "Please select a department.",
        # Cascade dropdown
        "choose_type":           "-- Choose Type --",
        "choose_dept_first_type":"-- Choose Department First --",
        "dept_types_hint":       "Leave all unchecked to allow all ticket types.",
        # ── Bulk action bar & deleted tickets table ─────────────────────────
        "sel_count_zero":        "0 selected",
        "choose_action":         "— Choose Action —",
        "waiting_customer":      "Waiting for Customer",
        "waiting_vendor":        "Waiting for Vendor",
        "reopened":              "Reopened",
        "apply":                 "Apply",
        "deleted_at_col":        "Deleted At",
        # ── Flash messages ────────────────────────────────────────────────────
        "err_fields_required":   "All fields are required",
        "err_email_taken":       "Email is already in use",
        "err_username_taken":    "Username is already taken",
        "err_pw_no_match":       "Passwords do not match",
        "err_pw_min_len":        "Password must be at least 10 characters",
        "err_pw_upper":          "Password must contain at least one uppercase letter",
        "err_pw_digit":          "Password must contain at least one number",
        "err_pw_special":        "Password must contain at least one special character (!@#$%^&*...)",
        "flash_reset_sent":      "Password reset link has been sent to your email.",
        "flash_reset_fallback":  "If that email exists in our system, a reset link has been sent.",
        "flash_reset_no_smtp":   "Email service is not configured. Please contact the administrator to reset your password.",
        "err_reset_invalid":     "The password reset link is invalid or has expired (1 hour limit).",
        "err_user_not_found":    "User not found.",
        "flash_pw_updated":      "Password updated successfully. Please log in.",
        "flash_welcome":         "Welcome {name}! Account created successfully. Please log in.",
        "err_invalid_priority":  "Invalid priority value.",
        "err_title_empty":       "Title cannot be empty.",
        "err_desc_empty":        "Description cannot be empty.",
        "err_ticket_create":     "Could not create ticket — please try again.",
        "err_attach_closed":     "Cannot add attachments to a closed ticket",
        "err_file_not_found":    "File not found",
        "err_invalid_request":   "Invalid request. Please try again.",
        "err_invalid_status":    "Invalid status value.",
        "err_invalid_assignee":  "Invalid assignee.",
        "err_assignee_inactive": "Selected assignee is not valid, inactive, or not an agent.",
        "err_csrf":              "CSRF validation failed.",
        "err_no_tickets_selected": "No tickets selected.",
        "err_no_valid_tickets":  "No valid tickets found.",
        "err_dept_name_required": "Department name is required.",
        "flash_ticket_restored": "Ticket [{num}] restored successfully.",
        "flash_dept_restored":   "Department \'{name}\' restored.",
        "err_openpyxl":          "openpyxl is not installed. Run: pip install openpyxl",
    },
    "ar": {
        "app_title":         "نظام التذاكر",
        "dashboard":         "لوحة التحكم",
        "all_tickets":       "كل التذاكر",
        "users":             "المستخدمون",
        "notifications":     "الإشعارات",
        "logout":            "تسجيل الخروج",
        "login":             "دخول",
        "please_login":      "يرجى تسجيل الدخول أولاً",
        "welcome":           "مرحباً،",
        "total_tickets":     "إجمالي التذاكر",
        "open":              "مفتوحة",
        "in_progress":       "قيد التنفيذ",
        "closed":            "مغلقة",
        "my_tickets":        "تذاكري",
        "assigned_to_me":    "المُعيَّنة لي",
        "new_ticket":        "تذكرة جديدة",
        "ticket_number":     "رقم التذكرة",
        "title":             "العنوان",
        "type":              "النوع",
        "priority":          "الأولوية",
        "status":            "الحالة",
        "date":              "التاريخ",
        "view":              "عرض",
        "no_tickets":        "لا توجد تذاكر بعد",
        "open_ticket":       "فتح تذكرة جديدة",
        "section":           "القسم",
        "choose_dept":       "-- اختر القسم --",
        "assign_to":         "تعيين إلى",
        "choose_dept_first": "-- اختر القسم أولاً --",
        "choose_assignee":   "-- اختر المُعيَّن --",
        "no_agents_in_dept": "-- لا يوجد وكلاء في هذا القسم --",
        "description":       "الوصف",
        "cancel":            "إلغاء",
        "send_ticket":       "إرسال التذكرة",
        "comments":          "التعليقات",
        "no_comments":       "لا توجد تعليقات بعد",
        "add_comment":       "أضف تعليقاً…",
        "send":              "إرسال",
        "ticket_details":    "تفاصيل التذكرة",
        "dept":              "القسم",
        "assignee":          "المسؤول",
        "unassigned":        "غير محدد",
        "sla_deadline":      "موعد SLA",
        "actions":           "إجراءات",
        "change_status":     "تغيير الحالة",
        "save_changes":      "حفظ التغييرات",
        "control_panel":     "لوحة التحكم",
        "breached":          "تجاوز SLA",
        "resolved":          "محلولة",
        "critical_open":     "حرجة (مفتوحة)",
        "unassigned_lbl":    "غير مُعيَّنة",
        "urgent_tickets":    "تذاكر تحتاج تدخل فوري",
        "sla_breached_lbl":  "تذاكر تجاوزت SLA",
        "no_sla_breach":     "لا توجد خروقات SLA",
        "no_urgent":         "لا توجد تذاكر عاجلة",
        "tickets_by_dept":   "التذاكر حسب القسم",
        "no_data":           "لا توجد بيانات",
        "all_tickets_btn":   "كل التذاكر",
        "all_tickets_title": "كل التذاكر",
        "filter":            "تصفية",
        "clear":             "مسح",
        "all_statuses":      "كل الحالات",
        "all_priorities":    "كل الأولويات",
        "all_depts":         "كل الأقسام",
        "requester":         "مقدّم الطلب",
        "no_tickets_found":  "لا توجد تذاكر",
        "manage_users":      "المستخدمون",
        "new_user":          "مستخدم جديد",
        "name":              "الاسم",
        "email":             "البريد الإلكتروني",
        "role":              "الدور",
        "active":            "نشط",
        "disabled":          "معطّل",
        "edit":              "تعديل",
        "setup_title":       "إعداد النظام — المرة الأولى",
        "setup_subtitle":    "أنشئ حساب المسؤول الرئيسي",
        "full_name":         "الاسم الكامل",
        "password":          "كلمة المرور",
        "confirm_password":  "تأكيد كلمة المرور",
        "create_account":    "إنشاء الحساب وبدء النظام",
        "one_time_page":     "هذه الصفحة تظهر مرة واحدة فقط",
        "login_title":       "نظام التذاكر الداخلي",
        "login_subtitle":    "سجّل دخولك للمتابعة",
        "sign_in":           "دخول",
        "flash_login_error": "البريد أو اسم المستخدم أو كلمة المرور غير صحيحة",
        "flash_logged_out":  "تم تسجيل الخروج بنجاح",
        "flash_ticket_ok":   "تم فتح التذكرة {num} بنجاح",
        "flash_update_ok":   "تم تحديث التذكرة بنجاح",
        "flash_bulk_done":   "تم تطبيق الإجراء الجماعي على {count} تذكرة.",
        "flash_comment_ok":  "تم إضافة التعليق",
        "flash_user_ok":     "تم إنشاء المستخدم {name}",
        "flash_user_upd":    "تم تحديث المستخدم",
        "err_403_title":     "403 — غير مسموح",
        "err_403_msg":       "ليس لديك صلاحية للوصول إلى هذه الصفحة.",
        "err_404_title":     "404 — الصفحة غير موجودة",
        "err_500_title":     "500 — خطأ في الخادم",
        "err_500_body":      "حدث خطأ من جانبنا. يرجى المحاولة مرة أخرى أو التواصل مع الدعم الفني.",
        "back_home":         "العودة للرئيسية",
        "no_notifications":  "لا توجد إشعارات",
        "na":                "—",
        # Password policy
        "pw_policy_title":   "متطلبات كلمة المرور",
        "pw_min_chars":      "10 أحرف على الأقل",
        "pw_uppercase":      "حرف كبير واحد على الأقل (A-Z)",
        "pw_digit":          "رقم واحد على الأقل (0-9)",
        "pw_special":        "رمز خاص واحد على الأقل (!@#$%^&*...)",
        # Password reset
        "forgot_password":   "نسيت كلمة المرور؟",
        "reset_password":    "إعادة تعيين كلمة المرور",
        "reset_email_label": "أدخل بريدك الإلكتروني",
        "send_reset_link":   "إرسال رابط الإعادة",
        "new_password":      "كلمة المرور الجديدة",
        "confirm_new_pw":    "تأكيد كلمة المرور الجديدة",
        "update_password":   "تحديث كلمة المرور",
        "back_to_login":     "العودة لتسجيل الدخول",
        "reset_link_note":   "سيظهر رابط الإعادة هنا (الإرسال بالبريد قادم في v2)",
        # Username
        "username_lbl":      "اسم المستخدم",
        "email_or_username": "الإيميل أو اسم المستخدم",
        "username_field":    "اسم المستخدم (اختياري — يُستخدم للدخول)",
        "username_placeholder": "مثال: ahmed_it",
        "leave_blank_pw":    "اتركه فارغاً للإبقاء على الحالي",
        # Attachments
        "attachments":       "المرفقات",
        "upload_file":       "رفع ملف",
        "no_attachments":    "لا توجد مرفقات",
        "download":          "تحميل",
        "file_too_large":    "الملف كبير جداً (الحد الأقصى 10 ميجا)",
        "file_type_not_allowed": "نوع الملف غير مسموح (PDF، JPG، PNG، DOCX فقط)",
        "upload_ok":         "تم رفع الملف بنجاح",
        "upload_error":      "فشل الرفع — يرجى المحاولة مرة أخرى",
        # Reopen
        # 429 / Rate limit
        "err_429_title":     "429 — طلبات كثيرة",
        "err_429_msg":       "لقد تجاوزت الحد المسموح من المحاولات. انتظر بضعة دقائق وحاول مجدداً.",
        # Inline file-type hint
        "file_types_hint":   "PDF، JPG، PNG، DOCX — الحد الأقصى 10 ميجا",
        # SLA badge
        "sla_breached_badge":"تجاوز SLA",
        "reopen_ticket":     "إعادة فتح التذكرة",
        "flash_reopen_ok":   "تمت إعادة فتح التذكرة بنجاح",
        "reopen_not_allowed":"يمكن إعادة فتح التذاكر المحلولة أو المغلقة فقط",
        # Audit log / History
        "audit_trail":       "سجل النشاط",
        "no_history":        "لا يوجد نشاط مسجّل",
        "hist_actor":        "بواسطة",
        "hist_action":       "الإجراء",
        "hist_old":          "قبل",
        "hist_new":          "بعد",
        "hist_time":         "الوقت",
        "hist_sla":          "SLA",
        "hist_breached":     "خرق",
        "hist_ok":           "سليم",
        "recent_activity":   "آخر النشاطات",
        "action_created":    "إنشاء تذكرة",
        "action_status_change": "تغيير الحالة",
        "action_reassign":   "إعادة تعيين",
        "action_comment_added": "إضافة تعليق",
        "action_attachment_uploaded": "رفع مرفق",
        # Phase 4 — Reports & Search
        "reports":               "التقارير",
        "reports_title":         "تقارير النظام",
        "search":                "بحث",
        "search_placeholder":    "ابحث بالعنوان أو الوصف أو رقم التذكرة…",
        "search_results":        "نتائج البحث",
        "no_search_results":     "لا توجد تذاكر تطابق بحثك",
        "search_query_label":    "نص البحث",
        "rpt_total_by_type":     "التذاكر حسب النوع",
        "rpt_avg_resolution":    "متوسط وقت الحل (ساعات)",
        "rpt_per_agent":         "التذاكر لكل موظف",
        "rpt_sla_compliance":    "نسبة الالتزام بـ SLA",
        "rpt_overdue":           "التذاكر المتأخرة",
        "rpt_dept":              "القسم",
        "rpt_count":             "العدد",
        "rpt_agent":             "الموظف",
        "rpt_open":              "مفتوحة",
        "rpt_resolved":          "محلولة",
        "rpt_closed":            "مغلقة",
        "rpt_avg_hrs":           "متوسط الساعات",
        "rpt_compliance_pct":    "نسبة الالتزام %",
        "rpt_no_data":           "لا توجد بيانات",
        "rpt_ticket_type":       "نوع التذكرة",
        "rpt_resolved_within":   "محلولة ضمن SLA",
        "rpt_total_resolved":    "إجمالي المحلولة",
        "rpt_overdue_count":     "عدد المتأخرة",
        "rpt_priority":          "الأولوية",
        "search_tip":            "تلميح: البحث يشمل العنوان والوصف ورقم التذكرة",
        # Department management
        "departments":           "الأقسام",
        "new_department":        "قسم جديد",
        "edit_department":       "تعديل القسم",
        "dept_name_lbl":         "اسم القسم",
        "dept_manager_lbl":      "مدير القسم (اختياري)",
        "no_manager":            "— بدون مدير —",
        "flash_dept_created":    "تم إنشاء القسم '{name}'",
        "flash_dept_updated":    "تم تحديث القسم",
        "flash_dept_deleted":    "تم حذف القسم",
        "flash_dept_has_tickets":"لا يمكن الحذف — يوجد تذاكر نشطة. يرجى إعادة تعيينها أولاً.",
        "dept_already_exists":   "يوجد قسم بهذا الاسم بالفعل",
        # Delete ticket
        "delete_ticket":         "حذف التذكرة",
        "flash_ticket_deleted":  "تم نقل التذكرة [{num}] إلى المهملات",
        "confirm_delete":        "هل أنت متأكد من حذف هذه التذكرة؟",
        # Admin filter assignee
        "all_assignees":         "كل المسؤولين",
        # per-agent avg hours
        "rpt_avg_agent_hrs":     "متوسط وقت الحل (ساعات)",
        # Availability
        "available":             "متاح",
        "unavailable":           "غير متاح",
        "on_leave":              "في إجازة",
        "not_on_leave":          "غير في إجازة",
        "availability_on":       "أنت متاح — اضغط لتعيين غير متاح",
        "availability_off":      "أنت غير متاح — اضغط لتعيين متاح",
        "on_leave_note":         "في إجازة رسمية — تواصل مع المسؤول لإلغائها",
        "auto_assigned":         "تعيين تلقائي",
        # ── النسخ الاحتياطي والاستعادة ────────────────────────────────
        "backups":               "النسخ الاحتياطي",
        "backups_title":         "النسخ الاحتياطي والاستعادة",
        "backup_create":         "إنشاء نسخة احتياطية الآن",
        "backup_id":             "رقم",
        "backup_date":           "التاريخ",
        "backup_size":           "الحجم (KB)",
        "backup_source":         "المصدر",
        "backup_email":          "البريد الإلكتروني",
        "backup_drive":          "درايف",
        "backup_actions":        "الإجراءات",
        "backup_source_auto":    "تلقائي",
        "backup_source_manual":  "يدوي",
        "backup_email_sent":     "تم الإرسال ✓",
        "backup_email_none":     "لم يُرسل",
        "backup_drive_saved":    "محفوظ ✓",
        "backup_drive_none":     "لم يُحفظ",
        "backup_download":       "تحميل",
        "backup_restore":        "استعادة",
        "backup_confirm_restore":"هل أنت متأكد من استعادة هذه النسخة؟ ستُستبدل جميع البيانات الحالية.",
        "backup_no_records":     "لا توجد نسخ احتياطية بعد — اضغط «إنشاء نسخة احتياطية الآن» للبدء.",
        "flash_backup_created":  "تم إنشاء النسخة الاحتياطية بنجاح",
        "flash_backup_restored": "تمت استعادة النسخة الاحتياطية بنجاح",
        "flash_backup_error":    "فشل إنشاء النسخة الاحتياطية — راجع سجل الخادم",
        "backup_delete":         "حذف",
        "backup_confirm_delete": "هل أنت متأكد من حذف هذه النسخة الاحتياطية؟ لا يمكن التراجع عن هذا الإجراء.",
        "flash_backup_deleted":  "تم حذف النسخة الاحتياطية بنجاح",
        "flash_delete_error":    "فشل الحذف — راجع سجل الخادم",
        "flash_restore_error":   "فشلت عملية الاستعادة — راجع سجل الخادم",
        # أنواع التذاكر (تُعرض في الواجهة)
        "ttype_it_support":      "طلب دعم تقني",
        "ttype_hr_request":      "طلب موارد بشرية",
        "ttype_complaint":       "شكوى",
        "ttype_general":         "عام",
        # التحقق من نموذج التذكرة
        "err_type_required":     "يرجى اختيار نوع التذكرة.",
        "err_dept_required":     "يرجى اختيار القسم.",
        # القائمة المتسلسلة
        "choose_type":           "-- اختر النوع --",
        "choose_dept_first_type":"-- اختر القسم أولاً --",
        "dept_types_hint":       "اتركها كلها فارغة للسماح بجميع أنواع التذاكر.",
        # ── Bulk action bar & deleted tickets table ─────────────────────────
        "sel_count_zero":        "0 محدد",
        "choose_action":         "— اختر إجراء —",
        "waiting_customer":      "في انتظار العميل",
        "waiting_vendor":        "في انتظار المورد",
        "reopened":              "معاد فتحها",
        "apply":                 "تطبيق",
        "deleted_at_col":        "تاريخ الحذف",
        # ── Flash messages ────────────────────────────────────────────────────
        "err_fields_required":   "جميع الحقول مطلوبة",
        "err_email_taken":       "البريد الإلكتروني مستخدم بالفعل",
        "err_username_taken":    "اسم المستخدم مستخدم بالفعل",
        "err_pw_no_match":       "كلمتا المرور غير متطابقتين",
        "err_pw_min_len":        "يجب أن تكون كلمة المرور 10 أحرف على الأقل",
        "err_pw_upper":          "يجب أن تحتوي كلمة المرور على حرف كبير واحد على الأقل",
        "err_pw_digit":          "يجب أن تحتوي كلمة المرور على رقم واحد على الأقل",
        "err_pw_special":        "يجب أن تحتوي كلمة المرور على رمز خاص واحد على الأقل (!@#$%^&*...)",
        "flash_reset_sent":      "تم إرسال رابط إعادة التعيين إلى بريدك الإلكتروني.",
        "flash_reset_fallback":  "إذا كان البريد موجوداً في النظام، تم إرسال رابط إعادة التعيين.",
        "flash_reset_no_smtp":   "خدمة البريد الإلكتروني غير مُعدَّة. يرجى التواصل مع المسؤول لإعادة تعيين كلمة المرور.",
        "err_reset_invalid":     "رابط إعادة التعيين غير صالح أو انتهت صلاحيته (حد أقصى ساعة).",
        "err_user_not_found":    "المستخدم غير موجود.",
        "flash_pw_updated":      "تم تحديث كلمة المرور بنجاح. يرجى تسجيل الدخول.",
        "flash_welcome":         "مرحباً {name}! تم إنشاء الحساب بنجاح. يرجى تسجيل الدخول.",
        "err_invalid_priority":  "قيمة الأولوية غير صالحة.",
        "err_title_empty":       "العنوان لا يمكن أن يكون فارغاً.",
        "err_desc_empty":        "الوصف لا يمكن أن يكون فارغاً.",
        "err_ticket_create":     "تعذّر إنشاء التذكرة — يرجى المحاولة مرة أخرى.",
        "err_attach_closed":     "لا يمكن إضافة مرفقات لتذكرة مغلقة",
        "err_file_not_found":    "الملف غير موجود",
        "err_invalid_request":   "طلب غير صالح. يرجى المحاولة مرة أخرى.",
        "err_invalid_status":    "قيمة الحالة غير صالحة.",
        "err_invalid_assignee":  "المُعيَّن غير صالح.",
        "err_assignee_inactive": "المُعيَّن المختار غير صالح أو غير نشط أو ليس وكيلاً.",
        "err_csrf":              "فشل التحقق الأمني. يرجى المحاولة مرة أخرى.",
        "err_no_tickets_selected": "لم يتم تحديد أي تذاكر.",
        "err_no_valid_tickets":  "لا توجد تذاكر صالحة.",
        "err_dept_name_required": "اسم القسم مطلوب.",
        "flash_ticket_restored": "تم استعادة التذكرة [{num}] بنجاح.",
        "flash_dept_restored":   "تم استعادة القسم \'{name}\' بنجاح.",
        "err_openpyxl":          "مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl",
    },
}

def get_lang():
    try:
        return flask_session.get("lang", "en")
    except RuntimeError:
        # Outside request context (e.g. unit tests, background jobs)
        return "en"

def t(key, **kwargs):
    lang = get_lang()
    text = TRANSLATIONS.get(lang, TRANSLATIONS["en"]).get(key, key)
    if kwargs:
        text = text.format(**kwargs)
    return text


# ── Jinja2 custom filter: parse JSON string safely in templates ───────────
import json as _json_filter
@app.template_filter("from_json")
def from_json_filter(value):
    """Safely parse a JSON string in templates; returns [] on failure."""
    if not value:
        return []
    try:
        return _json_filter.loads(value)
    except (ValueError, TypeError):
        return []


# ── Timezone helpers ──────────────────────────────────────────────────────
# All datetimes stored in the DB are naive UTC (tzinfo stripped intentionally
# to stay compatible with both PostgreSQL and SQLite).
# These helpers convert between UTC-naive and the app's configured local tz.

def _get_app_tz():
    """
    Return a ZoneInfo object for the configured APP_TIMEZONE.
    Falls back to UTC if zoneinfo is unavailable or the tz name is invalid.
    """
    tz_name = app.config.get("APP_TIMEZONE", "Africa/Cairo")
    if ZoneInfo is None:
        return timezone.utc
    try:
        return ZoneInfo(tz_name)
    except Exception:
        app.logger.warning(
            f"[TZ] Invalid APP_TIMEZONE={tz_name!r} — falling back to UTC."
        )
        return timezone.utc


def utc_now() -> datetime:
    """
    Return the current time as a naive UTC datetime (tzinfo=None).
    This is the canonical source of 'now' for all DB writes — replaces the
    scattered utc_now() calls.
    """
    return datetime.now(timezone.utc).replace(tzinfo=None)


def utc_to_local(dt: datetime) -> datetime:
    """
    Convert a naive UTC datetime (as stored in DB) to a naive local datetime
    in the app's configured timezone — used for display only.
    Returns dt unchanged if conversion fails.
    """
    if dt is None:
        return dt
    tz = _get_app_tz()
    if tz is timezone.utc:
        return dt   # no conversion needed
    try:
        return dt.replace(tzinfo=timezone.utc).astimezone(tz).replace(tzinfo=None)
    except Exception:
        return dt


def local_now() -> datetime:
    """
    Return the current time as a naive datetime in the app's local timezone.
    Used for display strings (e.g. report timestamps, backup filenames).
    """
    return utc_to_local(utc_now())


def local_to_utc(dt: datetime) -> datetime:
    """
    Convert a naive local datetime back to a naive UTC datetime.
    Used by add_business_hours() to ensure SLA calculation runs in local time.
    """
    if dt is None:
        return dt
    tz = _get_app_tz()
    if tz is timezone.utc:
        return dt
    try:
        return dt.replace(tzinfo=tz).astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        return dt


@app.template_filter("localtime")
def localtime_filter(dt, fmt="%Y-%m-%d %H:%M"):
    """
    Jinja2 filter: convert a naive UTC datetime (DB-stored) to the app's
    local timezone and return a formatted string.

    Usage in templates (replaces all .strftime() calls on datetime columns):
        {{ ticket.created_at | localtime }}
        {{ ticket.created_at | localtime('%Y-%m-%d') }}
        {{ ticket.sla_deadline | localtime }}
    """
    if dt is None:
        return "—"
    local_dt = utc_to_local(dt)
    try:
        return local_dt.strftime(fmt)
    except Exception:
        return str(dt)

# ── Jinja2 context processor: expose t() and lang to all templates ────────
@app.context_processor
def inject_i18n():
    return dict(t=t, lang=get_lang())


# ── Logging ───────────────────────────────────
log_dir = app.config.get("LOG_FOLDER", "/tmp/tickets_logs")
try:
    os.makedirs(log_dir, exist_ok=True)
    if not app.debug:
        handler = RotatingFileHandler(
            os.path.join(log_dir, "app.log"),
            maxBytes=10 * 1024 * 1024,
            backupCount=5,
        )
        handler.setLevel(logging.WARNING)
        app.logger.addHandler(handler)
except OSError:
    # لو المجلد مش writable (non-root container) — نكمل بدون file handler
    # Railway بتعرض اللوجز في الـ Dashboard بتاعتها من stdout تلقائياً
    pass


# ─────────────────────────────────────────────
# MODELS
# ─────────────────────────────────────────────

class Department(db.Model):
    __tablename__ = "departments"

    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)  # uniqueness enforced at app level (is_deleted-aware)
    manager_id = db.Column(
        db.Integer,
        db.ForeignKey("users.id", use_alter=True, name="fk_dept_manager_id"),
        nullable=True,
    )
    is_deleted = db.Column(db.Boolean, default=False, nullable=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: utc_now())
    # JSON-encoded list of allowed ticket types, e.g. '["IT Support","General"]'
    # NULL means all types are allowed (backward-compatible default)
    # NEW FIELD — run: flask db migrate -m "dept allowed_types" && flask db upgrade
    allowed_types = db.Column(db.Text, nullable=True)

    # relationships defined after User to avoid forward-ref issues
    users   = db.relationship("User",   back_populates="department",
                              foreign_keys="User.department_id", lazy="dynamic")
    tickets = db.relationship("Ticket", back_populates="department", lazy="dynamic")

    def __repr__(self):
        return f"<Department {self.name}>"


class User(UserMixin, db.Model):
    __tablename__ = "users"

    id              = db.Column(db.Integer, primary_key=True)
    name            = db.Column(db.String(100), nullable=False)
    username        = db.Column(db.String(60), nullable=True, unique=True, index=True)  # optional unique handle
    # ↑ NEW FIELD — run: flask db migrate -m "add username to users" && flask db upgrade
    email           = db.Column(db.String(150), nullable=False, unique=True)
    password_hash   = db.Column(db.String(255), nullable=False)
    role            = db.Column(db.String(20), nullable=False, default="employee")  # admin / manager / employee
    department_id   = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=True)
    active              = db.Column(db.Boolean, default=True, nullable=False)
    # Availability fields — used by auto-assign logic
    # on_leave:     set by Admin (vacation / official absence)
    # is_available: set by the engineer themselves (busy in meeting, etc.)
    # Auto-assign only picks users where active=True AND on_leave=False AND is_available=True
    # run: flask db migrate -m "add on_leave is_available to users" && flask db upgrade
    on_leave            = db.Column(db.Boolean, default=False, nullable=False)
    is_available        = db.Column(db.Boolean, default=True,  nullable=False)
    created_at          = db.Column(db.DateTime, default=lambda: utc_now())
    # Tracks last password change — reset tokens issued before this timestamp are invalid (one-time use)
    # run: flask db migrate -m "add password_changed_at to users" && flask db upgrade
    password_changed_at = db.Column(db.DateTime, nullable=True)

    department      = db.relationship("Department", back_populates="users",
                                      foreign_keys=[department_id])
    tickets_created = db.relationship("Ticket", back_populates="creator",
                                      foreign_keys="Ticket.created_by", lazy="dynamic")
    tickets_assigned= db.relationship("Ticket", back_populates="assignee",
                                      foreign_keys="Ticket.assigned_to", lazy="dynamic")
    comments        = db.relationship("Comment",       back_populates="author",    lazy="dynamic")
    history_actions = db.relationship("TicketHistory", back_populates="actor",     lazy="dynamic")
    notifications   = db.relationship("Notification",  back_populates="recipient", lazy="dynamic")

    # Flask-Login: deactivated users cannot log in
    def get_id(self):
        return str(self.id)

    @property
    def is_active(self):
        return self.active

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
        # Stamp timestamp so any outstanding reset token is invalidated (one-time use)
        self.password_changed_at = utc_now()

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f"<User {self.email} [{self.role}]>"


# ── SLA windows in business hours ────────────────────────────────────────────
# Critical = 2 real hours (24/7 — no business-hours filter applied)
# All others = business hours counted below
SLA_HOURS = {
    "Low":      5 * 8,   # 5 business days → 40 bh
    "Medium":   2 * 8,   # 2 business days → 16 bh
    "High":     8,       # same day        →  8 bh
    "Critical": 2,       # 2 real hours (always-on)
}

# ── Business-hours configuration ─────────────────────────────────────────────
# Adjust these to match the client's working schedule.
# WORK_DAYS: 0=Monday … 6=Sunday  (Egypt: Sat–Thu, weekends=Fri+Sat)
BUSINESS_START_HOUR = int(os.environ.get("BH_START", 9))   # 09:00
BUSINESS_END_HOUR   = int(os.environ.get("BH_END",   17))  # 17:00
BUSINESS_WORK_DAYS  = {
    int(d) for d in
    os.environ.get("BH_WORK_DAYS", "0,1,2,3,5,6").split(",")
}   # default: Mon Tue Wed Thu Sat Sun  (Friday off = Egyptian weekend)


def add_business_hours(start: datetime, hours: float) -> datetime:
    """
    Return a datetime that is `hours` business hours after `start`.

    Rules:
    - `start` is a naive UTC datetime (as stored in the DB).
    - It is first converted to the app's local timezone so that business-hours
      windows (BUSINESS_START_HOUR / BUSINESS_END_HOUR / BUSINESS_WORK_DAYS)
      are evaluated against local wall-clock time, not UTC.
      Without this, a ticket created at 10:00 Cairo time (= 07:00 UTC) would
      be treated as having been created before business hours open (09:00),
      making the SLA deadline wrong by up to 3 hours.
    - The result is converted back to naive UTC before being returned, so it
      can be stored / compared with other UTC datetimes consistently.
    - If `start` falls outside business hours it is snapped forward to the
      next business-hours opening before counting begins.
    - Critical tickets pass hours=2 and skip this function (they use raw time).
    - Works correctly across weekends, overnight boundaries, and DST transitions.
    """
    # ── Convert UTC-naive input to local-naive for BH arithmetic ─────────────
    local_start = utc_to_local(start)

    bh_seconds = int(hours * 3600)
    current = local_start

    def _is_work_day(dt: datetime) -> bool:
        return dt.weekday() in BUSINESS_WORK_DAYS

    def _day_open(dt: datetime) -> datetime:
        return dt.replace(hour=BUSINESS_START_HOUR, minute=0, second=0, microsecond=0)

    def _day_close(dt: datetime) -> datetime:
        return dt.replace(hour=BUSINESS_END_HOUR, minute=0, second=0, microsecond=0)

    # Snap to next business opening if needed
    if not _is_work_day(current) or current >= _day_close(current):
        # Move to next day's opening
        current += timedelta(days=1)
        current = _day_open(current)
        while not _is_work_day(current):
            current += timedelta(days=1)
        current = _day_open(current)
    elif current < _day_open(current):
        current = _day_open(current)

    remaining = bh_seconds
    while remaining > 0:
        close = _day_close(current)
        available = int((close - current).total_seconds())
        if available <= 0:
            # Already past closing — jump to next working day opening
            current += timedelta(days=1)
            current = _day_open(current)
            while not _is_work_day(current):
                current += timedelta(days=1)
            current = _day_open(current)
            continue

        if remaining <= available:
            current += timedelta(seconds=remaining)
            remaining = 0
        else:
            remaining -= available
            current = _day_open(current + timedelta(days=1))
            while not _is_work_day(current):
                current += timedelta(days=1)
            current = _day_open(current)

    # ── Convert local-naive result back to UTC-naive for storage/comparison ──
    return local_to_utc(current)


class Ticket(db.Model):
    __tablename__ = "tickets"

    id            = db.Column(db.Integer, primary_key=True)
    ticket_number = db.Column(db.String(20), unique=True, nullable=True, index=True)
    title         = db.Column(db.String(200), nullable=False)
    description   = db.Column(db.Text, nullable=False)
    type          = db.Column(db.String(50), nullable=False)    # IT / HR / Complaint / General
    priority      = db.Column(db.String(20), nullable=False)    # Low / Medium / High / Critical
    status        = db.Column(db.String(30), nullable=False, default="Open")
    created_by    = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    assigned_to   = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=True)
    created_at    = db.Column(db.DateTime, default=lambda: utc_now(), index=True)
    updated_at    = db.Column(db.DateTime, default=lambda: utc_now(), onupdate=lambda: utc_now())
    is_deleted    = db.Column(db.Boolean, default=False, nullable=False, index=True)
    deleted_at    = db.Column(db.DateTime, nullable=True)
    sla_breached  = db.Column(db.Boolean, default=False, nullable=False)

    creator    = db.relationship("User",       back_populates="tickets_created",
                                 foreign_keys=[created_by])
    assignee   = db.relationship("User",       back_populates="tickets_assigned",
                                 foreign_keys=[assigned_to])
    department = db.relationship("Department", back_populates="tickets")
    comments   = db.relationship("Comment",       back_populates="ticket",
                                 cascade="all, delete-orphan", lazy="dynamic")
    history    = db.relationship("TicketHistory", back_populates="ticket",
                                 lazy="dynamic")  # No cascade — audit log is protected from deletion
    attachments= db.relationship("Attachment",    back_populates="ticket",
                                 cascade="all, delete-orphan", lazy="dynamic")
    notifications = db.relationship("Notification", back_populates="ticket", lazy="dynamic")

    # ── Composite indexes (created via migration) ──────────────
    __table_args__ = (
        db.Index("idx_tickets_status_assigned", "status", "assigned_to"),
        db.Index("idx_tickets_is_deleted_status", "is_deleted", "status"),
        # NOTE (TC-141): The PostgreSQL GIN full-text index is intentionally NOT
        # declared here.  SQLAlchemy's postgresql_using='gin' only strips the
        # USING clause — it still emits a CREATE INDEX that calls to_tsvector(),
        # which SQLite cannot execute.  The index is created safely by
        # ensure_columns() at startup, which checks the dialect first.
    )

    @property
    def sla_deadline(self):
        hours = SLA_HOURS.get(self.priority, 40)
        if self.priority == "Critical":
            # Critical is always-on: count raw wall-clock hours, not business hours
            return self.created_at + timedelta(hours=hours)
        return add_business_hours(self.created_at, hours)

    @property
    def is_overdue(self):
        return (not self.sla_breached) and (utc_now() > self.sla_deadline) \
               and self.status not in ("Resolved", "Closed")

    def __repr__(self):
        return f"<Ticket {self.ticket_number} [{self.status}]>"


class Comment(db.Model):
    __tablename__ = "comments"

    id         = db.Column(db.Integer, primary_key=True)
    ticket_id  = db.Column(db.Integer, db.ForeignKey("tickets.id"), nullable=False)
    user_id    = db.Column(db.Integer, db.ForeignKey("users.id"),   nullable=False)
    body       = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: utc_now())

    ticket = db.relationship("Ticket", back_populates="comments")
    author = db.relationship("User",   back_populates="comments")

    def __repr__(self):
        return f"<Comment ticket={self.ticket_id} by={self.user_id}>"


class TicketHistory(db.Model):
    __tablename__ = "ticket_history"

    id          = db.Column(db.Integer, primary_key=True)
    ticket_id   = db.Column(db.Integer, db.ForeignKey("tickets.id"), nullable=False, index=True)
    changed_by  = db.Column(db.Integer, db.ForeignKey("users.id"),   nullable=False)
    action      = db.Column(db.String(50), nullable=False)   # status_change / reassign / comment_added / created
    old_value   = db.Column(db.String(200), nullable=True)
    new_value   = db.Column(db.String(200), nullable=True)
    sla_breached= db.Column(db.Boolean, default=False)
    created_at  = db.Column(db.DateTime, default=lambda: utc_now())

    ticket = db.relationship("Ticket", back_populates="history")
    actor  = db.relationship("User",   back_populates="history_actions")

    def __repr__(self):
        return f"<History ticket={self.ticket_id} action={self.action}>"


class Attachment(db.Model):
    __tablename__ = "attachments"

    id            = db.Column(db.Integer, primary_key=True)
    ticket_id     = db.Column(db.Integer, db.ForeignKey("tickets.id"), nullable=False)
    uploaded_by   = db.Column(db.Integer, db.ForeignKey("users.id"),   nullable=False)
    filename      = db.Column(db.String(255), nullable=True)      # UUID name on disk (legacy); None when stored in DB
    original_name = db.Column(db.String(255), nullable=False)     # shown to users
    file_size     = db.Column(db.Integer, nullable=False)
    mime_type     = db.Column(db.String(100), nullable=False)
    file_data     = db.Column(db.LargeBinary, nullable=True)      # file bytes stored in Neon (Railway deployment)
    created_at    = db.Column(db.DateTime, default=lambda: utc_now())

    ticket   = db.relationship("Ticket", back_populates="attachments")
    uploader = db.relationship("User")

    def __repr__(self):
        return f"<Attachment {self.original_name}>"


class Notification(db.Model):
    __tablename__ = "notifications"

    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey("users.id"),   nullable=False)
    ticket_id  = db.Column(db.Integer, db.ForeignKey("tickets.id"), nullable=True)
    message    = db.Column(db.String(255), nullable=False)
    is_read    = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: utc_now())

    recipient = db.relationship("User",   back_populates="notifications")
    ticket    = db.relationship("Ticket", back_populates="notifications")

    def __repr__(self):
        return f"<Notification user={self.user_id} read={self.is_read}>"


# ─────────────────────────────────────────────
# BACKUP MODEL
# ─────────────────────────────────────────────

class Backup(db.Model):
    __tablename__ = "backups"

    id         = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime,
                           default=lambda: utc_now(),
                           nullable=False)
    size_kb    = db.Column(db.Integer,     nullable=True)
    source     = db.Column(db.String(20),  default="auto",  nullable=False)  # auto / manual
    gdrive_id  = db.Column(db.String(200), nullable=True)   # Google Drive file ID; None if Drive not configured
    email_sent = db.Column(db.Boolean,     default=False,   nullable=False)  # True if backup email was sent successfully
    data       = db.Column(db.Text,        nullable=False)   # full JSON snapshot of all tables

    def __repr__(self):
        return f"<Backup id={self.id} source={self.source} size={self.size_kb}KB>"


# ─────────────────────────────────────────────
# TICKET COUNTER MODEL
# ─────────────────────────────────────────────

class TicketCounter(db.Model):
    """
    Monotonic per-year counter for ticket numbers.

    A single row per calendar year holds the last-issued sequence number.
    generate_ticket_number() atomically increments it with
    UPDATE ... RETURNING (works on both SQLite ≥ 3.35 and PostgreSQL),
    which is the only safe way to avoid duplicates under concurrent load —
    threading locks alone cannot protect across the generate→commit gap.
    """
    __tablename__ = "ticket_counter"

    year        = db.Column(db.Integer, primary_key=True)
    last_number = db.Column(db.Integer, default=0, nullable=False)

    def __repr__(self):
        return f"<TicketCounter year={self.year} last={self.last_number}>"


# ─────────────────────────────────────────────
# FLASK-LOGIN USER LOADER
# ─────────────────────────────────────────────

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def generate_ticket_number():
    """
    Generate TKT-YYYY-NNNN using an atomic database counter (TicketCounter table).

    Strategy
    --------
    A threading.Lock alone cannot prevent duplicate ticket numbers because it only
    protects the *generation* step.  Between lock-release and db.session.commit()
    another thread can read the same un-committed count and return the same number.

    The fix: delegate sequencing entirely to the database with an atomic
    UPDATE … RETURNING.  The DB engine serialises concurrent UPDATEs on the
    same row, so two threads/processes can never claim the same number.

    Works on:
      • SQLite ≥ 3.35  (released March 2021) — supports RETURNING
      • PostgreSQL      — native support for UPDATE … RETURNING

    Gaps in the sequence (e.g. after a rolled-back ticket creation) are
    acceptable; uniqueness is the only invariant that matters.
    """
    from sqlalchemy import text

    year = utc_now().year

    # ── Step 1: try atomic increment on existing row ──────────────────
    row = db.session.execute(
        text(
            "UPDATE ticket_counter "
            "SET last_number = last_number + 1 "
            "WHERE year = :y "
            "RETURNING last_number"
        ),
        {"y": year},
    ).fetchone()

    if row is None:
        # ── Step 2: first ticket of this year — create counter row ────
        # INSERT OR IGNORE (SQLite) / INSERT … ON CONFLICT DO NOTHING (PG)
        # then retry the UPDATE so we always get the RETURNING value.
        dialect = db.engine.dialect.name          # "sqlite" or "postgresql"
        if dialect == "sqlite":
            db.session.execute(
                text("INSERT OR IGNORE INTO ticket_counter (year, last_number) VALUES (:y, 0)"),
                {"y": year},
            )
        else:  # postgresql (and any other ANSI-SQL engine)
            db.session.execute(
                text(
                    "INSERT INTO ticket_counter (year, last_number) VALUES (:y, 0) "
                    "ON CONFLICT (year) DO NOTHING"
                ),
                {"y": year},
            )
        row = db.session.execute(
            text(
                "UPDATE ticket_counter "
                "SET last_number = last_number + 1 "
                "WHERE year = :y "
                "RETURNING last_number"
            ),
            {"y": year},
        ).fetchone()

    if row is None:
        # Extremely rare fallback — should never happen in practice
        import uuid
        return f"TKT-{year}-{uuid.uuid4().hex[:6].upper()}"

    return f"TKT-{year}-{row[0]:04d}"


def write_history(ticket, action, old_value, new_value, actor_id):
    """
    Write one row to ticket_history; calculate sla_breached at write time.

    Logic for calculating sla_breached:
    - We use new_value (the new status) not the current ticket.status,
      because this function is called *before* ticket.status is updated in the route.
    - If the action is status_change to Resolved or Closed:
      sla_breached = False (the SLA may have been breached earlier by the scheduler,
      but this history row records that the ticket was resolved — not a new breach).
    - In any other case: compare the current time against the deadline.
    """
    # Effective status to evaluate = new_value if status_change action, otherwise current ticket status
    effective_status = str(new_value) if action == "status_change" else ticket.status
    breached = (
        utc_now() > ticket.sla_deadline
        and effective_status not in ("Resolved", "Closed")
    )
    entry = TicketHistory(
        ticket_id    = ticket.id,
        changed_by   = actor_id,
        action       = action,
        old_value    = str(old_value) if old_value is not None else None,
        new_value    = str(new_value) if new_value is not None else None,
        sla_breached = breached,
    )
    db.session.add(entry)
    return entry


def send_email(to: str, subject: str, body_text: str, body_html: Optional[str] = None) -> bool:
    """
    Send an email via Flask-Mail.  Returns True on success, False on any error.

    Silently skipped (returns False without raising) if MAIL_SERVER is not configured —
    this keeps the app fully functional in environments without SMTP credentials.
    """
    if not app.config.get("MAIL_SERVER"):
        return False
    try:
        msg = MailMessage(
            subject=subject,
            recipients=[to],
            body=body_text,
            html=body_html,
        )
        mail.send(msg)
        return True
    except Exception as exc:
        app.logger.warning(f"[Email] Failed to send to {to!r}: {exc}")
        return False


# Events that warrant an email in addition to the in-app notification.
# Comment out any line to silence that specific category.
_EMAIL_EVENTS = {
    "assigned",       # ticket assigned to you
    "sla_breach",     # SLA breached
    "sla_escalation", # SLA escalation to manager / admin
    "resolved",       # ticket resolved
    "reopened",       # ticket reopened
    "mention",        # you were @mentioned in a comment
}


def send_notification(user_id, ticket_id, message, event: str = ""):
    """
    Write one in-app notification row and, for important events, send an email.

    Parameters
    ----------
    user_id   : int   — recipient user id
    ticket_id : int   — related ticket id  (may be None for system messages)
    message   : str   — notification text
    event     : str   — optional event category (see _EMAIL_EVENTS above)
    """
    notif = Notification(user_id=user_id, ticket_id=ticket_id, message=message)
    db.session.add(notif)

    # ── Optional email delivery ───────────────────────────────────────────────
    if event in _EMAIL_EVENTS and app.config.get("MAIL_SERVER"):
        try:
            recipient = db.session.get(User, user_id)
            if recipient and recipient.email:
                ticket_url = ""
                if ticket_id:
                    base_url = app.config.get("APP_BASE_URL", "")
                    if base_url:
                        with app.test_request_context(base_url=base_url):
                            ticket_url = url_for(
                                "employee.ticket_detail",
                                ticket_id=ticket_id,
                                _external=True,
                            )
                subject = f"[Ticket System] {message[:80]}"
                body_text = f"{message}\n\n{ticket_url}" if ticket_url else message
                body_html = (
                    f"<p>{message}</p>"
                    + (f'<p><a href="{ticket_url}">View ticket</a></p>' if ticket_url else "")
                )
                send_email(recipient.email, subject, body_text, body_html)
        except Exception as exc:
            app.logger.warning(f"[send_notification] email side-effect failed: {exc}")


def auto_assign_ticket(ticket):
    """
    Auto-assign a ticket to the most available agent in the ticket's department.

    Eligibility criteria (all three must be True):
        active=True       — user is still employed
        on_leave=False    — not on official leave (set by Admin)
        is_available=True — marked themselves as available (set by engineer)

    Tie-breaking: picks the eligible agent with the fewest open tickets
    (status not in Resolved/Closed) in any department — reflects true workload.

    Returns the assigned User object, or None if no eligible agent exists.
    If None: ticket stays Unassigned and the department Manager is notified.
    """
    if not ticket.department_id:
        return None

    open_statuses = ("Open", "In Progress", "Waiting for Customer",
                     "Waiting for Vendor", "Reopened")

    # All eligible agents in the same department.
    # Roles eligible for assignment: manager (department lead) + employee (support staff).
    # Admin is excluded — admins are system-level users, not department-level agents.
    # If no manager/employee is available, the function returns None and the department
    # manager is notified so they can assign manually.
    candidates = User.query.filter_by(
        department_id = ticket.department_id,
        active        = True,
        on_leave      = False,
        is_available  = True,
    ).filter(User.role.in_(["manager", "employee"])).all()

    if not candidates:
        # No one available — notify the department manager if exists
        dept = ticket.department
        if dept and dept.manager_id:
            send_notification(
                dept.manager_id, ticket.id,
                f"\u26a0 No available agents in {dept.name} — "
                f"ticket [{ticket.ticket_number}] is unassigned."
            )
        return None

    # Count open tickets per candidate (workload across all departments)
    def open_ticket_count(user):
        return user.tickets_assigned.filter(
            Ticket.status.in_(open_statuses),
            Ticket.is_deleted == False,
        ).count()

    chosen = min(candidates, key=open_ticket_count)
    return chosen


def process_mentions(comment_body, ticket, commenter_id):
    """
    Scan comment body for @username patterns.
    Resolve each to an active User by their UNIQUE username field,
    send an in-app notification.

    NOTE: searches by User.username (the short login handle, e.g. @ahmed_it),
    NOT by User.name (full name like "Ahmed Ali") — full names contain spaces
    and can never be matched by a single @word token.
    Returns list of mentioned user IDs.
    """
    mentioned_ids = []
    handles = _re.findall(r"@(\w+)", comment_body)
    for handle in set(handles):
        # Match against the unique username field (case-insensitive)
        user = User.query.filter(
            db.func.lower(User.username) == handle.lower(),
            User.active == True,
        ).first()
        if user and user.id != commenter_id:
            send_notification(
                user.id,
                ticket.id,
                f"You were mentioned in [{ticket.ticket_number}] by {current_user.name}",
            )
            mentioned_ids.append(user.id)
    return mentioned_ids


# ─────────────────────────────────────────────
# SLA BACKGROUND JOB (APScheduler)
# ─────────────────────────────────────────────

def check_sla_breaches():
    """
    Runs every 10 minutes.
    Flags SLA-breached tickets and:
      • Notifies the assignee (or creator)
      • Escalates to the department Manager
      • Escalates to all Admins if ticket is Critical

    Multi-worker safety (Gunicorn):
      On PostgreSQL, uses WITH FOR UPDATE SKIP LOCKED so that if multiple
      workers fire at the same time, each worker exclusively locks a different
      subset of rows — no ticket is processed twice, no duplicate notifications.
      On SQLite (dev/testing), FOR UPDATE is silently ignored; single-process
      dev server means there is no race to worry about.
    """
    with app.app_context():
        now = utc_now()
        open_statuses = ("Open", "In Progress", "Waiting for Customer", "Waiting for Vendor", "Reopened")

        is_postgres = db.engine.dialect.name == "postgresql"
        base_q = Ticket.query.filter(
            Ticket.is_deleted == False,
            Ticket.sla_breached == False,
            Ticket.status.in_(open_statuses),
        ).options(selectinload(Ticket.department))  # selectinload issues a separate SELECT after the main query —
        # avoids a JOIN so only the tickets table is locked by FOR UPDATE SKIP LOCKED.
        # joinedload here would generate SELECT tickets JOIN departments FOR UPDATE SKIP LOCKED,
        # which causes PostgreSQL to also lock department rows — unintended side effect.
        # SKIP LOCKED: each Gunicorn worker skips rows another worker is already
        # updating — prevents duplicate notifications and duplicate flag writes.
        if is_postgres:
            base_q = base_q.with_for_update(skip_locked=True)

        tickets = base_q.all()

        breached_count = 0
        for ticket in tickets:
            if now > ticket.sla_deadline:
                ticket.sla_breached = True
                breached_count += 1

                msg = (f"⚠ SLA Breached: [{ticket.ticket_number}] "
                       f"{ticket.title[:60]} — {ticket.priority} priority")

                # 1. Notify assignee or creator
                target = ticket.assigned_to or ticket.created_by
                send_notification(target, ticket.id, msg, event="sla_breach")

                # 2. Escalate to department Manager
                mgr_id = ticket.department.manager_id if ticket.department else None
                if mgr_id and mgr_id != target:
                    send_notification(mgr_id, ticket.id, f"[ESCALATION] {msg}", event="sla_escalation")

                # 3. Escalate all Admins if Critical
                if ticket.priority == "Critical":
                    admins = User.query.filter_by(role="admin", active=True).all()
                    for admin in admins:
                        if admin.id not in (target, mgr_id):
                            send_notification(admin.id, ticket.id,
                                              f"[CRITICAL ESCALATION] {msg}", event="sla_escalation")

        if breached_count:
            db.session.commit()
            app.logger.warning(f"SLA check: {breached_count} ticket(s) marked as breached.")


def check_waiting_for_customer_reminders():
    """
    Runs every 6 hours.
    Sends a reminder if a ticket has been 'Waiting for Customer' for 3+ days.
    Bumps updated_at after reminder so next alert fires 3 days later.

    Multi-worker safety: same SKIP LOCKED strategy as check_sla_breaches.
    Bumping updated_at inside the lock window means the second worker sees
    updated_at > threshold and skips the row naturally even without locks,
    but the lock guarantees atomicity during the current run.
    """
    with app.app_context():
        threshold = utc_now() - timedelta(days=3)

        is_postgres = db.engine.dialect.name == "postgresql"
        base_q = Ticket.query.filter(
            Ticket.is_deleted == False,
            Ticket.status == "Waiting for Customer",
            Ticket.updated_at <= threshold,
        )
        if is_postgres:
            base_q = base_q.with_for_update(skip_locked=True)

        stalled = base_q.all()

        reminded = 0
        for ticket in stalled:
            target = ticket.assigned_to or ticket.created_by
            msg = (f"⏰ Reminder: [{ticket.ticket_number}] has been "
                   f"'Waiting for Customer' for 3+ days. Please follow up.")
            send_notification(target, ticket.id, msg)
            # Bump updated_at to avoid repeat spam within the same 3-day window
            ticket.updated_at = utc_now()
            reminded += 1

        if reminded:
            db.session.commit()
            app.logger.info(f"Waiting-for-customer reminders sent: {reminded}")


scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(check_sla_breaches, "interval", minutes=10, id="sla_check")
scheduler.add_job(check_waiting_for_customer_reminders, "interval", hours=6, id="wfc_reminder")
scheduler.add_job(
    func=lambda: create_backup(source="auto"),
    trigger="cron",
    hour=22, minute=0,
    id="daily_backup",
    replace_existing=True,
)
# In debug mode, Werkzeug reloads twice (parent + child process).
# WERKZEUG_RUN_MAIN = "true" only in the actual child process.
# In production (no debug), run directly.
# This condition prevents the scheduler from running twice in development.
#
# Multi-worker safety (Gunicorn):
# Gunicorn --workers N creates N processes each importing app.py.
# With not app.debug = True in all workers, every worker would start its own
# scheduler → duplicate SLA notifications + duplicate daily backups.
#
# The Procfile for Railway must use --workers 1 --threads 4 (single process,
# multi-threaded) so the scheduler only runs once.  The SCHEDULER_ENABLED env
# var is a safety net: set it to "false" on any additional worker you add later.
_scheduler_enabled = os.environ.get("SCHEDULER_ENABLED", "true").lower() != "false"
if _scheduler_enabled and (os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug):
    scheduler.start()


# ─────────────────────────────────────────────
# BACKUP HELPERS
# ─────────────────────────────────────────────

def upload_to_gdrive(json_bytes: bytes, timestamp) -> Optional[str]:
    """
    Upload a JSON backup to Google Drive using OAuth2 refresh token.
    Returns the Google Drive file ID on success, or None if Drive is
    not configured or the upload fails (non-fatal — backup still saves
    to Neon regardless).

    Required environment variables:
        GDRIVE_CLIENT_ID      — OAuth2 client ID
        GDRIVE_CLIENT_SECRET  — OAuth2 client secret
        GDRIVE_REFRESH_TOKEN  — OAuth2 refresh token
        GDRIVE_FOLDER_ID      — ID of the target Drive folder
    """
    folder_id     = os.environ.get("GDRIVE_FOLDER_ID")
    client_id     = os.environ.get("GDRIVE_CLIENT_ID")
    client_secret = os.environ.get("GDRIVE_CLIENT_SECRET")
    refresh_token = os.environ.get("GDRIVE_REFRESH_TOKEN")
    if not all([folder_id, client_id, client_secret, refresh_token]):
        return None          # Google Drive not configured — silently skip
    try:
        from googleapiclient.discovery import build
        from google.oauth2.credentials import Credentials
        from googleapiclient.http import MediaInMemoryUpload

        creds = Credentials(
            token=None,
            refresh_token=refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=client_id,
            client_secret=client_secret,
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        service  = build("drive", "v3", credentials=creds)
        filename = f"backup_{utc_to_local(timestamp).strftime('%Y%m%d_%H%M%S')}.json.gz"
        media    = MediaInMemoryUpload(json_bytes, mimetype="application/gzip")
        file_meta = {"name": filename, "parents": [folder_id]}
        result = service.files().create(
            body=file_meta, media_body=media, fields="id"
        ).execute()
        return result.get("id")
    except Exception as exc:
        app.logger.error(f"[Backup] Google Drive upload failed: {exc}")
        return None


def create_backup(source: str = "auto") -> Backup | None:
    """
    Create a full JSON snapshot of all application data and persist it.

    Steps:
        1. Serialise every table row to a dict (using only primitive types).
        2. Dump to JSON and measure the byte size.
        3. Save a Backup record to Neon (always).
        4. Attempt to upload to Google Drive (optional — non-fatal if it fails).
        5. Delete Backup records older than 30 days to cap DB growth.

    Returns the new Backup instance, or None on error.
    Note: the daily_backup scheduler job calls this via a lambda so that
    the function reference is resolved at call-time, after it is defined.
    """
    import json as _json
    try:
        with app.app_context():
            now = utc_now()

            # ── 1. Serialise all tables ──────────────────────────────
            data = {
                "created_at": now.isoformat(),
                "version":    "1.0",
                "departments": [
                    {
                        "id":            d.id,
                        "name":          d.name,
                        "manager_id":    d.manager_id,
                        "is_deleted":    d.is_deleted,
                        "deleted_at":    d.deleted_at.isoformat() if d.deleted_at else None,
                        "created_at":    d.created_at.isoformat() if d.created_at else None,
                        "allowed_types": d.allowed_types,
                    }
                    for d in Department.query.all()
                ],
                "users": [
                    {
                        "id":                   u.id,
                        "name":                 u.name,
                        "username":             u.username,
                        "email":                u.email,
                        "password_hash":        u.password_hash,
                        "role":                 u.role,
                        "department_id":        u.department_id,
                        "active":               u.active,
                        "on_leave":             u.on_leave,
                        "is_available":         u.is_available,
                        "created_at":           u.created_at.isoformat() if u.created_at else None,
                        "password_changed_at":  u.password_changed_at.isoformat()
                                                if u.password_changed_at else None,
                    }
                    for u in User.query.all()
                ],
                "tickets": [
                    {
                        "id":            t.id,
                        "ticket_number": t.ticket_number,
                        "title":         t.title,
                        "description":   t.description,
                        "type":          t.type,
                        "priority":      t.priority,
                        "status":        t.status,
                        "created_by":    t.created_by,
                        "assigned_to":   t.assigned_to,
                        "department_id": t.department_id,
                        "created_at":    t.created_at.isoformat() if t.created_at else None,
                        "updated_at":    t.updated_at.isoformat() if t.updated_at else None,
                        "is_deleted":    t.is_deleted,
                        "deleted_at":    t.deleted_at.isoformat() if t.deleted_at else None,
                        "sla_breached":  t.sla_breached,
                    }
                    for t in Ticket.query.all()
                ],
                "comments": [
                    {
                        "id":         c.id,
                        "ticket_id":  c.ticket_id,
                        "user_id":    c.user_id,
                        "body":       c.body,
                        "created_at": c.created_at.isoformat() if c.created_at else None,
                    }
                    for c in Comment.query.all()
                ],
                "ticket_history": [
                    {
                        "id":          h.id,
                        "ticket_id":   h.ticket_id,
                        "changed_by":  h.changed_by,
                        "action":      h.action,
                        "old_value":   h.old_value,
                        "new_value":   h.new_value,
                        "sla_breached":h.sla_breached,
                        "created_at":  h.created_at.isoformat() if h.created_at else None,
                    }
                    for h in TicketHistory.query.all()
                ],
                "attachments": [
                    {
                        "id":            a.id,
                        "ticket_id":     a.ticket_id,
                        "uploaded_by":   a.uploaded_by,
                        "filename":      a.filename,
                        "original_name": a.original_name,
                        "file_size":     a.file_size,
                        "mime_type":     a.mime_type,
                        "created_at":    a.created_at.isoformat() if a.created_at else None,
                        # file_data stored as base64 so it survives JSON serialization.
                        # On Railway deployments filename=None and bytes live in file_data.
                        # Without this field restore produces records with both NULL —
                        # every download would return "File not found".
                        "file_data_b64": __import__("base64").b64encode(a.file_data).decode("ascii")
                                         if a.file_data is not None else None,
                    }
                    for a in Attachment.query.all()
                ],
                "notifications": [
                    {
                        "id":         n.id,
                        "user_id":    n.user_id,
                        "ticket_id":  n.ticket_id,
                        "message":    n.message,
                        "is_read":    n.is_read,
                        "created_at": n.created_at.isoformat() if n.created_at else None,
                    }
                    for n in Notification.query.all()
                ],
            }

            # ── 2. Dump to JSON then compress ───────────────────────
            json_str   = _json.dumps(data, ensure_ascii=False, indent=2)
            json_bytes = json_str.encode("utf-8")
            gzip_bytes = gzip.compress(json_bytes, compresslevel=6)
            size_kb      = len(json_bytes)  // 1024
            size_gz_kb   = len(gzip_bytes)  // 1024

            # ── 3. Save Backup record to Neon ────────────────────────
            backup = Backup(data=json_str, size_kb=size_kb, source=source)
            db.session.add(backup)
            db.session.flush()          # assign backup.id before Drive upload

            # ── 4. Upload to Google Drive (optional) ─────────────────
            gdrive_id = upload_to_gdrive(gzip_bytes, backup.created_at)
            if gdrive_id:
                backup.gdrive_id = gdrive_id

            db.session.commit()

            # ── 5. Prune backups older than 30 days ──────────────────
            cutoff = now - timedelta(days=30)
            old_backups = Backup.query.filter(Backup.created_at < cutoff).all()
            for old in old_backups:
                db.session.delete(old)
            if old_backups:
                db.session.commit()
                app.logger.info(f"[Backup] Pruned {len(old_backups)} old backup(s)")

            app.logger.info(
                f"[Backup] Created — id={backup.id} "
                f"size={size_kb}KB → compressed={size_gz_kb}KB "
                f"source={source} drive={'yes' if gdrive_id else 'no'}"
            )

            # ── 6. Send backup as email attachment (background thread) ──
            import threading
            backup_thread = threading.Thread(
                target=send_backup_email,
                args=(backup.id, json_bytes, gzip_bytes),
                daemon=True,
            )
            backup_thread.start()

            return backup

    except Exception as exc:
        app.logger.error(f"[Backup] create_backup failed: {exc}")
        db.session.rollback()
        return None


def send_backup_email(backup_id: int, json_bytes: bytes, gzip_bytes: bytes) -> None:
    """
    Send the backup as a gzip-compressed JSON attachment to BACKUP_MAIL_TO.
    Runs in a background thread — accepts backup_id (not the ORM object)
    to avoid detached-instance errors across thread boundaries.
    Silently skipped if MAIL_SERVER or BACKUP_MAIL_TO are not configured.

    The email attachment intentionally excludes file_data_b64 (binary
    attachment content) to keep the email size well within SMTP limits
    (Gmail = 25 MB, many servers = 10 MB).  The full backup including
    file bytes is always stored in Neon and can be restored from there.

    json_bytes  — uncompressed JSON (used to strip file_data_b64 before sending)
    gzip_bytes  — compressed bytes that become the .json.gz attachment
    """
    import json as _json_email
    backup_to = os.environ.get("BACKUP_MAIL_TO", "")
    if not app.config.get("MAIL_SERVER") or not backup_to:
        return

    with app.app_context():
        try:
            backup = db.session.get(Backup, backup_id)
            if not backup:
                return

            # ── Build a stripped copy: remove file_data_b64 from every
            #    attachment entry so the email stays small.
            #    Strip from json_bytes (plain JSON), then re-compress for
            #    the attachment so the recipient gets a clean .json.gz file.
            try:
                data_for_email = _json_email.loads(json_bytes.decode("utf-8"))
                for att in data_for_email.get("attachments", []):
                    att["file_data_b64"] = None   # strip binary — restore from Neon
                stripped_json_bytes = _json_email.dumps(
                    data_for_email, ensure_ascii=False, indent=2
                ).encode("utf-8")
                email_gz_bytes = gzip.compress(stripped_json_bytes, compresslevel=6)
            except Exception as _strip_err:
                # Fallback: if stripping fails for any reason, skip the email
                # rather than send an oversized attachment that SMTP will reject.
                app.logger.warning(
                    f"[Backup] Could not strip file_data_b64 for email "
                    f"(backup_id={backup_id}): {_strip_err} — email skipped."
                )
                return

            email_size_kb = len(email_gz_bytes) // 1024
            _backup_local = utc_to_local(backup.created_at)
            filename = f"backup_{_backup_local.strftime('%Y%m%d_%H%M%S')}.json.gz"
            subject  = f"[Ticket System] Backup — {_backup_local.strftime('%Y-%m-%d %H:%M')} ({backup.source})"
            body     = (
                f"نسخة احتياطية تلقائية من نظام التذاكر\n\n"
                f"التاريخ   : {_backup_local.strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"الحجم الكامل : {backup.size_kb} KB (محفوظ في Neon)\n"
                f"حجم الملف المرفق : {email_size_kb} KB (مضغوط، بدون محتوى المرفقات الثنائية)\n"
                f"المصدر    : {backup.source}\n\n"
                f"الملف المرفق يحتوي على نسخة مضغوطة من جميع البيانات الهيكلية.\n"
                f"محتوى المرفقات الثنائية محفوظ في قاعدة البيانات ويمكن استعادته منها.\n"
                f"احتفظ به في مكان آمن."
            )
            msg = MailMessage(
                subject=subject,
                recipients=[backup_to],
                body=body,
            )
            msg.attach(
                filename=filename,
                content_type="application/gzip",
                data=email_gz_bytes,
            )
            mail.send(msg)
            backup.email_sent = True
            db.session.commit()
            app.logger.info(
                f"[Backup] Email sent to {backup_to!r} — {filename} "
                f"({email_size_kb} KB compressed / {backup.size_kb} KB full)"
            )
        except Exception as exc:
            app.logger.warning(f"[Backup] Email send failed: {exc}")


def restore_from_backup(backup: Backup) -> None:
    """
    Restore the database from a Backup record.

    Strategy
    --------
    • Everything runs inside a single SQLAlchemy transaction.
      On any exception the session is rolled back — the live data
      is never partially overwritten.
    • We delete child rows before parent rows (respects FK constraints)
      then re-insert parent rows before child rows.
    • Original PKs from the snapshot are preserved so all FK references
      remain intact.
    • After inserting we reset each table's auto-increment sequence so
      the next INSERT gets a fresh ID above the highest restored one.
    • Attachment *records* are restored (filename/metadata) but the
      actual files on disk are not — they are flagged as unavailable.
    • Backup records themselves are never touched (they survive the restore).

    Raises
    ------
    Exception — propagated to the caller so the route can rollback & flash.
    """
    import json as _json

    data = _json.loads(backup.data)

    with db.session.begin_nested():   # savepoint — outer transaction wraps everything

        # ── DELETE in child-first order ──────────────────────────────
        db.session.execute(db.text("DELETE FROM notifications"))
        db.session.execute(db.text("DELETE FROM ticket_history"))
        db.session.execute(db.text("DELETE FROM attachments"))
        db.session.execute(db.text("DELETE FROM comments"))
        db.session.execute(db.text("DELETE FROM tickets"))
        db.session.execute(db.text("DELETE FROM users"))
        db.session.execute(db.text("DELETE FROM departments"))
        db.session.execute(db.text("DELETE FROM ticket_counter"))    # reset so counter stays in sync with restored ticket numbers

        # ── INSERT in parent-first order ─────────────────────────────

        # 1. Departments (manager_id FK points to users — insert without it first)
        for d in data.get("departments", []):
            db.session.execute(db.text(
                "INSERT INTO departments (id, name, manager_id, is_deleted, deleted_at, created_at, allowed_types) "
                "VALUES (:id, :name, NULL, :is_deleted, :deleted_at, :created_at, :allowed_types)"
            ), {
                "id":            d["id"],
                "name":          d["name"],
                "is_deleted":    d["is_deleted"],
                "deleted_at":    d.get("deleted_at"),
                "created_at":    d.get("created_at"),
                "allowed_types": d.get("allowed_types"),
            })

        # 2. Users
        for u in data.get("users", []):
            db.session.execute(db.text(
                "INSERT INTO users "
                "(id, name, username, email, password_hash, role, department_id, "
                " active, on_leave, is_available, created_at, password_changed_at) "
                "VALUES "
                "(:id, :name, :username, :email, :password_hash, :role, :department_id, "
                " :active, :on_leave, :is_available, :created_at, :password_changed_at)"
            ), {
                "id":                  u["id"],
                "name":                u["name"],
                "username":            u.get("username"),
                "email":               u["email"],
                "password_hash":       u["password_hash"],
                "role":                u["role"],
                "department_id":       u.get("department_id"),
                "active":              u["active"],
                "on_leave":            u.get("on_leave", False),
                "is_available":        u.get("is_available", True),
                "created_at":          u.get("created_at"),
                "password_changed_at": u.get("password_changed_at"),
            })

        # 3. Update departments.manager_id now that users exist
        for d in data.get("departments", []):
            if d.get("manager_id"):
                db.session.execute(db.text(
                    "UPDATE departments SET manager_id = :mid WHERE id = :id"
                ), {"mid": d["manager_id"], "id": d["id"]})

        # 4. Tickets
        for t in data.get("tickets", []):
            db.session.execute(db.text(
                "INSERT INTO tickets "
                "(id, ticket_number, title, description, type, priority, status, "
                " created_by, assigned_to, department_id, created_at, updated_at, "
                " is_deleted, deleted_at, sla_breached) "
                "VALUES "
                "(:id, :ticket_number, :title, :description, :type, :priority, :status, "
                " :created_by, :assigned_to, :department_id, :created_at, :updated_at, "
                " :is_deleted, :deleted_at, :sla_breached)"
            ), {
                "id":            t["id"],
                "ticket_number": t.get("ticket_number"),
                "title":         t["title"],
                "description":   t["description"],
                "type":          t["type"],
                "priority":      t["priority"],
                "status":        t["status"],
                "created_by":    t["created_by"],
                "assigned_to":   t.get("assigned_to"),
                "department_id": t.get("department_id"),
                "created_at":    t.get("created_at"),
                "updated_at":    t.get("updated_at"),
                "is_deleted":    t.get("is_deleted", False),
                "deleted_at":    t.get("deleted_at"),
                "sla_breached":  t.get("sla_breached", False),
            })

        # 5. Comments
        for c in data.get("comments", []):
            db.session.execute(db.text(
                "INSERT INTO comments (id, ticket_id, user_id, body, created_at) "
                "VALUES (:id, :ticket_id, :user_id, :body, :created_at)"
            ), {
                "id":         c["id"],
                "ticket_id":  c["ticket_id"],
                "user_id":    c["user_id"],
                "body":       c["body"],
                "created_at": c.get("created_at"),
            })

        # 6. TicketHistory
        for h in data.get("ticket_history", []):
            db.session.execute(db.text(
                "INSERT INTO ticket_history "
                "(id, ticket_id, changed_by, action, old_value, new_value, sla_breached, created_at) "
                "VALUES "
                "(:id, :ticket_id, :changed_by, :action, :old_value, :new_value, :sla_breached, :created_at)"
            ), {
                "id":          h["id"],
                "ticket_id":   h["ticket_id"],
                "changed_by":  h["changed_by"],
                "action":      h["action"],
                "old_value":   h.get("old_value"),
                "new_value":   h.get("new_value"),
                "sla_breached":h.get("sla_breached", False),
                "created_at":  h.get("created_at"),
            })

        # 7. Attachment records — restores both metadata and file bytes.
        # Backups created before this fix will not have "file_data_b64" in their JSON;
        # in that case file_data stays NULL (legacy behaviour, same as before).
        # Backups created after this fix carry the bytes as base64 and are fully restored.
        import base64 as _b64
        for a in data.get("attachments", []):
            raw_b64   = a.get("file_data_b64")
            file_bytes = _b64.b64decode(raw_b64) if raw_b64 is not None else None
            db.session.execute(db.text(
                "INSERT INTO attachments "
                "(id, ticket_id, uploaded_by, filename, original_name, file_size, mime_type, created_at, file_data) "
                "VALUES "
                "(:id, :ticket_id, :uploaded_by, :filename, :original_name, :file_size, :mime_type, :created_at, :file_data)"
            ), {
                "id":            a["id"],
                "ticket_id":     a["ticket_id"],
                "uploaded_by":   a["uploaded_by"],
                "filename":      a["filename"],
                "original_name": a["original_name"],
                "file_size":     a["file_size"],
                "mime_type":     a["mime_type"],
                "created_at":    a.get("created_at"),
                "file_data":     file_bytes,
            })

        # 8. Notifications
        for n in data.get("notifications", []):
            db.session.execute(db.text(
                "INSERT INTO notifications "
                "(id, user_id, ticket_id, message, is_read, created_at) "
                "VALUES "
                "(:id, :user_id, :ticket_id, :message, :is_read, :created_at)"
            ), {
                "id":         n["id"],
                "user_id":    n["user_id"],
                "ticket_id":  n.get("ticket_id"),
                "message":    n["message"],
                "is_read":    n["is_read"],
                "created_at": n.get("created_at"),
            })

        # ── Rebuild ticket_counter from restored ticket numbers ──────────────────
        # Parse every ticket_number (format TKT-YYYY-NNNN), find the highest
        # sequence number per year, and insert the correct counter rows.
        # Works with backups created before this counter table existed.
        import re as _re
        _counter_map: dict[int, int] = {}
        for _t in data.get("tickets", []):
            _m = _re.match(r"TKT-(\d{4})-(\d+)", _t.get("ticket_number") or "")
            if _m:
                _y, _n = int(_m.group(1)), int(_m.group(2))
                _counter_map[_y] = max(_counter_map.get(_y, 0), _n)
        for _year, _last in _counter_map.items():
            db.session.execute(
                db.text("INSERT INTO ticket_counter (year, last_number) VALUES (:y, :n)"),
                {"y": _year, "n": _last},
            )

        # ── Reset PostgreSQL sequences ───────────────────────────────
        # Only needed on PostgreSQL; silently skip on SQLite (dev).
        # Use db.engine.dialect.name for consistency with the rest of the codebase
        # (check_sla_breaches, reports, etc.) instead of a fragile string check on the URI.
        if db.engine.dialect.name == "postgresql":
            for table, col in [
                ("departments",   "id"),
                ("users",         "id"),
                ("tickets",       "id"),
                ("comments",      "id"),
                ("ticket_history","id"),
                ("attachments",   "id"),
                ("notifications", "id"),
            ]:
                db.session.execute(db.text(
                    f"SELECT setval(pg_get_serial_sequence('{table}', '{col}'), "
                    f"COALESCE((SELECT MAX({col}) FROM {table}), 0) + 1, false)"
                ))

    # Outer commit
    db.session.commit()


# ─────────────────────────────────────────────
# TEMPLATES (self-bootstrapping)
# ─────────────────────────────────────────────

TEMPLATES = {

# ── base.html ─────────────────────────────────
"templates/base.html": """<!DOCTYPE html>
<html lang="{{ 'ar' if lang == 'ar' else 'en' }}" dir="{{ 'rtl' if lang == 'ar' else 'ltr' }}">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{% block title %}{{ t('app_title') }}{% endblock %}</title>
  {% if lang == 'ar' %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.rtl.min.css">
  {% else %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
  {% endif %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
  <script src="https://unpkg.com/htmx.org@1.9.10" defer></script>
  <link rel="stylesheet" href="{{ url_for('static', filename='css/style.css') }}">
</head>
<body>

{% if current_user.is_authenticated %}
<nav class="navbar navbar-expand-lg navbar-dark bg-primary">
  <div class="container-fluid">
    <a class="navbar-brand fw-bold" href="{{ url_for('main.dashboard') }}">
      <i class="bi bi-ticket-perforated-fill me-1"></i> {{ t('app_title') }}
    </a>
    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navMain">
      <span class="navbar-toggler-icon"></span>
    </button>
    <div class="collapse navbar-collapse" id="navMain">
      <ul class="navbar-nav me-auto">
        <li class="nav-item">
          {% if current_user.role in ('admin', 'manager') %}
          <a class="nav-link" href="{{ url_for('admin.overview') }}">
          {% else %}
          <a class="nav-link" href="{{ url_for('main.dashboard') }}">
          {% endif %}
            <i class="bi bi-speedometer2"></i> {{ t('dashboard') }}
          </a>
        </li>
        {% if current_user.role in ('admin', 'manager') %}
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.tickets') }}">
            <i class="bi bi-list-task"></i> {{ t('all_tickets') }}
          </a>
        </li>
        {% endif %}
        {% if current_user.role == 'admin' %}
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.users') }}">
            <i class="bi bi-people-fill"></i> {{ t('users') }}
          </a>
        </li>
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.departments') }}">
            <i class="bi bi-diagram-3-fill"></i> {{ t('departments') }}
          </a>
        </li>
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.backups_list') }}">
            <i class="bi bi-shield-lock-fill"></i> {{ t('backups') }}
          </a>
        </li>
        {% endif %}
        {% if current_user.role in ('admin', 'manager') %}
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.reports') }}">
            <i class="bi bi-bar-chart-fill"></i> {{ t('reports') }}
          </a>
        </li>
        <li class="nav-item">
          <a class="nav-link" href="{{ url_for('admin.search') }}">
            <i class="bi bi-search"></i> {{ t('search') }}
          </a>
        </li>
        {% if current_user.role == 'admin' %}
        <li class="nav-item">
          <a class="nav-link text-danger" href="{{ url_for('admin.deleted_tickets') }}">
            <i class="bi bi-trash3"></i> Deleted
          </a>
        </li>
        {% endif %}
        {% endif %}
      </ul>
      <ul class="navbar-nav">
        <!-- Notification Badge -->
        {% set unread = unread_count %}
        <li class="nav-item me-2">
          <a class="nav-link position-relative" href="{{ url_for('main.notifications') }}">
            <i class="bi bi-bell-fill fs-5"></i>
            {% if unread > 0 %}
            <span class="position-absolute top-0 start-100 translate-middle badge rounded-pill bg-danger">
              {{ unread }}
            </span>
            {% endif %}
          </a>
        </li>
        <!-- Language Switcher -->
        <li class="nav-item me-2 d-flex align-items-center">
          {% if lang == 'ar' %}
          <a href="{{ url_for('main.set_lang', lang='en') }}" class="btn btn-sm btn-outline-light" title="Switch to English">EN</a>
          {% else %}
          <a href="{{ url_for('main.set_lang', lang='ar') }}" class="btn btn-sm btn-outline-light" title="التبديل للعربية">AR</a>
          {% endif %}
        </li>
        {% if current_user.role in ('admin', 'manager') %}
        <!-- Availability Toggle — only for agents who can be assigned tickets -->
        <li class="nav-item me-2 d-flex align-items-center">
          <form method="POST" action="{{ url_for('main.toggle_availability') }}" class="m-0">
            {{ csrf_token_input() | safe }}
            {% if current_user.is_available and not current_user.on_leave %}
            <button type="submit" class="btn btn-sm btn-success" title="{{ t('availability_on') }}">
              <i class="bi bi-circle-fill"></i> {{ t('available') }}
            </button>
            {% elif current_user.on_leave %}
            <span class="btn btn-sm btn-secondary disabled" title="{{ t('on_leave_note') }}">
              <i class="bi bi-calendar-x"></i> {{ t('on_leave') }}
            </span>
            {% else %}
            <button type="submit" class="btn btn-sm btn-outline-warning" title="{{ t('availability_off') }}">
              <i class="bi bi-circle"></i> {{ t('unavailable') }}
            </button>
            {% endif %}
          </form>
        </li>
        {% endif %}
        <li class="nav-item dropdown">
          <a class="nav-link dropdown-toggle" href="#" data-bs-toggle="dropdown">
            <i class="bi bi-person-circle"></i> {{ current_user.name }}
          </a>
          <ul class="dropdown-menu dropdown-menu-end">
            <li><span class="dropdown-item-text text-muted small">{{ current_user.email }}</span></li>
            <li><hr class="dropdown-divider"></li>
            <li>
              <a class="dropdown-item text-danger" href="{{ url_for('auth.logout') }}">
                <i class="bi bi-box-arrow-right"></i> {{ t('logout') }}
              </a>
            </li>
          </ul>
        </li>
      </ul>
    </div>
  </div>
</nav>
{% endif %}

<div class="container-fluid py-4 px-4">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
    <div class="alert alert-{{ cat }} alert-dismissible fade show" role="alert">
      {{ msg }}
      <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
    </div>
    {% endfor %}
  {% endwith %}

  {% block content %}{% endblock %}
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
""",

# ── login.html ─────────────────────────────────
"templates/login.html": """{% extends 'base.html' %}
{% block title %}{{ t('login') }}{% endblock %}
{% block content %}
<div class="row justify-content-center mt-5">
  <div class="col-md-5 col-lg-4">
    <div class="card shadow-sm border-0">
      <div class="card-body p-4">
        <div class="text-center mb-4">
          <i class="bi bi-ticket-perforated-fill text-primary" style="font-size:2.5rem"></i>
          <h4 class="mt-2 fw-bold">{{ t('login_title') }}</h4>
          <p class="text-muted small">{{ t('login_subtitle') }}</p>
        </div>
        <form method="POST" action="{{ url_for('auth.login') }}">
          {{ form.hidden_tag() }}
          <div class="mb-3">
            <label class="form-label">{{ t('email') }} / {{ t('username_lbl') }}</label>
            <input type="text" name="login_input" class="form-control" required autofocus
                   placeholder="{{ t('email_or_username') }}"
                   value="{{ request.form.get('login_input', '') }}">
          </div>
          <div class="mb-3">
            <label class="form-label">{{ t('password') }}</label>
            <div class="input-group">
              <input type="password" name="password" id="login_pw" class="form-control" required>
              <button type="button" class="btn btn-outline-secondary"
                      onclick="var e=document.getElementById('login_pw');var i=document.getElementById('login_pw_ico');if(e.type==='password'){e.type='text';i.className='bi bi-eye-slash';}else{e.type='password';i.className='bi bi-eye';}">
                <i id="login_pw_ico" class="bi bi-eye"></i>
              </button>
            </div>
          </div>
          <div class="d-grid mt-4">
            <button type="submit" class="btn btn-primary btn-lg">
              <i class="bi bi-box-arrow-in-right"></i> {{ t('sign_in') }}
            </button>
          </div>
          <div class="text-center mt-3">
            <a href="{{ url_for('auth.forgot_password') }}" class="text-muted small">
              {{ t('forgot_password') }}
            </a>
          </div>
        </form>
        <div class="text-center mt-3">
          {% if lang == 'ar' %}
          <a href="{{ url_for('main.set_lang', lang='en') }}" class="btn btn-sm btn-outline-secondary">Switch to English</a>
          {% else %}
          <a href="{{ url_for('main.set_lang', lang='ar') }}" class="btn btn-sm btn-outline-secondary">العربية</a>
          {% endif %}
        </div>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

# ── dashboard_employee.html ────────────────────
"templates/dashboard_employee.html": """{% extends 'base.html' %}
{% block title %}{{ t('dashboard') }}{% endblock %}
{% block content %}
<h4 class="fw-bold mb-4"><i class="bi bi-speedometer2 text-primary"></i> {{ t('welcome') }} {{ current_user.name }}</h4>

<div class="row g-3 mb-4">
  {% for label, count, color, icon in stats %}
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-{{ color }} bg-opacity-10">
          <i class="bi bi-{{ icon }} text-{{ color }} fs-4"></i>
        </div>
        <div>
          <div class="fs-4 fw-bold">{{ count }}</div>
          <div class="text-muted small">{{ label }}</div>
        </div>
      </div>
    </div>
  </div>
  {% endfor %}
</div>

<div class="d-flex justify-content-between align-items-center mb-3">
  <h5 class="fw-semibold mb-0">{{ t('my_tickets') }}</h5>
  <a href="{{ url_for('employee.new_ticket') }}" class="btn btn-primary btn-sm">
    <i class="bi bi-plus-lg"></i> {{ t('new_ticket') }}
  </a>
</div>

<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>{{ t('ticket_number') }}</th>
          <th>{{ t('title') }}</th>
          <th>{{ t('type') }}</th>
          <th>{{ t('priority') }}</th>
          <th>{{ t('status') }}</th>
          <th>{{ t('date') }}</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
        {% for tk in tickets.items %}
        <tr>
          <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
          <td>{{ tk.title|truncate(50) }}</td>
          <td>{{ tk.type }}</td>
          <td>
            <span class="badge bg-{{ priority_color(tk.priority) }}{{ ' priority-high' if tk.priority == 'High' else '' }}">{{ tk.priority }}</span>
          </td>
          <td>
            <span class="badge bg-{{ status_color(tk.status) }}">{{ tk.status }}</span>
            {% if tk.sla_breached %}<i class="bi bi-exclamation-triangle-fill text-danger ms-1" title="SLA Breached"></i>{% endif %}
          </td>
          <td class="text-muted small">{{ tk.created_at | localtime('%Y-%m-%d') }}</td>
          <td>
            <a href="{{ url_for('employee.ticket_detail', ticket_id=tk.id) }}"
               class="btn btn-outline-primary btn-sm">{{ t('view') }}</a>
          </td>
        </tr>
        {% else %}
        <tr><td colspan="7" class="text-center text-muted py-4">{{ t('no_tickets') }}</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% if tickets.pages > 1 %}
  <div class="card-footer d-flex justify-content-center">
    <nav>
      <ul class="pagination mb-0">
        {% for p in tickets.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2) %}
          {% if p %}
          <li class="page-item {% if p == tickets.page %}active{% endif %}">
            <a class="page-link" href="{{ url_for('main.dashboard', page=p) }}">{{ p }}</a>
          </li>
          {% else %}
          <li class="page-item disabled"><span class="page-link">…</span></li>
          {% endif %}
        {% endfor %}
      </ul>
    </nav>
  </div>
  {% endif %}
</div>

{# ── Tickets assigned to this employee ── #}
{% if assigned_tickets.total > 0 %}
<div class="d-flex justify-content-between align-items-center mb-3 mt-5">
  <h5 class="fw-semibold mb-0">
    <i class="bi bi-person-check text-info"></i> {{ t('assigned_to_me') }}
    <span class="badge bg-info ms-1">{{ assigned_tickets.total }}</span>
  </h5>
</div>
<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>{{ t('ticket_number') }}</th>
          <th>{{ t('title') }}</th>
          <th>{{ t('type') }}</th>
          <th>{{ t('priority') }}</th>
          <th>{{ t('status') }}</th>
          <th>{{ t('date') }}</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
        {% for tk in assigned_tickets.items %}
        <tr>
          <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
          <td>{{ tk.title|truncate(50) }}</td>
          <td>{{ tk.type }}</td>
          <td>
            <span class="badge bg-{{ priority_color(tk.priority) }}{{ ' priority-high' if tk.priority == 'High' else '' }}">{{ tk.priority }}</span>
          </td>
          <td>
            <span class="badge bg-{{ status_color(tk.status) }}">{{ tk.status }}</span>
            {% if tk.sla_breached %}<i class="bi bi-exclamation-triangle-fill text-danger ms-1" title="SLA Breached"></i>{% endif %}
          </td>
          <td class="text-muted small">{{ tk.created_at | localtime('%Y-%m-%d') }}</td>
          <td>
            <a href="{{ url_for('employee.ticket_detail', ticket_id=tk.id) }}"
               class="btn btn-outline-info btn-sm">{{ t('view') }}</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% if assigned_tickets.pages > 1 %}
  <div class="card-footer d-flex justify-content-center">
    <nav>
      <ul class="pagination mb-0">
        {% for p in assigned_tickets.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2) %}
          {% if p %}
          <li class="page-item {% if p == assigned_tickets.page %}active{% endif %}">
            <a class="page-link" href="{{ url_for('main.dashboard', apage=p) }}">{{ p }}</a>
          </li>
          {% else %}
          <li class="page-item disabled"><span class="page-link">…</span></li>
          {% endif %}
        {% endfor %}
      </ul>
    </nav>
  </div>
  {% endif %}
</div>
{% endif %}
{% endblock %}
""",

# ── new_ticket.html ────────────────────────────
"templates/new_ticket.html": """{% extends 'base.html' %}
{% block title %}{{ t('new_ticket') }}{% endblock %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-lg-7">
    <div class="card border-0 shadow-sm">
      <div class="card-header bg-primary text-white fw-semibold d-flex justify-content-between align-items-center">
        <span><i class="bi bi-plus-circle"></i> {{ t('open_ticket') }}</span>
        <span id="ticket-number-preview"
              hx-get="{{ url_for('api.preview_ticket_number') }}"
              hx-trigger="load"
              hx-swap="innerHTML">
          <span class="badge bg-secondary bg-opacity-50 fs-6 font-monospace">
            <i class="bi bi-hourglass-split"></i> …
          </span>
        </span>
      </div>
      <div class="card-body p-4">
        <form method="POST" action="{{ url_for('employee.new_ticket') }}" enctype="multipart/form-data">
          {{ form.hidden_tag() }}

          <div class="mb-3">
            <label class="form-label">{{ t('title') }} <span class="text-danger">*</span></label>
            <input type="text" name="title" class="form-control" required
                   value="{{ request.form.get('title', '') }}">
          </div>

          {# ── Step 1: Department (mandatory — drives the type cascade) ── #}
          <div class="mb-3">
            <label class="form-label">{{ t('section') }} <span class="text-danger">*</span></label>
            {% if current_user.role in ('admin', 'manager') %}
            <select name="department_id" id="dept_select" class="form-select" required
                    hx-get="{{ url_for('api.dept_ticket_types') }}"
                    hx-target="#type_select"
                    hx-trigger="change"
                    hx-include="[name='department_id']"
                    hx-on:change="htmx.ajax('GET','{{ url_for('api.dept_employees_by_select') }}',{target:'#assigned_to_select',values:{department_id:this.value}})">
              <option value="">{{ t('choose_dept') }}</option>
              {% for d in departments %}
              <option value="{{ d.id }}"
                {% if request.form.get('department_id')|int == d.id %}selected{% endif %}>
                {{ d.name }}
              </option>
              {% endfor %}
            </select>
            {% else %}
            <select name="department_id" id="dept_select" class="form-select" required
                    hx-get="{{ url_for('api.dept_ticket_types') }}"
                    hx-target="#type_select"
                    hx-trigger="change"
                    hx-include="[name='department_id']">
              <option value="">{{ t('choose_dept') }}</option>
              {% for d in departments %}
              <option value="{{ d.id }}"
                {% if request.form.get('department_id')|int == d.id %}selected
                {% elif not request.form.get('department_id') and current_user.department_id == d.id %}selected
                {% endif %}>
                {{ d.name }}
              </option>
              {% endfor %}
            </select>
            {% endif %}
          </div>

          {# ── Step 2: Type (mandatory — populated via HTMX after dept chosen) ── #}
          <div class="row g-3 mb-3">
            <div class="col-md-6">
              <label class="form-label">{{ t('type') }} <span class="text-danger">*</span></label>
              <select name="type" id="type_select" class="form-select" required>
                {% if request.form.get('type') %}
                  {# Re-render full list on POST-back so the selected value is preserved #}
                  <option value="">{{ t('choose_type') }}</option>
                  {% for tt in ticket_types %}
                  <option value="{{ tt }}"
                    {% if request.form.get('type') == tt %}selected{% endif %}>
                    {{ t(ticket_type_i18n.get(tt, tt)) }}
                  </option>
                  {% endfor %}
                {% else %}
                  <option value="">{{ t('choose_dept_first_type') }}</option>
                {% endif %}
              </select>
            </div>
            {% if current_user.role in ('admin', 'manager') %}
            <div class="col-md-6">
              <label class="form-label">{{ t('priority') }}</label>
              <select name="priority" class="form-select">
                {% for p in priorities %}
                <option value="{{ p }}"
                  {% if request.form.get('priority') == p %}selected{% endif %}>{{ p }}</option>
                {% endfor %}
              </select>
            </div>
            {% else %}
            {# Employees always submit Low — priority is set by manager after review #}
            <input type="hidden" name="priority" value="Low">
            {% endif %}
          </div>

          {% if current_user.role in ('admin', 'manager') %}
          <div class="mb-3">
            <label class="form-label">{{ t('assign_to') }}</label>
            <select id="assigned_to_select" name="assigned_to" class="form-select">
              <option value="">{{ t('choose_dept_first') }}</option>
            </select>
          </div>
          {% endif %}

          <div class="mb-3">
            <label class="form-label">{{ t('description') }} <span class="text-danger">*</span></label>
            <textarea name="description" class="form-control" rows="5" required>{{ request.form.get('description', '') }}</textarea>
          </div>

          <div class="mb-3">
            <label class="form-label">{{ t('attachments') }} <span class="text-muted small">({{ t('file_types_hint') }})</span></label>
            <input type="file" name="attachment" class="form-control"
                   accept=".pdf,.jpg,.jpeg,.png,.docx">
          </div>

          <div class="d-flex gap-2 justify-content-end">
            <a href="{{ url_for('main.dashboard') }}" class="btn btn-outline-secondary">{{ t('cancel') }}</a>
            <button type="submit" class="btn btn-primary">
              <i class="bi bi-send"></i> {{ t('send_ticket') }}
            </button>
          </div>
        </form>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

# ── ticket_detail.html ─────────────────────────
"templates/ticket_detail.html": """{% extends 'base.html' %}
{% block title %}{{ ticket.ticket_number }}{% endblock %}
{% block content %}
<div class="row g-4">
  <!-- Main ticket info -->
  <div class="col-lg-8">
    <div class="card border-0 shadow-sm mb-4">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span class="fw-bold">
          <span class="badge bg-secondary me-2">{{ ticket.ticket_number }}</span>
          {{ ticket.title }}
        </span>
        <span class="badge bg-{{ status_color(ticket.status) }} fs-6">{{ ticket.status }}</span>
      </div>
      <div class="card-body">
        <p class="mb-1 text-muted small">
          <i class="bi bi-person"></i> {{ ticket.creator.name }} &nbsp;|&nbsp;
          <i class="bi bi-calendar3"></i> {{ ticket.created_at | localtime }} &nbsp;|&nbsp;
          <i class="bi bi-flag"></i>
          <span class="badge bg-{{ priority_color(ticket.priority) }}{{ ' priority-high' if ticket.priority == 'High' else '' }}">{{ ticket.priority }}</span>
          {% if ticket.sla_breached %}
          <span class="badge bg-danger ms-1"><i class="bi bi-exclamation-triangle-fill"></i> {{ t('sla_breached_badge') }}</span>
          {% endif %}
        </p>
        <hr>
        <p class="mt-3" style="white-space:pre-wrap">{{ ticket.description }}</p>
      </div>
    </div>

    <!-- Comments -->
    <div class="card border-0 shadow-sm mb-4">
      <div class="card-header fw-semibold">
        <i class="bi bi-chat-dots"></i> {{ t('comments') }} ({{ ticket.comments.count() }})
      </div>
      <div class="card-body">
        {% for c in ticket.comments.order_by('created_at') %}
        <div class="d-flex gap-3 mb-3">
          <div class="rounded-circle bg-primary text-white d-flex align-items-center justify-content-center flex-shrink-0"
               style="width:38px;height:38px;font-size:.85rem">
            {{ c.author.name[:1] }}
          </div>
          <div class="flex-grow-1">
            <div class="fw-semibold small">{{ c.author.name }}
              <span class="text-muted fw-normal">· {{ c.created_at | localtime }}</span>
            </div>
            <p class="mb-0" style="white-space:pre-wrap">{{ c.body }}</p>
          </div>
        </div>
        {% else %}
        <p class="text-muted text-center">{{ t('no_comments') }}</p>
        {% endfor %}

        {% if ticket.status not in ('Closed',) %}
        <hr>
        <form method="POST" action="{{ url_for('employee.add_comment', ticket_id=ticket.id) }}">
          {{ form.hidden_tag() }}
          <div class="mb-2">
            <textarea name="body" class="form-control" rows="3" placeholder="{{ t('add_comment') }}" required></textarea>
          </div>
          <button type="submit" class="btn btn-primary btn-sm">
            <i class="bi bi-send"></i> {{ t('send') }}
          </button>
        </form>
        {% endif %}
      </div>
    </div>

    <!-- Attachments -->
    <div class="card border-0 shadow-sm mb-4">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span class="fw-semibold"><i class="bi bi-paperclip"></i> {{ t('attachments') }} ({{ ticket.attachments.count() }})</span>
      </div>
      <div class="card-body">
        {% for att in ticket.attachments.order_by('created_at') %}
        <div class="d-flex align-items-center justify-content-between py-2 border-bottom">
          <div class="d-flex align-items-center gap-2">
            {% if att.mime_type == 'application/pdf' %}
            <i class="bi bi-file-earmark-pdf text-danger fs-5"></i>
            {% elif att.mime_type.startswith('image/') %}
            <i class="bi bi-file-earmark-image text-primary fs-5"></i>
            {% elif 'word' in att.mime_type or 'docx' in att.mime_type %}
            <i class="bi bi-file-earmark-word text-info fs-5"></i>
            {% else %}
            <i class="bi bi-file-earmark fs-5"></i>
            {% endif %}
            <div>
              <div class="small fw-semibold">{{ att.original_name }}</div>
              <div class="text-muted" style="font-size:0.75rem">
                {{ (att.file_size / 1024)|round(1) }} KB &nbsp;·&nbsp;
                {{ att.uploader.name }} &nbsp;·&nbsp;
                {{ att.created_at | localtime }}
              </div>
            </div>
          </div>
          <a href="{{ url_for('employee.download_attachment', attachment_id=att.id) }}"
             class="btn btn-outline-secondary btn-sm">
            <i class="bi bi-download"></i> {{ t('download') }}
          </a>
        </div>
        {% else %}
        <p class="text-muted text-center mb-0">{{ t('no_attachments') }}</p>
        {% endfor %}

        {% if ticket.status not in ('Closed',) %}
        <form method="POST"
              action="{{ url_for('employee.upload_attachment', ticket_id=ticket.id) }}"
              enctype="multipart/form-data"
              class="mt-3">
          {{ form.hidden_tag() }}
          <div class="input-group input-group-sm">
            <input type="file" name="attachment" class="form-control form-control-sm"
                   accept=".pdf,.jpg,.jpeg,.png,.docx" required>
            <button type="submit" class="btn btn-outline-primary btn-sm">
              <i class="bi bi-upload"></i> {{ t('upload_file') }}
            </button>
          </div>
          <div class="form-text">{{ t('file_types_hint') }}</div>
        </form>
        {% endif %}
      </div>
    </div>
  </div>

  <!-- Sidebar -->
  <div class="col-lg-4">
    <div class="card border-0 shadow-sm mb-3">
      <div class="card-header fw-semibold">{{ t('ticket_details') }}</div>
      <ul class="list-group list-group-flush">
        <li class="list-group-item d-flex justify-content-between">
          <span class="text-muted">{{ t('type') }}</span><strong>{{ ticket.type }}</strong>
        </li>
        <li class="list-group-item d-flex justify-content-between">
          <span class="text-muted">{{ t('dept') }}</span>
          <strong>{{ ticket.department.name if ticket.department else t('na') }}</strong>
        </li>
        <li class="list-group-item d-flex justify-content-between">
          <span class="text-muted">{{ t('assignee') }}</span>
          <strong>{{ ticket.assignee.name if ticket.assignee else t('unassigned') }}</strong>
        </li>
        <li class="list-group-item d-flex justify-content-between">
          <span class="text-muted">{{ t('sla_deadline') }}</span>
          <strong class="{{ 'text-danger' if ticket.sla_breached else '' }}">
            {{ ticket.sla_deadline | localtime }}
          </strong>
        </li>
      </ul>
    </div>

    {% if current_user.role in ('admin', 'manager') %}
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">{{ t('actions') }}</div>
      <div class="card-body">
        <form method="POST" action="{{ url_for('admin.update_ticket', ticket_id=ticket.id) }}">
          {{ form.hidden_tag() }}
          <div class="mb-3">
            <label class="form-label small">{{ t('change_status') }}</label>
            <select name="status" class="form-select form-select-sm">
              {% for s in statuses %}
              <option value="{{ s }}" {% if ticket.status == s %}selected{% endif %}>{{ s }}</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label small">{{ t('assign_to') }}</label>
            <select name="assigned_to" class="form-select form-select-sm">
              <option value="">-- {{ t('unassigned') }} --</option>
              {% for u in agents %}
              <option value="{{ u.id }}"
                {% if ticket.assigned_to == u.id %}selected{% endif %}>{{ u.name }}</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label small">{{ t('priority') }}</label>
            <select name="priority" class="form-select form-select-sm">
              {% for p in priorities %}
              <option value="{{ p }}" {% if ticket.priority == p %}selected{% endif %}>{{ p }}</option>
              {% endfor %}
            </select>
          </div>
          <button type="submit" class="btn btn-primary btn-sm w-100">
            <i class="bi bi-check2"></i> {{ t('save_changes') }}
          </button>
        </form>
      </div>
    </div>
    {% endif %}

    {# ── Reopen button — shown to ticket creator when status is Resolved or Closed ── #}
    {% if current_user.id == ticket.created_by and ticket.status in ('Resolved', 'Closed') %}
    <div class="card border-0 shadow-sm mt-3">
      <div class="card-body">
        <form method="POST" action="{{ url_for('employee.reopen_ticket', ticket_id=ticket.id) }}">
          {{ form.hidden_tag() }}
          <button type="submit" class="btn btn-warning w-100">
            <i class="bi bi-arrow-counterclockwise"></i> {{ t('reopen_ticket') }}
          </button>
        </form>
      </div>
    </div>
    {% endif %}

    {# ── Delete ticket — Admin only ── #}
    {% if current_user.role == 'admin' %}
    <div class="card border-0 shadow-sm mt-3 border-danger">
      <div class="card-body">
        <form method="POST"
              action="{{ url_for('admin.delete_ticket', ticket_id=ticket.id) }}"
              onsubmit="return confirm('{{ t('confirm_delete') }}')">
          {{ form.hidden_tag() }}
          <button type="submit" class="btn btn-outline-danger w-100">
            <i class="bi bi-trash3"></i> {{ t('delete_ticket') }}
          </button>
        </form>
      </div>
    </div>
    {% endif %}
  </div>
</div>

{# ── Activity Log / Audit Trail ── #}
<div class="row mt-4">
  <div class="col-12">
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">
        <i class="bi bi-clock-history text-secondary me-1"></i> {{ t('audit_trail') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('hist_time') }}</th>
              {# Employees see a generic "By" column; admins/managers see the real actor name #}
              {% if current_user.role != 'employee' %}
              <th>{{ t('hist_actor') }}</th>
              {% endif %}
              <th>{{ t('hist_action') }}</th>
              <th>{{ t('hist_old') }}</th>
              <th>{{ t('hist_new') }}</th>
              <th>{{ t('hist_sla') }}</th>
            </tr>
          </thead>
          <tbody>
            {# visible_history is pre-filtered in the route based on current_user.role:
               - employees  : created / status_change / comment_added only (no reassign)
               - admin/mgr  : all actions #}
            {% for h in visible_history %}
            <tr>
              <td class="text-muted small text-nowrap">{{ h.created_at | localtime }}</td>
              {% if current_user.role != 'employee' %}
              {# Full name visible only to admin / manager #}
              <td class="small">{{ h.actor.name if h.actor else '—' }}</td>
              {% endif %}
              <td>
                {% if h.action == 'created' %}
                  <span class="badge bg-success">{{ t('action_created') }}</span>
                {% elif h.action == 'status_change' %}
                  <span class="badge bg-primary">{{ t('action_status_change') }}</span>
                {% elif h.action == 'reassign' %}
                  <span class="badge bg-info text-dark">{{ t('action_reassign') }}</span>
                {% elif h.action == 'comment_added' %}
                  <span class="badge bg-secondary">{{ t('action_comment_added') }}</span>
                {% elif h.action == 'attachment_uploaded' %}
                  <span class="badge bg-warning text-dark">{{ t('action_attachment_uploaded') }}</span>
                {% else %}
                  <span class="badge bg-light text-dark">{{ h.action }}</span>
                {% endif %}
              </td>
              <td class="small text-muted">{{ h.old_value or '—' }}</td>
              <td class="small">{{ h.new_value or '—' }}</td>
              <td>
                {% if h.sla_breached %}
                  <span class="badge bg-danger">{{ t('hist_breached') }}</span>
                {% else %}
                  <span class="badge bg-success">{{ t('hist_ok') }}</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
            <tr>
              <td colspan="{{ 5 if current_user.role == 'employee' else 6 }}"
                  class="text-center text-muted py-3">{{ t('no_history') }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

# ── dashboard_admin.html (overview) ───────────
"templates/dashboard_admin.html": """{% extends 'base.html' %}
{% block title %}{{ t('control_panel') }}{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
  <h4 class="fw-bold mb-0"><i class="bi bi-speedometer2 text-primary"></i> {{ t('control_panel') }}</h4>
  <span class="text-muted small">{{ now }}</span>
</div>

<!-- Stats Cards -->
<div class="row g-3 mb-4">
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-primary bg-opacity-10">
          <i class="bi bi-ticket-perforated text-primary fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.total }}</div>
          <div class="text-muted small">{{ t('total_tickets') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-warning bg-opacity-10">
          <i class="bi bi-folder2-open text-warning fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.open }}</div>
          <div class="text-muted small">{{ t('open') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-info bg-opacity-10">
          <i class="bi bi-arrow-repeat text-info fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.in_progress }}</div>
          <div class="text-muted small">{{ t('in_progress') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-danger bg-opacity-10">
          <i class="bi bi-exclamation-triangle text-danger fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.breached }}</div>
          <div class="text-muted small">{{ t('breached') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-success bg-opacity-10">
          <i class="bi bi-check-circle text-success fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.resolved }}</div>
          <div class="text-muted small">{{ t('resolved') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-secondary bg-opacity-10">
          <i class="bi bi-x-circle text-secondary fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.closed }}</div>
          <div class="text-muted small">{{ t('closed') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-danger bg-opacity-10">
          <i class="bi bi-fire text-danger fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.critical }}</div>
          <div class="text-muted small">{{ t('critical_open') }}</div>
        </div>
      </div>
    </div>
  </div>
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-warning bg-opacity-10">
          <i class="bi bi-person-check text-warning fs-4"></i>
        </div>
        <div>
          <div class="fs-3 fw-bold">{{ stats.unassigned }}</div>
          <div class="text-muted small">{{ t('unassigned_lbl') }}</div>
        </div>
      </div>
    </div>
  </div>
</div>

<div class="row g-4">
  <!-- Critical & High tickets -->
  <div class="col-lg-7">
    <div class="card border-0 shadow-sm">
      <div class="card-header d-flex justify-content-between align-items-center">
        <span class="fw-semibold"><i class="bi bi-fire text-danger me-1"></i> {{ t('urgent_tickets') }}</span>
        <a href="{{ url_for('admin.tickets') }}" class="btn btn-outline-primary btn-sm">{{ t('all_tickets_btn') }}</a>
      </div>
      <div class="table-responsive">
        <table class="table table-hover align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('ticket_number') }}</th>
              <th>{{ t('title') }}</th>
              <th>{{ t('priority') }}</th>
              <th>{{ t('status') }}</th>
              <th>{{ t('assignee') }}</th>
              <th></th>
            </tr>
          </thead>
          <tbody>
            {% for ut in urgent_tickets %}
            <tr {% if ut.sla_breached %}class="table-danger"{% endif %}>
              <td><span class="badge bg-secondary">{{ ut.ticket_number }}</span></td>
              <td>{{ ut.title|truncate(35) }}</td>
              <td><span class="badge bg-{{ priority_color(ut.priority) }}{{ ' priority-high' if ut.priority == 'High' else '' }}">{{ ut.priority }}</span></td>
              <td><span class="badge bg-{{ status_color(ut.status) }}">{{ ut.status }}</span></td>
              <td>{{ ut.assignee.name if ut.assignee else '—' }}</td>
              <td>
                <a href="{{ url_for('employee.ticket_detail', ticket_id=ut.id) }}"
                   class="btn btn-outline-primary btn-sm">{{ t('view') }}</a>
              </td>
            </tr>
            {% else %}
            <tr><td colspan="6" class="text-center text-muted py-3">{{ t('no_urgent') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- SLA Breached + Stats by dept -->
  <div class="col-lg-5">
    <!-- SLA Breached -->
    <div class="card border-0 shadow-sm mb-4">
      <div class="card-header fw-semibold">
        <i class="bi bi-exclamation-triangle-fill text-danger me-1"></i> {{ t('sla_breached_lbl') }}
      </div>
      <ul class="list-group list-group-flush">
        {% for bt in breached_tickets %}
        <li class="list-group-item d-flex justify-content-between align-items-center py-2">
          <div>
            <span class="badge bg-secondary me-1">{{ bt.ticket_number }}</span>
            <span class="small">{{ bt.title|truncate(30) }}</span>
          </div>
          <span class="badge bg-{{ priority_color(bt.priority) }}">{{ bt.priority }}</span>
        </li>
        {% else %}
        <li class="list-group-item text-center text-muted py-3">{{ t('no_sla_breach') }}</li>
        {% endfor %}
      </ul>
      {% if breached_tickets|length == 5 %}
      <div class="card-footer text-center">
        <a href="{{ url_for('admin.tickets') }}?sla=breached" class="small text-primary">{{ t('all_tickets_btn') }}</a>
      </div>
      {% endif %}
    </div>

    <!-- Tickets by Department -->
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">
        <i class="bi bi-diagram-3 text-primary me-1"></i> {{ t('tickets_by_dept') }}
      </div>
      <ul class="list-group list-group-flush">
        {% for dept_name, dept_count in dept_stats %}
        <li class="list-group-item d-flex justify-content-between align-items-center py-2">
          <span>{{ dept_name }}</span>
          <span class="badge bg-primary rounded-pill">{{ dept_count }}</span>
        </li>
        {% else %}
        <li class="list-group-item text-center text-muted py-3">{{ t('no_data') }}</li>
        {% endfor %}
      </ul>
    </div>
  </div>
</div>

{# ── Recent Activity (Audit Log) ── #}
<div class="row mt-4">
  <div class="col-12">
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">
        <i class="bi bi-clock-history text-secondary me-1"></i> {{ t('recent_activity') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('hist_time') }}</th>
              <th>{{ t('ticket_number') }}</th>
              <th>{{ t('hist_actor') }}</th>
              <th>{{ t('hist_action') }}</th>
              <th>{{ t('hist_old') }}</th>
              <th>{{ t('hist_new') }}</th>
              <th>{{ t('hist_sla') }}</th>
            </tr>
          </thead>
          <tbody>
            {% for h in recent_history %}
            <tr>
              <td class="text-muted small text-nowrap">{{ h.created_at | localtime }}</td>
              <td>
                <a href="{{ url_for('employee.ticket_detail', ticket_id=h.ticket_id) }}"
                   class="badge bg-secondary text-decoration-none">{{ h.ticket.ticket_number }}</a>
              </td>
              <td class="small">{{ h.actor.name if h.actor else '—' }}</td>
              <td>
                {% if h.action == 'created' %}
                  <span class="badge bg-success">{{ t('action_created') }}</span>
                {% elif h.action == 'status_change' %}
                  <span class="badge bg-primary">{{ t('action_status_change') }}</span>
                {% elif h.action == 'reassign' %}
                  <span class="badge bg-info text-dark">{{ t('action_reassign') }}</span>
                {% elif h.action == 'comment_added' %}
                  <span class="badge bg-secondary">{{ t('action_comment_added') }}</span>
                {% elif h.action == 'attachment_uploaded' %}
                  <span class="badge bg-warning text-dark">{{ t('action_attachment_uploaded') }}</span>
                {% else %}
                  <span class="badge bg-light text-dark">{{ h.action }}</span>
                {% endif %}
              </td>
              <td class="small text-muted">{{ h.old_value or '—' }}</td>
              <td class="small">{{ h.new_value or '—' }}</td>
              <td>
                {% if h.sla_breached %}
                  <span class="badge bg-danger">{{ t('hist_breached') }}</span>
                {% else %}
                  <span class="badge bg-success">{{ t('hist_ok') }}</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
            <tr><td colspan="7" class="text-center text-muted py-3">{{ t('no_history') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

# ── tickets_list.html ──────────────────────────
"templates/tickets_list.html": """{% extends 'base.html' %}
{% block title %}{{ t('all_tickets_title') }}{% endblock %}
{% block content %}
<h4 class="fw-bold mb-4"><i class="bi bi-list-task text-primary"></i> {{ t('all_tickets_title') }}</h4>

<div class="row g-3 mb-4">
  {% for label, count, color, icon in stats %}
  <div class="col-6 col-md-3">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-body d-flex align-items-center gap-3">
        <div class="rounded-circle p-3 bg-{{ color }} bg-opacity-10">
          <i class="bi bi-{{ icon }} text-{{ color }} fs-4"></i>
        </div>
        <div>
          <div class="fs-4 fw-bold">{{ count }}</div>
          <div class="text-muted small">{{ label }}</div>
        </div>
      </div>
    </div>
  </div>
  {% endfor %}
</div>

<!-- Filter bar -->
<form method="GET" class="row g-2 mb-3 align-items-end">
  <div class="col-auto">
    <select name="status" class="form-select form-select-sm">
      <option value="">{{ t('all_statuses') }}</option>
      {% for s in statuses %}
      <option value="{{ s }}" {% if request.args.get('status') == s %}selected{% endif %}>{{ s }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-auto">
    <select name="priority" class="form-select form-select-sm">
      <option value="">{{ t('all_priorities') }}</option>
      {% for p in priorities %}
      <option value="{{ p }}" {% if request.args.get('priority') == p %}selected{% endif %}>{{ p }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-auto">
    <select id="filter_dept" name="dept" class="form-select form-select-sm"
            hx-get="{{ url_for('api.filter_dept_agents') }}"
            hx-target="#filter_assignee_wrap"
            hx-trigger="change"
            hx-include="[name='dept']">
      <option value="">{{ t('all_depts') }}</option>
      {% for d in departments %}
      <option value="{{ d.id }}" {% if request.args.get('dept')|int == d.id %}selected{% endif %}>{{ d.name }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-auto" id="filter_assignee_wrap">
    <select name="assignee" class="form-select form-select-sm">
      <option value="">{{ t('all_assignees') }}</option>
      {% for u in all_agents %}
      <option value="{{ u.id }}" {% if request.args.get('assignee')|int == u.id %}selected{% endif %}>{{ u.name }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-auto">
    <button type="submit" class="btn btn-primary btn-sm">
      <i class="bi bi-funnel"></i> {{ t('filter') }}
    </button>
    <a href="{{ url_for('admin.tickets') }}" class="btn btn-outline-secondary btn-sm">{{ t('clear') }}</a>
  </div>
</form>

<div class="card border-0 shadow-sm">
  <div class="table-responsive">

  <!-- ── Bulk Actions Bar (3) ─────────────────────────────────────── -->
  <form id="bulkForm" method="POST" action="{{ url_for('admin.bulk_action') }}">
    {{ form.hidden_tag() }}
    <input type="hidden" name="ticket_ids" id="bulkIds">
    <div id="bulkBar"
         class="d-none align-items-center gap-2 px-3 py-2 bg-light border-bottom flex-wrap">
      <span class="fw-semibold text-muted small me-1" id="selCount">{{ t('sel_count_zero') }}</span>

      <select name="action" id="bulkAction" class="form-select form-select-sm w-auto">
        <option value="">{{ t('choose_action') }}</option>
        <option value="close">{{ t('closed') }}</option>
        <option value="change_status">{{ t('change_status') }}</option>
        <option value="assign">{{ t('assign_to') }}</option>
      </select>

      <select name="new_status" id="bulkStatus"
              class="form-select form-select-sm w-auto d-none">
        <option value="Open">{{ t('open') }}</option>
        <option value="In Progress">{{ t('in_progress') }}</option>
        <option value="Waiting for Customer">{{ t('waiting_customer') }}</option>
        <option value="Waiting for Vendor">{{ t('waiting_vendor') }}</option>
        <option value="Resolved">{{ t('resolved') }}</option>
        <option value="Reopened">{{ t('reopened') }}</option>
      </select>

      <select name="assigned_to" id="bulkAssign"
              class="form-select form-select-sm w-auto d-none">
        {% for u in all_agents %}
          <option value="{{ u.id }}">{{ u.name }}</option>
        {% endfor %}
      </select>

      <button type="submit" class="btn btn-sm btn-primary px-3"
              onclick="prepareBulk()">{{ t('apply') }}</button>
      <button type="button" class="btn btn-sm btn-outline-secondary"
              onclick="clearBulk()">{{ t('clear') }}</button>
    </div>
  </form>
  <!-- ─────────────────────────────────────────────────────────────── -->

    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th style="width:40px">
            <input type="checkbox" id="selectAll" class="form-check-input"
                   title="Select all">
          </th>
          <th>{{ t('ticket_number') }}</th>
          <th>{{ t('title') }}</th>
          <th>{{ t('requester') }}</th>
          <th>{{ t('dept') }}</th>
          <th>{{ t('priority') }}</th>
          <th>{{ t('status') }}</th>
          <th>{{ t('assignee') }}</th>
          <th>{{ t('date') }}</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
        {% for tk in tickets.items %}
        <tr {% if tk.sla_breached %}class="table-danger"{% endif %}>
          <td><input type="checkbox" class="form-check-input bulk-cb"
                     value="{{ tk.id }}"></td>
          <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
          <td>{{ tk.title|truncate(40) }}</td>
          <td>{{ tk.creator.name }}</td>
          <td>{{ tk.department.name if tk.department else '—' }}</td>
          <td><span class="badge bg-{{ priority_color(tk.priority) }}{{ ' priority-high' if tk.priority == 'High' else '' }}">{{ tk.priority }}</span></td>
          <td><span class="badge bg-{{ status_color(tk.status) }}">{{ tk.status }}</span></td>
          <td>{{ tk.assignee.name if tk.assignee else '—' }}</td>
          <td class="text-muted small">{{ tk.created_at | localtime('%Y-%m-%d') }}</td>
          <td class="text-nowrap">
            <a href="{{ url_for('employee.ticket_detail', ticket_id=tk.id) }}"
               class="btn btn-outline-primary btn-sm">{{ t('view') }}</a>
            {% if current_user.role == 'admin' %}
            <form method="POST"
                  action="{{ url_for('admin.delete_ticket', ticket_id=tk.id) }}"
                  class="d-inline"
                  onsubmit="return confirm('{{ t('confirm_delete') }}')">
              {{ form.hidden_tag() }}
              <button type="submit" class="btn btn-outline-danger btn-sm ms-1">
                <i class="bi bi-trash3"></i>
              </button>
            </form>
            {% endif %}
          </td>
        </tr>
        {% else %}
        <tr><td colspan="10" class="text-center text-muted py-4">{{ t('no_tickets_found') }}</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% if tickets.pages > 1 %}
  <div class="card-footer d-flex justify-content-center">
    <nav>
      <ul class="pagination mb-0">
        {% for p in tickets.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2) %}
          {% if p %}
          <li class="page-item {% if p == tickets.page %}active{% endif %}">
            <a class="page-link" href="{{ url_for('admin.tickets', page=p, **request.args) }}">{{ p }}</a>
          </li>
          {% else %}
          <li class="page-item disabled"><span class="page-link">…</span></li>
          {% endif %}
        {% endfor %}
      </ul>
    </nav>
  </div>
  {% endif %}
</div>

{% if current_user.role == 'admin' %}
<div class="mt-3 text-end">
  <a href="{{ url_for('admin.deleted_tickets') }}" class="btn btn-outline-danger btn-sm">
    <i class="bi bi-trash3"></i> View Deleted Tickets
  </a>
</div>
{% endif %}

<script>
/* ── Bulk Actions JS ─────────────────────────────────────────────── */
const bulkBar    = document.getElementById('bulkBar');
const selCount   = document.getElementById('selCount');
const bulkAction = document.getElementById('bulkAction');
const bulkStatus = document.getElementById('bulkStatus');
const bulkAssign = document.getElementById('bulkAssign');

function checkedBoxes() {
  return document.querySelectorAll('.bulk-cb:checked');
}

function updateBar() {
  const n = checkedBoxes().length;
  selCount.textContent = n + ' selected';
  if (n > 0) {
    bulkBar.classList.remove('d-none');
    bulkBar.classList.add('d-flex');
  } else {
    bulkBar.classList.add('d-none');
    bulkBar.classList.remove('d-flex');
  }
}

document.getElementById('selectAll').addEventListener('change', function () {
  document.querySelectorAll('.bulk-cb').forEach(cb => cb.checked = this.checked);
  updateBar();
});

document.querySelectorAll('.bulk-cb').forEach(cb => {
  cb.addEventListener('change', updateBar);
});

bulkAction.addEventListener('change', function () {
  bulkStatus.classList.toggle('d-none', this.value !== 'change_status');
  bulkAssign.classList.toggle('d-none', this.value !== 'assign');
});

function prepareBulk() {
  document.getElementById('bulkIds').value =
    [...checkedBoxes()].map(cb => cb.value).join(',');
}

function clearBulk() {
  document.querySelectorAll('.bulk-cb, #selectAll')
          .forEach(cb => cb.checked = false);
  updateBar();
}
/* ─────────────────────────────────────────────────────────────────── */
</script>
{% endblock %}
""",

# ── users.html (admin) ─────────────────────────
"templates/users.html": """{% extends 'base.html' %}
{% block title %}{{ t('manage_users') }}{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h4 class="fw-bold mb-0"><i class="bi bi-people-fill text-primary"></i> {{ t('manage_users') }}</h4>
  <a href="{{ url_for('admin.new_user') }}" class="btn btn-primary btn-sm">
    <i class="bi bi-person-plus"></i> {{ t('new_user') }}
  </a>
</div>
<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>{{ t('name') }}</th><th>{{ t('email') }}</th><th>{{ t('role') }}</th><th>{{ t('dept') }}</th><th>{{ t('status') }}</th><th></th>
        </tr>
      </thead>
      <tbody>
        {% for u in users %}
        <tr>
          <td>{{ u.name }}</td>
          <td>{{ u.email }}</td>
          <td><span class="badge bg-info text-dark">{{ u.role }}</span></td>
          <td>{{ u.department.name if u.department else '—' }}</td>
          <td>
            {% if u.active %}
            <span class="badge bg-success">{{ t('active') }}</span>
            {% else %}
            <span class="badge bg-secondary">{{ t('disabled') }}</span>
            {% endif %}
          </td>
          <td>
            <a href="{{ url_for('admin.edit_user', user_id=u.id) }}"
               class="btn btn-outline-secondary btn-sm">{{ t('edit') }}</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
{% endblock %}
""",

# ── notifications.html ─────────────────────────
"templates/notifications.html": """{% extends 'base.html' %}
{% block title %}{{ t('notifications') }}{% endblock %}
{% block content %}
<h4 class="fw-bold mb-3"><i class="bi bi-bell-fill text-primary"></i> {{ t('notifications') }}</h4>
<div class="list-group">
  {% for n in notifications %}
  <a href="{{ url_for('employee.ticket_detail', ticket_id=n.ticket_id) if n.ticket_id else '#' }}"
     class="list-group-item list-group-item-action {% if not n.is_read %}list-group-item-warning{% endif %}">
    <div class="d-flex justify-content-between">
      <span>{{ n.message }}</span>
      <small class="text-muted ms-3">{{ n.created_at | localtime }}</small>
    </div>
  </a>
  {% else %}
  <div class="list-group-item text-center text-muted py-4">{{ t('no_notifications') }}</div>
  {% endfor %}
</div>
{% endblock %}
""",

# ── 403.html / 404.html ────────────────────────
"templates/403.html": """{% extends 'base.html' %}
{% block title %}403{% endblock %}
{% block content %}
<div class="text-center py-5">
  <i class="bi bi-shield-x text-danger" style="font-size:4rem"></i>
  <h2 class="mt-3">{{ t('err_403_title') }}</h2>
  <p class="text-muted">{{ t('err_403_msg') }}</p>
  <a href="{{ url_for('main.dashboard') }}" class="btn btn-primary">{{ t('back_home') }}</a>
</div>
{% endblock %}
""",

"templates/forgot_password.html": """{% extends 'base.html' %}
{% block title %}{{ t('forgot_password') }}{% endblock %}
{% block content %}
<div class="row justify-content-center mt-5">
  <div class="col-md-5 col-lg-4">
    <div class="card border-0 shadow-sm">
      <div class="card-body p-4">
        <div class="text-center mb-4">
          <i class="bi bi-key-fill text-primary" style="font-size:2.5rem"></i>
          <h5 class="mt-2 fw-bold">{{ t('forgot_password') }}</h5>
          <p class="text-muted small">{{ t('reset_email_label') }}</p>
        </div>
        <form method="POST" action="{{ url_for('auth.forgot_password') }}">
          {{ form.hidden_tag() }}
          <div class="mb-3">
            <label class="form-label">{{ t('email') }}</label>
            <input type="email" name="email" class="form-control" required autofocus>
          </div>
          <div class="d-grid mt-4">
            <button type="submit" class="btn btn-primary">
              <i class="bi bi-envelope"></i> {{ t('send_reset_link') }}
            </button>
          </div>
        </form>
        <div class="text-center mt-3">
          <a href="{{ url_for('auth.login') }}" class="text-muted small">
            <i class="bi bi-arrow-left"></i> {{ t('back_to_login') }}
          </a>
        </div>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

"templates/reset_password.html": """{% extends 'base.html' %}
{% block title %}{{ t('reset_password') }}{% endblock %}
{% block content %}
<div class="row justify-content-center mt-5">
  <div class="col-md-5 col-lg-4">
    <div class="card border-0 shadow-sm">
      <div class="card-body p-4">
        <div class="text-center mb-4">
          <i class="bi bi-shield-lock-fill text-primary" style="font-size:2.5rem"></i>
          <h5 class="mt-2 fw-bold">{{ t('reset_password') }}</h5>
        </div>
        <form method="POST" action="{{ url_for('auth.reset_password', token=token) }}">
          {{ form.hidden_tag() }}
          <div class="mb-3">
            <label class="form-label">{{ t('new_password') }}</label>
            <input type="password" name="password" class="form-control" required autofocus>
            <div class="form-text text-muted mt-1">
              <strong>{{ t('pw_policy_title') }}:</strong><br>
              • {{ t('pw_min_chars') }}<br>
              • {{ t('pw_uppercase') }}<br>
              • {{ t('pw_digit') }}<br>
              • {{ t('pw_special') }}
            </div>
          </div>
          <div class="mb-3">
            <label class="form-label">{{ t('confirm_new_pw') }}</label>
            <input type="password" name="password2" class="form-control" required>
          </div>
          <div class="d-grid mt-4">
            <button type="submit" class="btn btn-primary">
              <i class="bi bi-check2-circle"></i> {{ t('update_password') }}
            </button>
          </div>
        </form>
        <div class="text-center mt-3">
          <a href="{{ url_for('auth.login') }}" class="text-muted small">
            <i class="bi bi-arrow-left"></i> {{ t('back_to_login') }}
          </a>
        </div>
      </div>
    </div>
  </div>
</div>
{% endblock %}
""",

"templates/429.html": """{% extends 'base.html' %}
{% block title %}{{ t('err_429_title') }}{% endblock %}
{% block content %}
<div class="text-center mt-5">
  <i class="bi bi-shield-exclamation text-warning" style="font-size:4rem"></i>
  <h2 class="mt-3">{{ t('err_429_title') }}</h2>
  <p class="text-muted">{{ t('err_429_msg') }}</p>
  <a href="{{ url_for('auth.login') }}" class="btn btn-primary mt-2">{{ t('back_to_login') }}</a>
</div>
{% endblock %}
""",

"templates/404.html": """{% extends 'base.html' %}
{% block title %}404{% endblock %}
{% block content %}
<div class="text-center py-5">
  <i class="bi bi-emoji-frown text-warning" style="font-size:4rem"></i>
  <h2 class="mt-3">{{ t('err_404_title') }}</h2>
  <a href="{{ url_for('main.dashboard') }}" class="btn btn-primary mt-2">{{ t('back_home') }}</a>
</div>
{% endblock %}
""",

# ── 500.html ───────────────────────────────────
"templates/500.html": """{% extends 'base.html' %}
{% block title %}500{% endblock %}
{% block content %}
<div class="text-center py-5">
  <i class="bi bi-exclamation-triangle-fill text-danger" style="font-size:4rem"></i>
  <h2 class="mt-3">{{ t('err_500_title') }}</h2>
  <p class="text-muted">{{ t('err_500_body') }}</p>
  <a href="{{ url_for('main.dashboard') }}" class="btn btn-primary mt-2">{{ t('back_home') }}</a>
</div>
{% endblock %}
""",

# ── setup.html ────────────────────────────────
"templates/setup.html": """<!DOCTYPE html>
<html lang="{{ 'ar' if lang == 'ar' else 'en' }}" dir="{{ 'rtl' if lang == 'ar' else 'ltr' }}">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ t('setup_title') }}</title>
  {% if lang == 'ar' %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.rtl.min.css">
  {% else %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
  {% endif %}
  <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
  <style>
    body { background: #f4f6f9; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
    .setup-card { max-width: 480px; margin: 80px auto; }
    .setup-icon { font-size: 3rem; color: #0d6efd; }
  </style>
</head>
<body>
<div class="setup-card">
  <div class="card border-0 shadow">
    <div class="card-body p-5">
      <div class="text-center mb-4">
        <i class="bi bi-shield-lock-fill setup-icon"></i>
        <h4 class="mt-3 fw-bold">{{ t('setup_title') }}</h4>
        <p class="text-muted small">{{ t('setup_subtitle') }}</p>
      </div>

      {% with messages = get_flashed_messages(with_categories=true) %}
        {% for cat, msg in messages %}
        <div class="alert alert-{{ cat }} alert-dismissible fade show" role="alert">
          {{ msg }}
          <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
        </div>
        {% endfor %}
      {% endwith %}

      <form method="POST" action="{{ url_for('setup.run_setup') }}">
        {{ form.hidden_tag() }}

        <div class="mb-3">
          <label class="form-label fw-semibold">{{ t('full_name') }} <span class="text-danger">*</span></label>
          <input type="text" name="name" class="form-control form-control-lg"
                 placeholder="e.g. John Smith" required value="{{ request.form.get('name', '') }}">
        </div>

        <div class="mb-3">
          <label class="form-label fw-semibold">{{ t('email') }} <span class="text-danger">*</span></label>
          <input type="email" name="email" class="form-control form-control-lg"
                 placeholder="admin@company.com" required value="{{ request.form.get('email', '') }}">
        </div>

        <div class="mb-3">
          <label class="form-label fw-semibold">{{ t('password') }} <span class="text-danger">*</span></label>
          <input type="password" name="password" class="form-control form-control-lg"
                 placeholder="Min. 10 characters" required>
          <div class="form-text text-muted">
            <strong>{{ t('pw_policy_title') }}:</strong>
            {{ t('pw_min_chars') }} · {{ t('pw_uppercase') }} · {{ t('pw_digit') }} · {{ t('pw_special') }}
          </div>
        </div>

        <div class="mb-4">
          <label class="form-label fw-semibold">{{ t('confirm_password') }} <span class="text-danger">*</span></label>
          <input type="password" name="password2" class="form-control form-control-lg"
                 placeholder="Re-enter password" required>
        </div>

        <div class="d-grid">
          <button type="submit" class="btn btn-primary btn-lg">
            <i class="bi bi-check2-circle me-1"></i> {{ t('create_account') }}
          </button>
        </div>
      </form>
      <div class="text-center mt-3">
        {% if lang == 'ar' %}
        <a href="{{ url_for('main.set_lang', lang='en') }}" class="btn btn-sm btn-outline-secondary">Switch to English</a>
        {% else %}
        <a href="{{ url_for('main.set_lang', lang='ar') }}" class="btn btn-sm btn-outline-secondary">العربية</a>
        {% endif %}
      </div>
    </div>
  </div>
  <p class="text-center text-muted small mt-3">
    <i class="bi bi-lock-fill"></i> {{ t('one_time_page') }}
  </p>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
""",

# ── style.css ──────────────────────────────────
"static/css/style.css": """
body {
  background: #f4f6f9;
  font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
}
.navbar { box-shadow: 0 2px 8px rgba(0,0,0,.12); }
.card   { border-radius: .75rem; }
.table th { font-weight: 600; font-size: .875rem; }
.badge  { font-size: .78rem; font-weight: 500; }
.list-group-item-warning { background: #fff8e1; }
/* Priority — High gets a border to distinguish it from Medium */
.badge.bg-warning { color: #000; }
.badge.priority-high { border: 1px solid #cc8800; font-weight: 700; }
/* Force LTR for inputs when page is LTR — Bootstrap RTL sometimes leaks direction */
:root { --input-dir: ltr; --input-align: left; }
[dir="rtl"] { --input-dir: rtl; --input-align: right; }

input, textarea, select {
  direction: var(--input-dir) !important;
  text-align: var(--input-align) !important;
}
""",
# ── reports.html ──────────────────────────────
"templates/reports.html": """{% extends 'base.html' %}
{% block title %}{{ t('reports_title') }}{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
  <h4 class="fw-bold mb-0"><i class="bi bi-bar-chart-fill text-primary"></i> {{ t('reports_title') }}</h4>
  <div class="d-flex align-items-center gap-2">
    <span class="text-muted small">{{ now }}</span>
    <a href="{{ url_for('admin.export_reports') }}"
       class="btn btn-sm btn-success d-flex align-items-center gap-1"
       title="Download full report as Excel workbook">
      <i class="bi bi-file-earmark-excel-fill"></i> Export Excel
    </a>
  </div>
</div>

<!-- Row 1: Tickets by Type + SLA Compliance -->
<div class="row g-4 mb-4">

  <!-- Tickets by Type -->
  <div class="col-lg-6">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header fw-semibold">
        <i class="bi bi-pie-chart text-primary me-1"></i> {{ t('rpt_total_by_type') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('rpt_ticket_type') }}</th>
              <th class="text-center">{{ t('rpt_count') }}</th>
              <th class="text-center">{{ t('rpt_open') }}</th>
              <th class="text-center">{{ t('rpt_resolved') }}</th>
              <th class="text-center">{{ t('rpt_closed') }}</th>
            </tr>
          </thead>
          <tbody>
            {% for row in by_type %}
            <tr>
              <td><span class="badge bg-secondary">{{ row.type }}</span></td>
              <td class="text-center fw-bold">{{ row.total }}</td>
              <td class="text-center text-warning">{{ row.open }}</td>
              <td class="text-center text-success">{{ row.resolved }}</td>
              <td class="text-center text-muted">{{ row.closed }}</td>
            </tr>
            {% else %}
            <tr><td colspan="5" class="text-center text-muted py-3">{{ t('rpt_no_data') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- SLA Compliance by Priority -->
  <div class="col-lg-6">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header fw-semibold">
        <i class="bi bi-shield-check text-success me-1"></i> {{ t('rpt_sla_compliance') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('rpt_priority') }}</th>
              <th class="text-center">{{ t('rpt_total_resolved') }}</th>
              <th class="text-center">{{ t('rpt_resolved_within') }}</th>
              <th class="text-center">{{ t('rpt_compliance_pct') }}</th>
            </tr>
          </thead>
          <tbody>
            {% for row in sla_compliance %}
            <tr>
              <td>
                <span class="badge bg-{{ priority_color(row.priority) }}">{{ row.priority }}</span>
              </td>
              <td class="text-center">{{ row.total_resolved }}</td>
              <td class="text-center text-success">{{ row.within_sla }}</td>
              <td class="text-center">
                {% if row.total_resolved > 0 %}
                  {% set pct = ((row.within_sla / row.total_resolved) * 100)|round(1) %}
                  <span class="fw-bold {{ 'text-success' if pct >= 80 else 'text-danger' }}">
                    {{ pct }}%
                  </span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
            <tr><td colspan="4" class="text-center text-muted py-3">{{ t('rpt_no_data') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<!-- Row 2: Avg Resolution Time by Dept + Tickets per Agent -->
<div class="row g-4 mb-4">

  <!-- Avg Resolution Time by Department -->
  <div class="col-lg-6">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header fw-semibold">
        <i class="bi bi-clock-history text-info me-1"></i> {{ t('rpt_avg_resolution') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('rpt_dept') }}</th>
              <th class="text-center">{{ t('rpt_resolved') }}</th>
              <th class="text-center">{{ t('rpt_avg_hrs') }}</th>
            </tr>
          </thead>
          <tbody>
            {% for row in avg_resolution %}
            <tr>
              <td>{{ row.dept }}</td>
              <td class="text-center">{{ row.resolved_count }}</td>
              <td class="text-center fw-bold">
                {% if row.avg_hours is not none %}
                  {{ row.avg_hours|round(1) }}h
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
            <tr><td colspan="3" class="text-center text-muted py-3">{{ t('rpt_no_data') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Tickets per Agent -->
  <div class="col-lg-6">
    <div class="card border-0 shadow-sm h-100">
      <div class="card-header fw-semibold">
        <i class="bi bi-person-badge text-warning me-1"></i> {{ t('rpt_per_agent') }}
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('rpt_agent') }}</th>
              <th class="text-center">{{ t('rpt_count') }}</th>
              <th class="text-center">{{ t('rpt_open') }}</th>
              <th class="text-center">{{ t('rpt_resolved') }}</th>
              <th class="text-center">{{ t('rpt_avg_agent_hrs') }}</th>
            </tr>
          </thead>
          <tbody>
            {% for row in per_agent %}
            <tr>
              <td>{{ row.name }}</td>
              <td class="text-center fw-bold">{{ row.total }}</td>
              <td class="text-center text-warning">{{ row.open }}</td>
              <td class="text-center text-success">{{ row.resolved }}</td>
              <td class="text-center">
                {% if row.avg_hours is not none %}
                  <span class="fw-semibold">{{ row.avg_hours }}h</span>
                {% else %}
                  <span class="text-muted">—</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
            <tr><td colspan="5" class="text-center text-muted py-3">{{ t('rpt_no_data') }}</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<!-- Row 3: Overdue Tickets -->
<div class="row g-4">
  <div class="col-12">
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">
        <i class="bi bi-exclamation-triangle-fill text-danger me-1"></i>
        {{ t('rpt_overdue') }}
        <span class="badge bg-danger ms-1">{{ overdue_tickets|length }}</span>
      </div>
      {% if overdue_tickets %}
      <div class="table-responsive">
        <table class="table table-sm align-middle mb-0">
          <thead class="table-light">
            <tr>
              <th>{{ t('ticket_number') }}</th>
              <th>{{ t('title') }}</th>
              <th>{{ t('priority') }}</th>
              <th>{{ t('dept') }}</th>
              <th>{{ t('assignee') }}</th>
              <th>{{ t('sla_deadline') }}</th>
              <th></th>
            </tr>
          </thead>
          <tbody>
            {% for tk in overdue_tickets %}
            <tr class="table-danger">
              <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
              <td>{{ tk.title|truncate(40) }}</td>
              <td><span class="badge bg-{{ priority_color(tk.priority) }}">{{ tk.priority }}</span></td>
              <td>{{ tk.department.name if tk.department else '—' }}</td>
              <td>{{ tk.assignee.name if tk.assignee else '—' }}</td>
              <td class="text-danger small">{{ tk.sla_deadline | localtime }}</td>
              <td>
                <a href="{{ url_for('employee.ticket_detail', ticket_id=tk.id) }}"
                   class="btn btn-outline-danger btn-sm">{{ t('view') }}</a>
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% else %}
      <div class="card-body text-center text-muted py-4">
        <i class="bi bi-check-circle-fill text-success fs-3 mb-2 d-block"></i>
        {{ t('rpt_no_data') }} — no overdue tickets
      </div>
      {% endif %}
    </div>
  </div>
</div>
{% endblock %}
""",

# ── deleted_tickets.html ────────────────────────
"templates/deleted_tickets.html": """{% extends 'base.html' %}
{% block title %}Deleted Tickets{% endblock %}
{% block content %}
<div class="container-fluid py-4">
  <div class="d-flex justify-content-between align-items-center mb-4">
    <h4 class="fw-bold text-danger">
      <i class="bi bi-trash3"></i> Deleted Tickets
    </h4>
    <a href="{{ url_for('admin.tickets') }}" class="btn btn-outline-secondary btn-sm">
      <i class="bi bi-arrow-left"></i> Back to All Tickets
    </a>
  </div>

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% for cat, msg in messages %}
      <div class="alert alert-{{ cat }} alert-dismissible fade show" role="alert">
        {{ msg }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button>
      </div>
    {% endfor %}
  {% endwith %}

  {% if tickets.items %}
  <div class="table-responsive">
    <table class="table table-hover align-middle">
      <thead class="table-light">
        <tr>
          <th>{{ t('ticket_number') }}</th>
          <th>{{ t('title') }}</th>
          <th>{{ t('type') }}</th>
          <th>{{ t('priority') }}</th>
          <th>{{ t('status') }}</th>
          <th>{{ t('deleted_at_col') }}</th>
          <th>{{ t('actions') }}</th>
        </tr>
      </thead>
      <tbody>
        {% for tk in tickets.items %}
        <tr>
          <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
          <td>{{ tk.title }}</td>
          <td>{{ tk.type }}</td>
          <td>
            <span class="badge bg-{{ priority_color(tk.priority) }}">{{ tk.priority }}</span>
          </td>
          <td>
            <span class="badge bg-{{ status_color(tk.status) }}">{{ tk.status }}</span>
          </td>
          <td>{{ tk.deleted_at | localtime if tk.deleted_at else '—' }}</td>
          <td>
            <form method="POST"
                  action="{{ url_for('admin.restore_ticket', ticket_id=tk.id) }}"
                  onsubmit="return confirm('Restore ticket {{ tk.ticket_number }}?')">
              {{ form.hidden_tag() }}
              <button type="submit" class="btn btn-sm btn-success">
                <i class="bi bi-arrow-counterclockwise"></i> Restore
              </button>
            </form>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  {% if tickets.pages > 1 %}
  <nav aria-label="Pagination">
    <ul class="pagination pagination-sm justify-content-center">
      {% for p in tickets.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2) %}
        {% if p %}
          <li class="page-item {% if p == tickets.page %}active{% endif %}">
            <a class="page-link" href="?page={{ p }}">{{ p }}</a>
          </li>
        {% else %}
          <li class="page-item disabled"><span class="page-link">…</span></li>
        {% endif %}
      {% endfor %}
    </ul>
  </nav>
  {% endif %}

  {% else %}
    <div class="alert alert-info">
      <i class="bi bi-info-circle"></i> No deleted tickets found.
    </div>
  {% endif %}
</div>
{% endblock %}
""",

# ── search_results.html ────────────────────────
"templates/search_results.html": """{% extends 'base.html' %}
{% block title %}{{ t('search_results') }}{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
  <h4 class="fw-bold mb-0">
    <i class="bi bi-search text-primary"></i> {{ t('search_results') }}
  </h4>
</div>

<!-- Search bar -->
<form method="GET" action="{{ url_for('admin.search') }}" class="mb-4">
  <div class="input-group input-group-lg shadow-sm">
    <span class="input-group-text bg-white border-end-0">
      <i class="bi bi-search text-muted"></i>
    </span>
    <input type="text" name="q" class="form-control border-start-0 ps-0"
           placeholder="{{ t('search_placeholder') }}"
           value="{{ query }}" autofocus>
    <button type="submit" class="btn btn-primary px-4">{{ t('search') }}</button>
    {% if query %}
    <a href="{{ url_for('admin.search') }}" class="btn btn-outline-secondary">{{ t('clear') }}</a>
    {% endif %}
  </div>
  <div class="form-text ms-1 mt-1">{{ t('search_tip') }}</div>
</form>

{% if query %}
<p class="text-muted mb-3">
  <strong>{{ results.total }}</strong> result(s) for
  <strong>"{{ query }}"</strong>
</p>
{% endif %}

<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>{{ t('ticket_number') }}</th>
          <th>{{ t('title') }}</th>
          <th>{{ t('requester') }}</th>
          <th>{{ t('dept') }}</th>
          <th>{{ t('priority') }}</th>
          <th>{{ t('status') }}</th>
          <th>{{ t('date') }}</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
        {% for tk in results.items %}
        <tr {% if tk.sla_breached %}class="table-danger"{% endif %}>
          <td><span class="badge bg-secondary">{{ tk.ticket_number }}</span></td>
          <td>{{ tk.title|truncate(45) }}</td>
          <td>{{ tk.creator.name }}</td>
          <td>{{ tk.department.name if tk.department else '—' }}</td>
          <td>
            <span class="badge bg-{{ priority_color(tk.priority) }}{{ ' priority-high' if tk.priority == 'High' else '' }}">
              {{ tk.priority }}
            </span>
          </td>
          <td><span class="badge bg-{{ status_color(tk.status) }}">{{ tk.status }}</span></td>
          <td class="text-muted small">{{ tk.created_at | localtime('%Y-%m-%d') }}</td>
          <td>
            <a href="{{ url_for('employee.ticket_detail', ticket_id=tk.id) }}"
               class="btn btn-outline-primary btn-sm">{{ t('view') }}</a>
          </td>
        </tr>
        {% else %}
        <tr>
          <td colspan="8" class="text-center text-muted py-5">
            <i class="bi bi-search fs-3 d-block mb-2"></i>
            {% if query %}{{ t('no_search_results') }}{% else %}{{ t('search_placeholder') }}{% endif %}
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <!-- Pagination -->
  {% if results.pages > 1 %}
  <div class="card-footer d-flex justify-content-center">
    <nav>
      <ul class="pagination mb-0">
        {% for p in results.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2) %}
          {% if p %}
          <li class="page-item {% if p == results.page %}active{% endif %}">
            <a class="page-link" href="{{ url_for('admin.search', q=query, page=p) }}">{{ p }}</a>
          </li>
          {% else %}
          <li class="page-item disabled"><span class="page-link">…</span></li>
          {% endif %}
        {% endfor %}
      </ul>
    </nav>
  </div>
  {% endif %}
</div>
{% endblock %}
""",

# ── departments.html (admin) ───────────────────
"templates/departments.html": """{% extends 'base.html' %}
{% block title %}{{ t('departments') }}{% endblock %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
  <h4 class="fw-bold mb-0">
    <i class="bi bi-diagram-3-fill text-primary"></i> {{ t('departments') }}
  </h4>
  <button class="btn btn-primary btn-sm" data-bs-toggle="modal" data-bs-target="#newDeptModal">
    <i class="bi bi-plus-lg"></i> {{ t('new_department') }}
  </button>
</div>

<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th>{{ t('dept_name_lbl') }}</th>
          <th>{{ t('dept_manager_lbl') }}</th>
          <th>{{ t('type') }}</th>
          <th class="text-center">{{ t('total_tickets') }}</th>
          <th class="text-center">{{ t('open') }}</th>
          <th class="text-center">{{ t('status') }}</th>
          <th></th>
        </tr>
      </thead>
      <tbody>
        {% for d, ticket_total, ticket_open in dept_rows %}
        {% set dept_types = d.allowed_types | from_json if d.allowed_types else [] %}
        <tr>
          <td class="fw-semibold">{{ d.name }}</td>
          <td>
            {% if d.manager_id %}
              {% set mgr = managers_map.get(d.manager_id) %}
              {{ mgr.name if mgr else '—' }}
            {% else %}
              <span class="text-muted">{{ t('no_manager') }}</span>
            {% endif %}
          </td>
          <td>
            {% if dept_types %}
              {% for tt in dept_types %}
                <span class="badge bg-light text-dark border me-1">{{ t(ticket_type_i18n.get(tt, tt)) }}</span>
              {% endfor %}
            {% else %}
              <span class="text-muted small">{{ t('na') }} ({{ t('rpt_total_by_type') }})</span>
            {% endif %}
          </td>
          <td class="text-center">{{ ticket_total }}</td>
          <td class="text-center">
            {% if ticket_open > 0 %}
              <span class="badge bg-warning text-dark">{{ ticket_open }}</span>
            {% else %}
              <span class="text-muted">0</span>
            {% endif %}
          </td>
          <td class="text-center">
            {% if d.is_deleted %}
              <span class="badge bg-secondary">{{ t('disabled') }}</span>
            {% else %}
              <span class="badge bg-success">{{ t('active') }}</span>
            {% endif %}
          </td>
          <td class="text-nowrap">
            <button class="btn btn-outline-secondary btn-sm"
                    data-bs-toggle="modal"
                    data-bs-target="#editDeptModal"
                    data-dept-id="{{ d.id }}"
                    data-dept-name="{{ d.name }}"
                    data-dept-manager="{{ d.manager_id or '' }}"
                    data-dept-types="{{ d.allowed_types or '' }}">
              <i class="bi bi-pencil"></i> {{ t('edit') }}
            </button>
            {% if not d.is_deleted %}
            <form method="POST"
                  action="{{ url_for('admin.delete_department', dept_id=d.id) }}"
                  class="d-inline"
                  onsubmit="return confirm('Delete department {{ d.name }}?')">
              {{ form.hidden_tag() }}
              <button type="submit" class="btn btn-outline-danger btn-sm ms-1">
                <i class="bi bi-trash3"></i>
              </button>
            </form>
            {% else %}
            <form method="POST"
                  action="{{ url_for('admin.restore_department', dept_id=d.id) }}"
                  class="d-inline">
              {{ form.hidden_tag() }}
              <button type="submit" class="btn btn-outline-success btn-sm ms-1">
                <i class="bi bi-arrow-counterclockwise"></i> Restore
              </button>
            </form>
            {% endif %}
          </td>
        </tr>
        {% else %}
        <tr>
          <td colspan="7" class="text-center text-muted py-4">{{ t('no_data') }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>

<!-- New Department Modal -->
<div class="modal fade" id="newDeptModal" tabindex="-1">
  <div class="modal-dialog">
    <div class="modal-content">
      <form method="POST" action="{{ url_for('admin.new_department') }}">
        {{ form.hidden_tag() }}
        <div class="modal-header">
          <h5 class="modal-title fw-bold">
            <i class="bi bi-plus-circle text-primary me-2"></i>{{ t('new_department') }}
          </h5>
          <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
        </div>
        <div class="modal-body">
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('dept_name_lbl') }} <span class="text-danger">*</span></label>
            <input type="text" name="name" class="form-control" required autofocus
                   placeholder="e.g. Finance">
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('dept_manager_lbl') }}</label>
            <select name="manager_id" class="form-select">
              <option value="">{{ t('no_manager') }}</option>
              {% for u in managers %}
              <option value="{{ u.id }}">{{ u.name }} ({{ u.role }})</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('type') }}
              <span class="text-muted fw-normal small ms-1">({{ t('na') }} = {{ t('rpt_total_by_type') }})</span>
            </label>
            <div class="d-flex flex-wrap gap-2 p-2 border rounded">
              {% for tt in ticket_types %}
              <div class="form-check">
                <input class="form-check-input" type="checkbox"
                       name="allowed_types" value="{{ tt }}"
                       id="new_type_{{ loop.index }}">
                <label class="form-check-label" for="new_type_{{ loop.index }}">
                  {{ t(ticket_type_i18n.get(tt, tt)) }}
                </label>
              </div>
              {% endfor %}
            </div>
            <div class="form-text">{{ t('dept_types_hint') }}</div>
          </div>
        </div>
        <div class="modal-footer">
          <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">{{ t('cancel') }}</button>
          <button type="submit" class="btn btn-primary">
            <i class="bi bi-check2-circle"></i> {{ t('new_department') }}
          </button>
        </div>
      </form>
    </div>
  </div>
</div>

<!-- Edit Department Modal -->
<div class="modal fade" id="editDeptModal" tabindex="-1">
  <div class="modal-dialog">
    <div class="modal-content">
      <form method="POST" action="{{ url_for('admin.edit_department') }}" id="editDeptForm">
        {{ form.hidden_tag() }}
        <input type="hidden" name="dept_id" id="editDeptId">
        <div class="modal-header">
          <h5 class="modal-title fw-bold">
            <i class="bi bi-pencil text-secondary me-2"></i>{{ t('edit_department') }}
          </h5>
          <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
        </div>
        <div class="modal-body">
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('dept_name_lbl') }} <span class="text-danger">*</span></label>
            <input type="text" name="name" id="editDeptName" class="form-control" required>
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('dept_manager_lbl') }}</label>
            <select name="manager_id" id="editDeptManager" class="form-select">
              <option value="">{{ t('no_manager') }}</option>
              {% for u in managers %}
              <option value="{{ u.id }}">{{ u.name }} ({{ u.role }})</option>
              {% endfor %}
            </select>
          </div>
          <div class="mb-3">
            <label class="form-label fw-semibold">{{ t('type') }}
              <span class="text-muted fw-normal small ms-1">({{ t('na') }} = {{ t('rpt_total_by_type') }})</span>
            </label>
            <div class="d-flex flex-wrap gap-2 p-2 border rounded" id="editTypeCheckboxes">
              {% for tt in ticket_types %}
              <div class="form-check">
                <input class="form-check-input edit-type-cb" type="checkbox"
                       name="allowed_types" value="{{ tt }}"
                       id="edit_type_{{ loop.index }}">
                <label class="form-check-label" for="edit_type_{{ loop.index }}">
                  {{ t(ticket_type_i18n.get(tt, tt)) }}
                </label>
              </div>
              {% endfor %}
            </div>
            <div class="form-text">{{ t('dept_types_hint') }}</div>
          </div>
        </div>
        <div class="modal-footer">
          <button type="button" class="btn btn-outline-secondary" data-bs-dismiss="modal">{{ t('cancel') }}</button>
          <button type="submit" class="btn btn-primary">
            <i class="bi bi-check2-circle"></i> {{ t('save_changes') }}
          </button>
        </div>
      </form>
    </div>
  </div>
</div>

<script>
document.getElementById('editDeptModal').addEventListener('show.bs.modal', function(e) {
  var btn = e.relatedTarget;
  document.getElementById('editDeptId').value   = btn.dataset.deptId;
  document.getElementById('editDeptName').value = btn.dataset.deptName;

  // Restore manager selection
  var mgr = btn.dataset.deptManager;
  var sel = document.getElementById('editDeptManager');
  for (var i = 0; i < sel.options.length; i++) {
    sel.options[i].selected = (sel.options[i].value == mgr);
  }

  // Restore allowed_types checkboxes
  var savedTypes = [];
  try { savedTypes = JSON.parse(btn.dataset.deptTypes || '[]'); } catch(e) {}
  document.querySelectorAll('.edit-type-cb').forEach(function(cb) {
    cb.checked = savedTypes.includes(cb.value);
  });
});
</script>
{% endblock %}
""",

# ── backups.html ──────────────────────────────
"templates/backups.html": """{% extends 'base.html' %}
{% block title %}{{ t('backups_title') }}{% endblock %}
{% block content %}

<div class="d-flex justify-content-between align-items-center mb-4">
  <h4 class="fw-bold mb-0">
    <i class="bi bi-shield-lock-fill text-success me-2"></i>{{ t('backups_title') }}
  </h4>
  <form method="POST" action="{{ url_for('admin.manual_backup') }}">
    {{ form.hidden_tag() }}
    <button type="submit" class="btn btn-success btn-sm">
      <i class="bi bi-cloud-arrow-up-fill me-1"></i>{{ t('backup_create') }}
    </button>
  </form>
</div>


{# ── Backups table ────────────────────────────────────────────── #}
<div class="card border-0 shadow-sm">
  <div class="table-responsive">
    <table class="table table-hover align-middle mb-0">
      <thead class="table-light">
        <tr>
          <th class="text-center" style="width:60px">{{ t('backup_id') }}</th>
          <th>{{ t('backup_date') }}</th>
          <th class="text-center">{{ t('backup_size') }}</th>
          <th class="text-center">{{ t('backup_source') }}</th>
          <th class="text-center">{{ t('backup_email') }}</th>
          <th class="text-center">{{ t('backup_drive') }}</th>
          <th class="text-end">{{ t('backup_actions') }}</th>
        </tr>
      </thead>
      <tbody>
        {% if backups %}
          {% for b in backups %}
          <tr>
            {# ID #}
            <td class="text-center text-muted small">#{{ b.id }}</td>

            {# Date #}
            <td>
              <span class="fw-semibold">{{ b.created_at | localtime('%Y-%m-%d') }}</span>
              <span class="text-muted small ms-1">{{ b.created_at | localtime('%H:%M:%S') }}</span>
            </td>

            {# Size #}
            <td class="text-center">
              {% if b.size_kb %}
                <span class="badge bg-secondary">{{ b.size_kb }} KB</span>
              {% else %}
                <span class="text-muted">—</span>
              {% endif %}
            </td>

            {# Source badge #}
            <td class="text-center">
              {% if b.source == 'manual' %}
                <span class="badge bg-primary">{{ t('backup_source_manual') }}</span>
              {% else %}
                <span class="badge bg-light text-dark border">{{ t('backup_source_auto') }}</span>
              {% endif %}
            </td>

            {# Email status #}
            <td class="text-center">
              {% if b.email_sent %}
                <span class="text-success small"><i class="bi bi-envelope-check-fill me-1"></i>{{ t('backup_email_sent') }}</span>
              {% else %}
                <span class="text-muted small"><i class="bi bi-envelope-slash me-1"></i>{{ t('backup_email_none') }}</span>
              {% endif %}
            </td>

            {# Google Drive status #}
            <td class="text-center">
              {% if b.gdrive_id %}
                <span class="text-success small"><i class="bi bi-google me-1"></i>{{ t('backup_drive_saved') }}</span>
              {% else %}
                <span class="text-muted small"><i class="bi bi-cloud-slash me-1"></i>{{ t('backup_drive_none') }}</span>
              {% endif %}
            </td>

            {# Actions #}
            <td class="text-end text-nowrap">
              {# Download #}
              <a href="{{ url_for('admin.download_backup', backup_id=b.id) }}"
                 class="btn btn-outline-secondary btn-sm me-1"
                 title="{{ t('backup_download') }}">
                <i class="bi bi-download"></i>
              </a>

              {# Restore — requires confirmation modal #}
              <button type="button"
                      class="btn btn-outline-danger btn-sm me-1"
                      title="{{ t('backup_restore') }}"
                      data-bs-toggle="modal"
                      data-bs-target="#restoreModal"
                      data-backup-date="{{ b.created_at | localtime }}"
                      data-restore-url="{{ url_for('admin.restore_backup', backup_id=b.id) }}">
                <i class="bi bi-arrow-counterclockwise"></i>
              </button>

              {# Delete — requires confirmation modal #}
              <button type="button"
                      class="btn btn-outline-secondary btn-sm"
                      title="{{ t('backup_delete') }}"
                      data-bs-toggle="modal"
                      data-bs-target="#deleteBackupModal"
                      data-backup-date="{{ b.created_at | localtime }}"
                      data-delete-url="{{ url_for('admin.delete_backup', backup_id=b.id) }}">
                <i class="bi bi-trash"></i>
              </button>
            </td>
          </tr>
          {% endfor %}
        {% else %}
          <tr>
            <td colspan="7" class="text-center text-muted py-5">
              <i class="bi bi-cloud-slash fs-2 d-block mb-2 opacity-25"></i>
              {{ t('backup_no_records') }}
            </td>
          </tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>

{# ── Restore Confirmation Modal ───────────────────────────────── #}
<div class="modal fade" id="restoreModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content border-0 shadow">
      <div class="modal-header bg-danger text-white">
        <h5 class="modal-title fw-bold">
          <i class="bi bi-exclamation-triangle-fill me-2"></i>
          {{ t('backup_restore') }}
        </h5>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <p class="mb-2">{{ t('backup_confirm_restore') }}</p>
        <p class="small text-muted mb-0">
          {% if lang == 'ar' %}
            النسخة: <strong id="restoreDate"></strong>
          {% else %}
            Snapshot: <strong id="restoreDate"></strong>
          {% endif %}
        </p>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-outline-secondary btn-sm" data-bs-dismiss="modal">
          {{ t('cancel') }}
        </button>
        <form method="POST" id="restoreForm">
          {{ form.hidden_tag() }}
          <button type="submit" class="btn btn-danger btn-sm">
            <i class="bi bi-arrow-counterclockwise me-1"></i>{{ t('backup_restore') }}
          </button>
        </form>
      </div>
    </div>
  </div>
</div>

{# ── Delete Backup Confirmation Modal ────────────────────────────── #}
<div class="modal fade" id="deleteBackupModal" tabindex="-1">
  <div class="modal-dialog modal-dialog-centered">
    <div class="modal-content border-0 shadow">
      <div class="modal-header bg-secondary text-white">
        <h5 class="modal-title fw-bold">
          <i class="bi bi-trash-fill me-2"></i>
          {{ t('backup_delete') }}
        </h5>
        <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
      </div>
      <div class="modal-body">
        <p class="mb-2">{{ t('backup_confirm_delete') }}</p>
        <p class="small text-muted mb-0">
          {% if lang == 'ar' %}
            النسخة: <strong id="deleteBackupDate"></strong>
          {% else %}
            Snapshot: <strong id="deleteBackupDate"></strong>
          {% endif %}
        </p>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-outline-secondary btn-sm" data-bs-dismiss="modal">
          {{ t('cancel') }}
        </button>
        <form method="POST" id="deleteBackupForm">
          {{ form.hidden_tag() }}
          <button type="submit" class="btn btn-secondary btn-sm">
            <i class="bi bi-trash me-1"></i>{{ t('backup_delete') }}
          </button>
        </form>
      </div>
    </div>
  </div>
</div>


<script>
document.getElementById('restoreModal').addEventListener('show.bs.modal', function (e) {
  var btn = e.relatedTarget;
  var dt  = btn.getAttribute('data-backup-date');
  document.getElementById('restoreDate').textContent = dt;
  document.getElementById('restoreForm').action = btn.getAttribute('data-restore-url');
});
document.getElementById('deleteBackupModal').addEventListener('show.bs.modal', function (e) {
  var btn = e.relatedTarget;
  var dt  = btn.getAttribute('data-backup-date');
  document.getElementById('deleteBackupDate').textContent = dt;
  document.getElementById('deleteBackupForm').action = btn.getAttribute('data-delete-url');
});
</script>

{% endblock %}
""",

}  # end TEMPLATES


def bootstrap_files():
    """
    Write template/static files to disk on every startup.
    Always overwrite — ensures code changes in TEMPLATES dict are reflected
    immediately without requiring manual deletion of the templates folder.
    Exception: style.css is only written if it does not exist (user may customise it).
    """
    for path, content in TEMPLATES.items():
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if path.endswith(".css") and os.path.exists(path):
            continue   # do not overwrite user-customised CSS
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)


# ─────────────────────────────────────────────
# TEMPLATE CONTEXT HELPERS
# ─────────────────────────────────────────────

@app.context_processor
def inject_helpers():
    def priority_color(p):
        return {"Low": "success", "Medium": "info", "High": "warning",
                "Critical": "danger"}.get(p, "secondary")

    def status_color(s):
        return {
            "Open": "primary", "In Progress": "info", "Resolved": "success",
            "Closed": "secondary", "Reopened": "warning",
            "Waiting for Customer": "warning", "Waiting for Vendor": "warning",
        }.get(s, "secondary")

    def csrf_token_input():
        from flask_wtf.csrf import generate_csrf
        return f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">'

    # Unread notification count — excludes notifications whose ticket has been
    # soft-deleted, so the badge never shows a count that leads to a 404 on click.
    unread_count = 0
    if current_user.is_authenticated:
        unread_count = (
            Notification.query
            .filter_by(user_id=current_user.id, is_read=False)
            .outerjoin(Ticket, Ticket.id == Notification.ticket_id)
            .filter(
                db.or_(
                    Notification.ticket_id == None,   # system notifications (no ticket)
                    Ticket.is_deleted == False,        # ticket still alive
                )
            )
            .count()
        )

    current_lang = get_lang()
    return dict(priority_color=priority_color, status_color=status_color,
                t=t, lang=current_lang, csrf_token_input=csrf_token_input,
                unread_count=unread_count)


# ─────────────────────────────────────────────
# ROLE DECORATORS
# ─────────────────────────────────────────────

from functools import wraps

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for("auth.login", next=request.url))
        if current_user.role != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated


def manager_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for("auth.login", next=request.url))
        if current_user.role not in ("admin", "manager"):
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────
# BLUEPRINTS / ROUTES
# ─────────────────────────────────────────────

from flask import Blueprint
from flask_wtf import FlaskForm   # bare form — only CSRF token

class EmptyForm(FlaskForm):
    pass


# ── SETUP (first-run wizard) ──────────────────
setup_bp = Blueprint("setup", __name__)

@setup_bp.route("/setup", methods=["GET", "POST"])
def run_setup():
    """Shown only once when the database has no users."""
    if User.query.count() > 0:
        return redirect(url_for("auth.login"))

    form = EmptyForm()
    if request.method == "POST" and form.validate_on_submit():
        # Second guard inside POST — prevents race condition if two requests arrive simultaneously
        if User.query.count() > 0:
            return redirect(url_for("auth.login"))

        name      = request.form.get("name", "").strip()
        email     = request.form.get("email", "").strip().lower()
        password  = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        # Validation
        if not name or not email or not password:
            flash(t("err_fields_required"), "danger")
        elif User.query.filter_by(email=email).first():
            flash(t("err_email_taken"), "danger")
        else:
            pw_errors = validate_password(password)
            if password != password2:
                pw_errors.append(t("err_pw_no_match"))
            if pw_errors:
                for err in pw_errors:
                    flash(err, "danger")
            else:
                # Create default departments
                dept_names = ["IT", "HR", "Finance", "General"]
                for dept_name in dept_names:
                    if not Department.query.filter_by(name=dept_name).first():
                        db.session.add(Department(name=dept_name))
                db.session.flush()

                # Create the admin account
                admin = User(
                    name   = name,
                    email  = email,
                    role   = "admin",
                    active = True,
                )
                admin.set_password(password)
                db.session.add(admin)
                db.session.commit()

                flash(t("flash_welcome", name=name), "success")
                return redirect(url_for("auth.login"))

    return render_template_string(TEMPLATES["templates/setup.html"], form=form)


# ── AUTH ──────────────────────────────────────
auth_bp = Blueprint("auth", __name__)

@auth_bp.route("/login", methods=["GET", "POST"])
@limiter.limit("10 per 15 minutes")
def login():
    # If database is empty — redirect to first-run setup page
    if User.query.count() == 0:
        return redirect(url_for("setup.run_setup"))
    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))
    form = EmptyForm()
    if request.method == "POST" and form.validate_on_submit():
        login_input = request.form.get("login_input", "").strip().lower()
        password    = request.form.get("password", "")
        # Match by email first, then by username
        user = (User.query.filter_by(email=login_input).first()
                or User.query.filter(
                    db.func.lower(User.username) == login_input
                ).first())
        if user and user.check_password(password) and user.active:
            login_user(user, remember=False)
            next_page = request.args.get("next")
            # Security: reject absolute URLs to prevent open redirect attacks
            # Only allow relative paths (start with /) within the same app
            if next_page and (not next_page.startswith("/") or next_page.startswith("//")):
                next_page = None
            return redirect(next_page or url_for("main.dashboard"))
        flash(t("flash_login_error"), "danger")
    return render_template_string(TEMPLATES["templates/login.html"], form=form)

@auth_bp.route("/logout")
@login_required
def logout():
    logout_user()
    flash(t("flash_logged_out"), "success")
    return redirect(url_for("auth.login"))


# ── MAIN ──────────────────────────────────────
# ── Password Reset Routes ────────────────────────────────────
@auth_bp.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])   # GET (view form) excluded — only POST submissions count
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = User.query.filter_by(email=email, active=True).first()
        # Always show success to prevent email enumeration
        if user:
            token = generate_reset_token(user.email)
            reset_url = url_for("auth.reset_password", token=token, _external=True)

            body_text = (
                f"Hello {user.name},\n\n"
                f"You requested a password reset for your Ticket System account.\n\n"
                f"Click the link below to reset your password (valid for 1 hour):\n"
                f"{reset_url}\n\n"
                f"If you did not request this, please ignore this email.\n"
            )
            body_html = (
                f"<p>Hello <strong>{user.name}</strong>,</p>"
                f"<p>You requested a password reset for your Ticket System account.</p>"
                f'<p><a href="{reset_url}" style="padding:8px 16px;background:#0d6efd;'
                f'color:#fff;border-radius:4px;text-decoration:none">Reset Password</a></p>'
                f"<p>This link is valid for <strong>1 hour</strong>. "
                f"If you did not request this, please ignore this email.</p>"
            )
            sent = send_email(
                to=user.email,
                subject="[Ticket System] Password Reset Request",
                body_text=body_text,
                body_html=body_html,
            )
            if sent:
                flash(t("flash_reset_sent"), "info")
            else:
                # SMTP not configured — log the reset URL server-side for admin use,
                # but never expose the token in the UI (security risk).
                app.logger.warning(
                    f"[PasswordReset] SMTP not configured — reset URL for {user.email}: {reset_url}"
                )
                flash(t("flash_reset_no_smtp"), "warning")
        else:
            flash(t("flash_reset_fallback"), "info")
        return redirect(url_for("auth.login"))
    return render_template_string(TEMPLATES["templates/forgot_password.html"], form=EmptyForm())



@auth_bp.route("/reset-password/<token>", methods=["GET", "POST"])
@limiter.limit("10 per hour")
def reset_password(token):
    email = verify_reset_token(token)
    if not email:
        flash(t("err_reset_invalid"), "danger")
        return redirect(url_for("auth.forgot_password"))

    user = User.query.filter_by(email=email).first()
    if not user:
        flash(t("err_user_not_found"), "danger")
        return redirect(url_for("auth.login"))

    if request.method == "POST":
        password  = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        errors = validate_password(password)
        if password != password2:
            errors.append(t("err_pw_no_match"))

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template_string(TEMPLATES["templates/reset_password.html"], form=EmptyForm(), token=token)

        user.set_password(password)
        db.session.commit()
        flash(t("flash_pw_updated"), "success")
        return redirect(url_for("auth.login"))

    return render_template_string(TEMPLATES["templates/reset_password.html"], form=EmptyForm(), token=token)


main_bp = Blueprint("main", __name__)


@main_bp.route("/toggle-availability", methods=["POST"])
@login_required
def toggle_availability():
    """
    Allows admin/manager users to toggle their own availability.
    on_leave users cannot toggle — Admin must clear the leave flag first.
    """
    if current_user.role not in ("admin", "manager"):
        abort(403)
    form = EmptyForm()
    if form.validate_on_submit():
        if current_user.on_leave:
            flash(t("on_leave_note"), "warning")
        else:
            current_user.is_available = not current_user.is_available
            db.session.commit()
    return redirect(request.referrer or url_for("main.dashboard"))


@main_bp.route("/set-lang/<lang>")
def set_lang(lang):
    if lang in ("en", "ar"):
        flask_session["lang"] = lang
    referrer = request.referrer
    if referrer and referrer.startswith(request.host_url):
        return redirect(referrer)
    return redirect(url_for("main.dashboard"))

@main_bp.route("/")
@login_required
def dashboard():
    if current_user.role in ("admin", "manager"):
        return redirect(url_for("admin.overview"))

    # Base query — tickets this employee created.
    # COUNT queries are far cheaper than loading all rows when a user has many tickets.
    base_q = Ticket.query.filter_by(created_by=current_user.id, is_deleted=False)

    stats = [
        (t("total_tickets"), base_q.count(),                                              "primary", "ticket-perforated"),
        (t("open"),          base_q.filter_by(status="Open").count(),                     "warning", "folder2-open"),
        (t("in_progress"),   base_q.filter_by(status="In Progress").count(),              "info",    "arrow-repeat"),
        (t("closed"),        base_q.filter_by(status="Closed").count(),                   "success", "check-circle"),
    ]

    page    = request.args.get("page", 1, type=int)
    tickets = base_q.order_by(Ticket.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )

    # Tickets assigned TO this employee (created by someone else).
    # Paginated separately — employee may be both creator and assignee on other tickets.
    apage           = request.args.get("apage", 1, type=int)
    assigned_tickets = (
        Ticket.query
        .filter(
            Ticket.assigned_to == current_user.id,
            Ticket.is_deleted  == False,
        )
        .order_by(Ticket.created_at.desc())
        .paginate(page=apage, per_page=20, error_out=False)
    )

    # Recent activity — history entries for tickets owned or assigned to this employee.
    visible_history = (
        TicketHistory.query
        .join(Ticket, Ticket.id == TicketHistory.ticket_id)
        .filter(
            db.or_(
                Ticket.created_by  == current_user.id,
                Ticket.assigned_to == current_user.id,
            ),
            Ticket.is_deleted   == False,
            TicketHistory.action.in_(["created", "status_change", "comment_added", "attachment_uploaded"]),
        )
        .options(
            joinedload(TicketHistory.actor),
            joinedload(TicketHistory.ticket),
        )
        .order_by(TicketHistory.created_at.desc())
        .limit(15)
        .all()
    )

    return render_template_string(TEMPLATES["templates/dashboard_employee.html"],
                                  tickets=tickets, stats=stats,
                                  assigned_tickets=assigned_tickets,
                                  visible_history=visible_history)

@main_bp.route("/notifications")
@login_required
def notifications():
    # Exclude notifications whose ticket has been soft-deleted —
    # showing them would display a link that returns 404 on click.
    # Notifications with ticket_id=None (system-level) are always shown.
    notifs = (
        Notification.query
        .filter_by(user_id=current_user.id)
        .outerjoin(Ticket, Ticket.id == Notification.ticket_id)
        .filter(
            db.or_(
                Notification.ticket_id == None,   # system notifications (no ticket)
                Ticket.is_deleted == False,        # ticket still alive
            )
        )
        .order_by(Notification.created_at.desc())
        .limit(50)
        .all()
    )
    # Mark all as read — including orphaned (deleted-ticket) notifications so they
    # don't keep inflating the badge after the ticket is gone.
    # Use a direct query instead of the dynamic relationship to avoid SQLAlchemy 2.x deprecation issues.
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({"is_read": True})
    db.session.commit()
    return render_template_string(TEMPLATES["templates/notifications.html"],
                                  notifications=notifs)


# ── EMPLOYEE ──────────────────────────────────
employee_bp = Blueprint("employee", __name__, url_prefix="/tickets")

TICKET_TYPES = ["IT Support", "HR Request", "Complaint", "General"]
PRIORITIES   = ["Low", "Medium", "High", "Critical"]

# ── Cascade: types allowed per department name keyword ──────────────────────
# Keys are lowercase substrings matched against department names.
# "default" is used when no keyword matches (covers custom department names).
# The raw English values MUST stay consistent with TICKET_TYPES above.
DEPT_TYPE_MAP = {
    "it":      ["IT Support", "General"],
    "tech":    ["IT Support", "General"],
    "hr":      ["HR Request", "General"],
    "human":   ["HR Request", "General"],
    "finance": ["Complaint", "General"],
    "account": ["Complaint", "General"],
    "legal":   ["Complaint", "General"],
    "default": TICKET_TYPES,   # all types for unrecognised departments
}

# Translation key map: raw value → i18n key
TICKET_TYPE_I18N = {
    "IT Support": "ttype_it_support",
    "HR Request": "ttype_hr_request",
    "Complaint":  "ttype_complaint",
    "General":    "ttype_general",
}

def get_types_for_dept(dept_name: str, dept_obj=None) -> list:
    """Return allowed ticket types for a department.

    Priority:
    1. dept_obj.allowed_types (DB column) — set by admin in the UI
    2. DEPT_TYPE_MAP keyword fallback (legacy / new depts not yet configured)
    """
    import json as _json
    if dept_obj is not None and dept_obj.allowed_types:
        try:
            types = _json.loads(dept_obj.allowed_types)
            if isinstance(types, list) and types:
                # Validate each entry against master whitelist to guard against stale data
                return [t for t in types if t in TICKET_TYPES] or TICKET_TYPES
        except (ValueError, TypeError):
            pass  # malformed JSON — fall through to keyword dict

    if not dept_name:
        return TICKET_TYPES
    lower = dept_name.lower()
    for keyword, types in DEPT_TYPE_MAP.items():
        if keyword != "default" and keyword in lower:
            return types
    return DEPT_TYPE_MAP["default"]

STATUSES     = ["Open", "In Progress", "Waiting for Customer",
                "Waiting for Vendor", "Resolved", "Closed", "Reopened"]

@employee_bp.route("/new", methods=["GET", "POST"])
@login_required
@limiter.limit("20 per hour")
def new_ticket():
    form = EmptyForm()
    departments = Department.query.filter_by(is_deleted=False).all()
    if request.method == "POST" and form.validate_on_submit():
        # assigned_to:
        #   admin/manager → manual pick from the form (already validated)
        #   employee      → auto-assign to the least-loaded available agent
        #                   in the selected department (set after ticket is created)
        validated_assigned = None
        if current_user.role in ("admin", "manager"):
            raw_assigned = request.form.get("assigned_to") or None
            if raw_assigned:
                try:
                    assignee_id = int(raw_assigned)
                    assignee = User.query.filter_by(id=assignee_id, active=True).first()
                    if assignee:
                        validated_assigned = assignee_id
                except (ValueError, TypeError):
                    pass  # Ignore invalid value

        # ── Validate department_id — must be an existing non-deleted department ──
        raw_dept = request.form.get("department_id") or None
        validated_dept = None
        if raw_dept:
            try:
                dept_id = int(raw_dept)
                dept_obj = Department.query.filter_by(id=dept_id, is_deleted=False).first()
                if dept_obj:
                    validated_dept = dept_id
            except (ValueError, TypeError):
                pass  # ignore invalid / injected value

        # ── Validate type and priority against whitelists ─────────────────
        raw_type     = request.form.get("type", "").strip()
        raw_priority = request.form.get("priority", "")

        # Department must be selected (mandatory)
        if not validated_dept:
            flash(t("err_dept_required"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )

        # Type must be selected and non-empty
        if not raw_type:
            flash(t("err_type_required"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )

        # Type must be in the global whitelist (guards against tampering)
        if raw_type not in TICKET_TYPES:
            flash(t("err_type_required"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )

        # Type must also be valid for the chosen department (cascade enforcement)
        dept_obj_for_validation = Department.query.filter_by(id=validated_dept, is_deleted=False).first()
        if dept_obj_for_validation:
            allowed_for_dept = get_types_for_dept(dept_obj_for_validation.name, dept_obj=dept_obj_for_validation)
            if raw_type not in allowed_for_dept:
                flash(t("err_type_required"), "danger")
                return render_template_string(
                    TEMPLATES["templates/new_ticket.html"],
                    form=form, departments=departments,
                    ticket_types=TICKET_TYPES,
                    ticket_type_i18n=TICKET_TYPE_I18N,
                    priorities=PRIORITIES,
                )

        # Priority is forced to "Low" for employees — only validate for admin/manager
        if current_user.role in ("admin", "manager") and raw_priority not in PRIORITIES:
            flash(t("err_invalid_priority"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )


        # ── Atomic ticket creation with IntegrityError retry ──────────────
        # generate_ticket_number() atomically increments a TicketCounter row
        # using UPDATE … RETURNING, so two concurrent requests always receive
        # distinct numbers.  The IntegrityError retry loop below is a last-resort
        # safety net (e.g. counter row missing) — under normal operation it never
        # triggers more than once.

        # ── Server-side blank validation (HTML required can be bypassed) ──
        _title_val = request.form.get("title", "").strip()
        _desc_val  = request.form.get("description", "").strip()
        if not _title_val:
            flash(t("err_title_empty"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )
        if not _desc_val:
            flash(t("err_desc_empty"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )

        from sqlalchemy.exc import IntegrityError as _IntegrityError
        ticket = None
        for _attempt in range(5):
            try:
                ticket = Ticket(
                    title         = _title_val,
                    description   = _desc_val,
                    type          = raw_type,
                    priority      = raw_priority if current_user.role in ("admin", "manager") else "Low",
                    status        = "Open",
                    created_by    = current_user.id,
                    department_id = validated_dept,
                    assigned_to   = validated_assigned,
                    ticket_number = generate_ticket_number(),
                )
                db.session.add(ticket)
                db.session.flush()   # raises IntegrityError if ticket_number is taken
                break                # success — exit retry loop
            except _IntegrityError:
                db.session.rollback()
                # expunge only the failed ticket — expunge_all() is unsafe in
                # threaded contexts as it evicts objects owned by other threads
                try:
                    if ticket is not None:
                        db.session.expunge(ticket)
                except Exception:
                    pass
                ticket = None
                continue             # regenerate and try again
        else:
            # Extremely unlikely (5 concurrent collisions in a row)
            app.logger.error("Failed to generate unique ticket_number after 5 attempts.")
            flash(t("err_ticket_create"), "danger")
            return render_template_string(
                TEMPLATES["templates/new_ticket.html"],
                form=form, departments=departments,
                ticket_types=TICKET_TYPES,
                ticket_type_i18n=TICKET_TYPE_I18N,
                priorities=PRIORITIES,
            )
        # ─────────────────────────────────────────────────────────────────

        write_history(ticket, "created", None, "Open", current_user.id)

        # ── Auto-assign for employees ─────────────────────────────────────
        # Admin/Manager: validated_assigned already set from the form above.
        # Employee: run auto_assign_ticket to find the least-loaded available
        #           agent in the selected department.
        if current_user.role == "employee":
            chosen = auto_assign_ticket(ticket)
            if chosen:
                validated_assigned = chosen.id
                ticket.assigned_to = chosen.id

        # Write history + notify whoever was assigned (manual or auto)
        if validated_assigned:
            assignee_obj = db.session.get(User, validated_assigned)
            write_history(ticket, "reassign", None, assignee_obj.name, current_user.id)
            send_notification(validated_assigned, ticket.id,
                              f"You have been assigned to new ticket [{ticket.ticket_number}]",
                              event="assigned")
        db.session.commit()

        # Optional file attachment on ticket creation
        file = request.files.get("attachment")
        if file and file.filename:
            attachment, att_error = _save_attachment(file, ticket)
            if att_error:
                flash(att_error, "warning")   # ticket is already saved — warn but don't block
            else:
                write_history(ticket, "attachment_uploaded", None,
                              f"attachment uploaded: {attachment.original_name}", current_user.id)
                db.session.commit()

        flash(t("flash_ticket_ok", num=ticket.ticket_number), "success")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket.id))
    return render_template_string(TEMPLATES["templates/new_ticket.html"],
                                  form=form, departments=departments,
                                  ticket_types=TICKET_TYPES,
                                  ticket_type_i18n=TICKET_TYPE_I18N,
                                  priorities=PRIORITIES)

@employee_bp.route("/<int:ticket_id>")
@login_required
def ticket_detail(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()
    # Access: owner, assignee, admin, manager
    if current_user.role == "employee" and \
       ticket.created_by != current_user.id and \
       ticket.assigned_to != current_user.id:
        abort(403)
    form    = EmptyForm()
    agents  = User.query.filter_by(active=True).filter(User.role.in_(["admin", "manager"])).all()

    # ── History visibility: employees only see their own actions + status changes.
    #    Reassign events (internal management) are hidden from employees entirely.
    #    joinedload(actor) prevents N+1: the template accesses h.actor.name on every row.
    if current_user.role == "employee":
        visible_history = (
            ticket.history
            .filter(TicketHistory.action.in_(["created", "status_change", "comment_added", "attachment_uploaded"]))
            .options(joinedload(TicketHistory.actor))
            .order_by(TicketHistory.created_at)
            .limit(200)   # guard: a long-lived ticket can accumulate hundreds of rows — cap to avoid memory/DOM bloat
            .all()
        )
    else:
        visible_history = (
            ticket.history
            .options(joinedload(TicketHistory.actor))
            .order_by(TicketHistory.created_at)
            .limit(200)   # same guard for admin/manager view
            .all()
        )

    return render_template_string(TEMPLATES["templates/ticket_detail.html"],
                                  ticket=ticket, form=form, agents=agents,
                                  statuses=STATUSES, priorities=PRIORITIES, visible_history=visible_history)

@employee_bp.route("/<int:ticket_id>/comment", methods=["POST"])
@login_required
@limiter.limit("60 per hour")
def add_comment(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()
    if current_user.role == "employee" and \
       ticket.created_by != current_user.id and \
       ticket.assigned_to != current_user.id:
        abort(403)
    form = EmptyForm()
    if form.validate_on_submit():
        body = request.form.get("body", "").strip()
        if body:
            comment = Comment(ticket_id=ticket.id, user_id=current_user.id, body=body)
            db.session.add(comment)
            write_history(ticket, "comment_added", None, (body[:120] + ("…" if len(body) > 120 else "")), current_user.id)
            # Notify assignee if commenter is the creator (and vice versa)
            notify_target = ticket.assigned_to if current_user.id == ticket.created_by else ticket.created_by
            if notify_target:
                send_notification(notify_target, ticket.id,
                                  f"New comment on [{ticket.ticket_number}] by {current_user.name}")
            # ── @Mention detection ───────────────────────────────────────
            process_mentions(body, ticket, current_user.id)
            # ─────────────────────────────────────────────────────────────
            db.session.commit()
            flash(t("flash_comment_ok"), "success")
    return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))


# ─────────────────────────────────────────────
# ATTACHMENT HELPERS & ROUTES
# ─────────────────────────────────────────────

ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png", "docx"}
ALLOWED_MIMETYPES  = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
MAX_FILE_BYTES = 10 * 1024 * 1024   # 10 MB


def _get_upload_folder():
    folder = app.config.get("UPLOAD_FOLDER", os.path.join(os.path.expanduser("~"), "tickets_uploads"))
    os.makedirs(folder, exist_ok=True)
    return folder


def _save_attachment(file_storage, ticket):
    """
    Validate, save to disk with UUID filename, write Attachment row.
    Returns (Attachment | None, error_message | None).
    Does NOT call db.session.commit() — caller handles the transaction.
    """
    original_name = file_storage.filename or ""
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""

    # Extension whitelist
    if ext not in ALLOWED_EXTENSIONS:
        return None, t("file_type_not_allowed")

    # Read file bytes once for size check and MIME sniffing
    file_bytes = file_storage.read()
    if len(file_bytes) > MAX_FILE_BYTES:
        return None, t("file_too_large")

    # ── Server-side MIME validation (python-magic reads actual bytes) ────
    ext_mime_map = {
        "pdf":  "application/pdf",
        "jpg":  "image/jpeg",
        "jpeg": "image/jpeg",
        "png":  "image/png",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
    if _MAGIC_AVAILABLE:
        detected_mime = _magic.from_buffer(file_bytes, mime=True)
    else:
        guessed, _ = mimetypes.guess_type(original_name)
        detected_mime = guessed or ext_mime_map.get(ext, "")

    if detected_mime not in ALLOWED_MIMETYPES:
        return None, t("file_type_not_allowed")
    # ─────────────────────────────────────────────────────────────────────

    # Store file bytes in the database (Railway has no persistent disk)
    attachment = Attachment(
        ticket_id     = ticket.id,
        uploaded_by   = current_user.id,
        filename      = None,
        original_name = secure_filename(original_name),
        file_size     = len(file_bytes),
        mime_type     = detected_mime,
        file_data     = file_bytes,
    )
    db.session.add(attachment)
    return attachment, None


@employee_bp.route("/<int:ticket_id>/upload", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
def upload_attachment(ticket_id):
    """Upload a file attachment to an existing ticket."""
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()

    # Access: owner, assignee, admin, manager
    if current_user.role == "employee" and \
       ticket.created_by != current_user.id and \
       ticket.assigned_to != current_user.id:
        abort(403)
    if ticket.status == "Closed":
        flash(t("err_attach_closed"), "warning")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    form = EmptyForm()
    if not form.validate_on_submit():
        flash(t("upload_error"), "danger")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    file = request.files.get("attachment")
    if not file or not file.filename:
        flash(t("upload_error"), "danger")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    attachment, error = _save_attachment(file, ticket)
    if error:
        flash(error, "danger")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    write_history(ticket, "attachment_uploaded", None,
                  f"attachment uploaded: {attachment.original_name}", current_user.id)
    db.session.commit()
    flash(t("upload_ok"), "success")
    return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))


@employee_bp.route("/attachments/<int:attachment_id>")
@login_required
def download_attachment(attachment_id):
    """
    Secure file download — authenticated users only.
    Employee can only download files belonging to their own tickets.
    Admin/Manager can download any attachment.
    """
    att = db.session.get(Attachment, attachment_id)
    if not att:
        abort(404)

    ticket = db.session.get(Ticket, att.ticket_id)
    if not ticket or ticket.is_deleted:
        abort(404)

    # Access control — employee can download if they created OR are assigned to the ticket
    if (current_user.role == "employee"
            and ticket.created_by != current_user.id
            and ticket.assigned_to != current_user.id):
        abort(403)

    import io

    # Primary: file stored in database (Railway deployment)
    if att.file_data is not None:
        return send_file(
            io.BytesIO(att.file_data),
            mimetype=att.mime_type,
            as_attachment=True,
            download_name=att.original_name,
        )

    # Fallback: file stored on disk (legacy / non-Railway deployments)
    if att.filename:
        disk_path = os.path.join(_get_upload_folder(), att.filename)
        if os.path.exists(disk_path):
            return send_file(
                disk_path,
                mimetype=att.mime_type,
                as_attachment=True,
                download_name=att.original_name,
            )

    flash(t("err_file_not_found"), "danger")
    return redirect(url_for("employee.ticket_detail", ticket_id=ticket.id))


@employee_bp.route("/<int:ticket_id>/reopen", methods=["POST"])
@login_required
def reopen_ticket(ticket_id):
    """Allow the ticket creator to reopen a Resolved or Closed ticket."""
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()

    # Only the creator (employee) can reopen — admins/managers use update_ticket.
    # Manager is further restricted to their own department tickets.
    if current_user.role == "manager" and ticket.department_id != current_user.department_id:
        abort(403)
    if current_user.role == "employee" and ticket.created_by != current_user.id:
        abort(403)

    form = EmptyForm()
    if not form.validate_on_submit():
        flash(t("err_invalid_request"), "danger")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    if ticket.status not in ("Resolved", "Closed"):
        flash(t("reopen_not_allowed"), "warning")
        return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

    write_history(ticket, "status_change", ticket.status, "Reopened", current_user.id)
    ticket.status = "Reopened"
    ticket.updated_at = utc_now()

    # Notify assignee if exists; otherwise notify all admins so nobody misses the reopen
    if ticket.assigned_to:
        send_notification(
            ticket.assigned_to, ticket.id,
            f"Ticket [{ticket.ticket_number}] has been reopened by {current_user.name}",
            event="reopened",
        )
    else:
        for admin_user in User.query.filter_by(role="admin", active=True).all():
            if admin_user.id != current_user.id:
                send_notification(
                    admin_user.id, ticket.id,
                    f"Unassigned ticket [{ticket.ticket_number}] was reopened by {current_user.name}",
                    event="reopened",
                )

    db.session.commit()
    flash(t("flash_reopen_ok"), "success")
    return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))


# ── ADMIN ─────────────────────────────────────
admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

@admin_bp.route("/overview")
@manager_required
def overview():
    """Main admin control panel — comprehensive overview."""
    from sqlalchemy import func

    base_q = Ticket.query.filter_by(is_deleted=False)
    if current_user.role == "manager":
        base_q = base_q.filter_by(department_id=current_user.department_id)

    open_statuses = ("Open", "In Progress", "Waiting for Customer", "Waiting for Vendor", "Reopened")

    stats = {
        "total":       base_q.count(),
        "open":        base_q.filter(Ticket.status == "Open").count(),
        "in_progress": base_q.filter(Ticket.status == "In Progress").count(),
        "breached":    base_q.filter(Ticket.sla_breached == True).count(),
        "resolved":    base_q.filter(Ticket.status == "Resolved").count(),
        "closed":      base_q.filter(Ticket.status == "Closed").count(),
        "critical":    base_q.filter(Ticket.priority == "Critical",
                                     Ticket.status.in_(open_statuses)).count(),
        "unassigned":  base_q.filter(Ticket.assigned_to == None,
                                     Ticket.status.in_(open_statuses)).count(),
    }

    # Last 8 Critical or High open tickets
    urgent_tickets = (base_q.filter(
        Ticket.priority.in_(["Critical", "High"]),
        Ticket.status.in_(open_statuses)
    ).options(
        joinedload(Ticket.assignee),    # template accesses ut.assignee.name — without this: 1 SELECT per row
    ).order_by(Ticket.created_at.asc()).limit(8).all())

    # Last 5 SLA-breached tickets
    breached_tickets = (base_q.filter(
        Ticket.sla_breached == True,
        Ticket.status.in_(open_statuses)
    ).order_by(Ticket.created_at.asc()).limit(5).all())

    # Stats by department
    dept_rows = (db.session.query(Department.name, func.count(Ticket.id))
                 .join(Ticket, Ticket.department_id == Department.id)
                 .filter(Ticket.is_deleted == False)
                 .group_by(Department.name)
                 .order_by(func.count(Ticket.id).desc())
                 .all())
    dept_stats = [(row[0], row[1]) for row in dept_rows]

    now = local_now().strftime("%Y-%m-%d %H:%M")

    # Last 20 history entries across all visible tickets
    # joinedload(actor) + joinedload(ticket) eliminates the N+1 problem:
    # without them, accessing h.actor.name and h.ticket.ticket_number in the template
    # triggers one extra SELECT per row (up to 40 extra queries for 20 rows).
    history_q = (TicketHistory.query
                 .join(Ticket, Ticket.id == TicketHistory.ticket_id)
                 .filter(Ticket.is_deleted == False)
                 .options(
                     joinedload(TicketHistory.actor),
                     joinedload(TicketHistory.ticket),
                 ))
    if current_user.role == "manager":
        history_q = history_q.filter(Ticket.department_id == current_user.department_id)
    recent_history = history_q.order_by(TicketHistory.created_at.desc()).limit(20).all()

    return render_template_string(TEMPLATES["templates/dashboard_admin.html"],
                                  stats=stats,
                                  urgent_tickets=urgent_tickets,
                                  breached_tickets=breached_tickets,
                                  dept_stats=dept_stats,
                                  recent_history=recent_history,
                                  now=now)


@admin_bp.route("/tickets")
@manager_required
def tickets():
    query = Ticket.query.filter_by(is_deleted=False)
    if current_user.role == "manager":
        query = query.filter_by(department_id=current_user.department_id)

    status   = request.args.get("status")
    priority = request.args.get("priority")
    dept     = request.args.get("dept")
    assignee = request.args.get("assignee")
    sla      = request.args.get("sla")
    if status:
        query = query.filter_by(status=status)
    if priority:
        query = query.filter_by(priority=priority)
    if dept:
        try:
            query = query.filter_by(department_id=int(dept))
        except (ValueError, TypeError):
            pass
    if assignee:
        try:
            query = query.filter_by(assigned_to=int(assignee))
        except (ValueError, TypeError):
            pass
    if sla == "breached":
        query = query.filter_by(sla_breached=True)

    tickets_page = query.options(
        joinedload(Ticket.creator),
        joinedload(Ticket.assignee),
        joinedload(Ticket.department),
    ).order_by(Ticket.created_at.desc()).paginate(
        page=request.args.get("page", 1, type=int), per_page=20, error_out=False
    )
    total    = Ticket.query.filter_by(is_deleted=False).count()
    open_cnt = Ticket.query.filter_by(is_deleted=False, status="Open").count()
    breach   = Ticket.query.filter_by(is_deleted=False, sla_breached=True).count()
    stats = [
        (t("total_tickets"), total,    "primary", "ticket-perforated"),
        (t("open"),          open_cnt, "warning",  "folder2-open"),
        ("SLA Breached",     breach,   "danger",   "exclamation-triangle"),
    ]
    departments = Department.query.filter_by(is_deleted=False).all()
    all_agents  = User.query.filter_by(active=True).filter(
        User.role.in_(["admin", "manager"])
    ).order_by(User.name).all()
    form = EmptyForm()
    return render_template_string(TEMPLATES["templates/tickets_list.html"],
                                  tickets=tickets_page, stats=stats,
                                  departments=departments,
                                  statuses=STATUSES, priorities=PRIORITIES,
                                  all_agents=all_agents, form=form)

@admin_bp.route("/tickets/<int:ticket_id>/update", methods=["POST"])
@manager_required
def update_ticket(ticket_id):
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()
    # Manager can only update tickets within their own department
    if current_user.role == "manager" and ticket.department_id != current_user.department_id:
        abort(403)
    form = EmptyForm()
    if form.validate_on_submit():
        new_status   = request.form.get("status")
        new_assigned = request.form.get("assigned_to") or None
        new_priority = request.form.get("priority")

        # Fix TC-044: validate new_status against whitelist — reject unknown/injected values
        if new_status and new_status not in STATUSES:
            flash(t("err_invalid_status"), "danger")
            return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

        if new_status and new_status != ticket.status:
            write_history(ticket, "status_change", ticket.status, new_status, current_user.id)
            # Notify creator on status change
            _ev = "resolved" if new_status in ("Resolved", "Closed") else ""
            send_notification(ticket.created_by, ticket.id,
                              f"Your ticket [{ticket.ticket_number}] status changed to: {new_status}",
                              event=_ev)
            ticket.status = new_status

        # Fix: validate and apply priority change
        if new_priority and new_priority not in PRIORITIES:
            flash(t("err_invalid_priority"), "danger")
            return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))

        if new_priority and new_priority != ticket.priority:
            write_history(ticket, "priority_change", ticket.priority, new_priority, current_user.id)
            ticket.priority = new_priority

        if new_assigned and str(new_assigned) != str(ticket.assigned_to or ""):
            # Fix TC-051: validate assignee exists and is active before saving
            try:
                assignee_id = int(new_assigned)
            except (ValueError, TypeError):
                flash(t("err_invalid_assignee"), "danger")
                return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))
            new_user = db.session.get(User, assignee_id)
            if not new_user or not new_user.active or new_user.role not in ("admin", "manager", "employee"):
                flash(t("err_assignee_inactive"), "danger")
                return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))
            old_name = ticket.assignee.name if ticket.assignee else "Unassigned"
            write_history(ticket, "reassign", old_name, new_user.name, current_user.id)
            ticket.assigned_to = assignee_id
            # Notify new assignee
            send_notification(assignee_id, ticket.id,
                              f"You have been assigned to ticket [{ticket.ticket_number}]",
                              event="assigned")
        elif new_assigned is None and ticket.assigned_to is not None:
            # Unassign: admin/manager explicitly cleared the assignee field
            old_name = ticket.assignee.name if ticket.assignee else "Unassigned"
            write_history(ticket, "reassign", old_name, "Unassigned", current_user.id)
            ticket.assigned_to = None

        ticket.updated_at = utc_now()
        db.session.commit()
        flash(t("flash_update_ok"), "success")
    return redirect(url_for("employee.ticket_detail", ticket_id=ticket_id))


# ─────────────────────────────────────────────
# BULK ACTIONS (3)
# ─────────────────────────────────────────────

@admin_bp.route("/tickets/bulk-action", methods=["POST"])
@manager_required
def bulk_action():
    """
    Apply a single action to multiple tickets at once.
    Form fields:
      ticket_ids  — comma-separated ticket IDs (built by JS checkboxes)
      action      — "assign" | "close" | "change_status"
      assigned_to — user ID  (when action == "assign")
      new_status  — status   (when action == "change_status")
    """
    form = EmptyForm()
    if not form.validate_on_submit():
        flash(t("err_csrf"), "danger")
        return redirect(url_for("admin.tickets"))

    raw_ids = request.form.get("ticket_ids", "")
    action  = request.form.get("action", "")

    try:
        ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()]
    except ValueError:
        ids = []

    if not ids:
        flash(t("err_no_tickets_selected"), "warning")
        return redirect(url_for("admin.tickets"))

    tickets_q = Ticket.query.filter(Ticket.id.in_(ids), Ticket.is_deleted == False)
    if current_user.role == "manager":
        tickets_q = tickets_q.filter(Ticket.department_id == current_user.department_id)

    tickets_list = tickets_q.all()
    if not tickets_list:
        flash(t("err_no_valid_tickets"), "warning")
        return redirect(url_for("admin.tickets"))

    updated = 0
    for ticket in tickets_list:
        changed = False   # track whether this ticket was actually modified

        if action == "assign":
            new_uid = request.form.get("assigned_to", "")
            if new_uid.isdigit():
                new_user = db.session.get(User, int(new_uid))
                if new_user and new_user.active and new_user.role in ("admin", "manager", "employee"):
                    old_name = ticket.assignee.name if ticket.assignee else "Unassigned"
                    write_history(ticket, "reassign", old_name, new_user.name, current_user.id)
                    ticket.assigned_to = new_user.id
                    send_notification(new_user.id, ticket.id,
                                      f"You were bulk-assigned to [{ticket.ticket_number}]",
                                      event="assigned")
                    updated += 1
                    changed = True

        elif action == "close":
            if ticket.status != "Closed":
                write_history(ticket, "status_change", ticket.status, "Closed", current_user.id)
                send_notification(ticket.created_by, ticket.id,
                                  f"Your ticket [{ticket.ticket_number}] was closed.",
                                  event="resolved")
                ticket.status = "Closed"
                updated += 1
                changed = True

        elif action == "change_status":
            new_status = request.form.get("new_status", "")
            if new_status in STATUSES and new_status != ticket.status:
                write_history(ticket, "status_change", ticket.status, new_status, current_user.id)
                _bulk_ev = "resolved" if new_status in ("Resolved", "Closed") else ""
                send_notification(ticket.created_by, ticket.id,
                                  f"Your ticket [{ticket.ticket_number}] status changed to: {new_status}",
                                  event=_bulk_ev)
                ticket.status = new_status
                updated += 1
                changed = True

        if changed:   # only stamp updated_at when a real change was applied
            ticket.updated_at = utc_now()

    db.session.commit()
    flash(t("flash_bulk_done", count=updated), "success")
    return redirect(url_for("admin.tickets"))


# ─────────────────────────────────────────────
# SOFT DELETE TICKET (Admin only)
# ─────────────────────────────────────────────

@admin_bp.route("/tickets/<int:ticket_id>/delete", methods=["POST"])
@admin_required
def delete_ticket(ticket_id):
    """
    Soft-delete a ticket: sets is_deleted=True and deleted_at timestamp.
    The ticket remains in the database and can be restored from the Deleted tab.
    Writes a history row before deletion so the audit trail is preserved.
    """
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=False).first_or_404()
    # Write audit row first (before status changes)
    write_history(ticket, "status_change", ticket.status, "Deleted", current_user.id)
    ticket.is_deleted = True
    ticket.deleted_at = utc_now()
    ticket.updated_at = utc_now()
    db.session.commit()
    flash(t("flash_ticket_deleted", num=ticket.ticket_number), "warning")
    # Redirect: if came from ticket_detail go to tickets list; else stay on list
    return redirect(url_for("admin.tickets"))


# ─────────────────────────────────────────────
# RESTORE DELETED TICKETS (4)
# ─────────────────────────────────────────────

@admin_bp.route("/tickets/deleted")
@admin_required
def deleted_tickets():
    """Show soft-deleted tickets — Admin only (Deleted Tab)."""
    page    = request.args.get("page", 1, type=int)
    tickets = (Ticket.query
               .filter_by(is_deleted=True)
               .order_by(Ticket.deleted_at.desc())
               .paginate(page=page, per_page=20, error_out=False))
    form = EmptyForm()
    return render_template_string(
        TEMPLATES["templates/deleted_tickets.html"],
        tickets=tickets, form=form,
    )


@admin_bp.route("/tickets/<int:ticket_id>/restore", methods=["POST"])
@admin_required
def restore_ticket(ticket_id):
    """Restore a soft-deleted ticket to its previous status."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)
    ticket = Ticket.query.filter_by(id=ticket_id, is_deleted=True).first_or_404()
    ticket.is_deleted = False
    ticket.deleted_at = None
    ticket.updated_at = utc_now()
    write_history(ticket, "status_change", "Deleted", ticket.status, current_user.id)
    db.session.commit()
    flash(t("flash_ticket_restored", num=ticket.ticket_number), "success")
    return redirect(url_for("admin.deleted_tickets"))


# ─────────────────────────────────────────────
# PHASE 4 — REPORTS
# ─────────────────────────────────────────────
@admin_bp.route("/reports")
@manager_required
def reports():
    """
    System reports — admin/manager only.
    Manager sees only their department data.
    """
    from sqlalchemy import func, case

    is_manager = current_user.role == "manager"
    dept_filter = current_user.department_id if is_manager else None
    now = local_now()

    # ── 1. Tickets by Type ──────────────────────────────────
    type_q = (
        db.session.query(
            Ticket.type,
            func.count(Ticket.id).label("total"),
            func.sum(case((Ticket.status == "Open", 1), else_=0)).label("open"),
            func.sum(case((Ticket.status == "Resolved", 1), else_=0)).label("resolved"),
            func.sum(case((Ticket.status == "Closed", 1), else_=0)).label("closed"),
        )
        .filter(Ticket.is_deleted == False)
    )
    if dept_filter:
        type_q = type_q.filter(Ticket.department_id == dept_filter)
    type_rows = type_q.group_by(Ticket.type).order_by(func.count(Ticket.id).desc()).all()
    by_type = [
        {"type": r.type, "total": r.total, "open": r.open or 0,
         "resolved": r.resolved or 0, "closed": r.closed or 0}
        for r in type_rows
    ]

    # ── 2. Avg Resolution Time by Department ────────────────
    # extract('epoch', ...) is PostgreSQL-only; SQLite uses strftime('%s').
    is_postgres = db.engine.dialect.name == "postgresql"
    if is_postgres:
        _diff_seconds = (
            func.extract("epoch", Ticket.updated_at) -
            func.extract("epoch", Ticket.created_at)
        )
    else:
        # SQLite: strftime('%s', ...) returns Unix timestamp as text → cast to int
        _diff_seconds = (
            db.cast(func.strftime("%s", Ticket.updated_at), db.Integer) -
            db.cast(func.strftime("%s", Ticket.created_at), db.Integer)
        )

    resolved_q = (
        db.session.query(
            Department.name.label("dept_name"),
            func.count(Ticket.id).label("resolved_count"),
            func.avg(_diff_seconds).label("avg_seconds"),
        )
        .join(Department, Department.id == Ticket.department_id)
        .filter(
            Ticket.is_deleted == False,
            Ticket.status.in_(["Resolved", "Closed"]),
        )
    )
    if dept_filter:
        resolved_q = resolved_q.filter(Ticket.department_id == dept_filter)
    resolved_rows = resolved_q.group_by(Department.name).order_by(Department.name).all()
    avg_resolution = [
        {"dept": r.dept_name, "resolved_count": r.resolved_count,
         "avg_hours": round(r.avg_seconds / 3600, 1) if r.avg_seconds else None}
        for r in resolved_rows
    ]

    # ── 3. Tickets per Agent ─────────────────────────────────
    if is_postgres:
        _agent_diff_seconds = (
            func.extract("epoch", Ticket.updated_at) -
            func.extract("epoch", Ticket.created_at)
        )
    else:
        _agent_diff_seconds = (
            db.cast(func.strftime("%s", Ticket.updated_at), db.Integer) -
            db.cast(func.strftime("%s", Ticket.created_at), db.Integer)
        )

    agent_q = (
        db.session.query(
            User.name.label("agent_name"),
            func.count(Ticket.id).label("total"),
            func.sum(case((Ticket.status == "Open", 1), else_=0)).label("open"),
            func.sum(case((Ticket.status.in_(["Resolved", "Closed"]), 1), else_=0)).label("resolved"),
            func.avg(
                case(
                    (Ticket.status.in_(["Resolved", "Closed"]),
                     _agent_diff_seconds),
                    else_=None,
                )
            ).label("avg_seconds"),
        )
        .join(Ticket, Ticket.assigned_to == User.id)
        .filter(Ticket.is_deleted == False, User.active == True)
    )
    if dept_filter:
        agent_q = agent_q.filter(Ticket.department_id == dept_filter)
    agent_rows = agent_q.group_by(User.name).order_by(func.count(Ticket.id).desc()).all()
    per_agent = [
        {"name": r.agent_name, "total": r.total,
         "open": r.open or 0, "resolved": r.resolved or 0,
         "avg_hours": round(r.avg_seconds / 3600, 1) if r.avg_seconds else None}
        for r in agent_rows
    ]

    # ── 4. SLA Compliance Rate by Priority ───────────────────
    sla_q = (
        db.session.query(
            Ticket.priority,
            func.count(Ticket.id).label("total_resolved"),
            func.sum(case((Ticket.sla_breached == False, 1), else_=0)).label("within_sla"),
        )
        .filter(Ticket.is_deleted == False, Ticket.status.in_(["Resolved", "Closed"]))
    )
    if dept_filter:
        sla_q = sla_q.filter(Ticket.department_id == dept_filter)
    sla_rows = sla_q.group_by(Ticket.priority).all()
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    sla_compliance = sorted(
        [{"priority": r.priority, "total_resolved": r.total_resolved,
          "within_sla": r.within_sla or 0} for r in sla_rows],
        key=lambda x: priority_order.get(x["priority"], 99)
    )

    # ── 5. Overdue Tickets ────────────────────────────────────
    open_statuses = ("Open", "In Progress", "Waiting for Customer",
                     "Waiting for Vendor", "Reopened")
    overdue_q = Ticket.query.filter(
        Ticket.is_deleted == False,
        Ticket.status.in_(open_statuses),
        Ticket.sla_breached == True,
    ).order_by(Ticket.created_at.asc())
    if dept_filter:
        overdue_q = overdue_q.filter(Ticket.department_id == dept_filter)
    overdue_tickets = overdue_q.options(
        joinedload(Ticket.department),  # template: tk.department.name  — without this: 1 SELECT per row
        joinedload(Ticket.assignee),    # template: tk.assignee.name    — without this: 1 SELECT per row
    ).all()

    return render_template_string(
        TEMPLATES["templates/reports.html"],
        by_type=by_type,
        avg_resolution=avg_resolution,
        per_agent=per_agent,
        sla_compliance=sla_compliance,
        overdue_tickets=overdue_tickets,
        now=now.strftime("%Y-%m-%d %H:%M"),
    )


# ─────────────────────────────────────────────
# PHASE 4 — EXCEL EXPORT
# ─────────────────────────────────────────────
@admin_bp.route("/reports/export")
@manager_required
def export_reports():
    """
    Export the full reports dataset as a multi-sheet Excel workbook.

    Sheets:
        1. Tickets by Type
        2. Avg Resolution by Department
        3. Tickets per Agent
        4. SLA Compliance by Priority
        5. Overdue Tickets
        6. All Tickets (raw data dump)

    Access: admin / manager only (same as reports page).
    Manager sees only their department data.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash(t("err_openpyxl"), "danger")
        return redirect(url_for("admin.reports"))
    from sqlalchemy import func, case

    is_manager = current_user.role == "manager"
    dept_filter = current_user.department_id if is_manager else None
    now = local_now()

    # ── Shared style helpers ────────────────────────────────────────────────
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")   # light blue
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN_BORDER  = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header_row(ws, row_idx, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font       = HEADER_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = CENTER
            cell.border     = THIN_BORDER

    def style_data_row(ws, row_idx, col_count, alternate=False):
        fill = ALT_FILL if alternate else None
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=c)
            if fill:
                cell.fill = fill
            cell.alignment = LEFT
            cell.border    = THIN_BORDER

    def auto_width(ws):
        """Set column widths based on content (max 50)."""
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 1: Tickets by Type
    # ────────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Tickets by Type")
    headers1 = ["Type", "Total", "Open", "Resolved", "Closed"]
    ws1.append(headers1)
    style_header_row(ws1, 1, len(headers1))

    type_q = (
        db.session.query(
            Ticket.type,
            func.count(Ticket.id).label("total"),
            func.sum(case((Ticket.status == "Open", 1), else_=0)).label("open"),
            func.sum(case((Ticket.status == "Resolved", 1), else_=0)).label("resolved"),
            func.sum(case((Ticket.status == "Closed", 1), else_=0)).label("closed"),
        ).filter(Ticket.is_deleted == False)
    )
    if dept_filter:
        type_q = type_q.filter(Ticket.department_id == dept_filter)
    for i, r in enumerate(type_q.group_by(Ticket.type).order_by(func.count(Ticket.id).desc()).all(), start=2):
        ws1.append([r.type, r.total, r.open or 0, r.resolved or 0, r.closed or 0])
        style_data_row(ws1, i, len(headers1), alternate=(i % 2 == 0))
    auto_width(ws1)

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 2: Avg Resolution by Department
    # ────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Avg Resolution by Dept")
    headers2 = ["Department", "Resolved Count", "Avg Resolution (hours)"]
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))

    is_postgres = db.engine.dialect.name == "postgresql"
    if is_postgres:
        _diff = func.extract("epoch", Ticket.updated_at) - func.extract("epoch", Ticket.created_at)
    else:
        _diff = (
            db.cast(func.strftime("%s", Ticket.updated_at), db.Integer) -
            db.cast(func.strftime("%s", Ticket.created_at), db.Integer)
        )
    res_q = (
        db.session.query(
            Department.name.label("dept_name"),
            func.count(Ticket.id).label("resolved_count"),
            func.avg(_diff).label("avg_seconds"),
        )
        .join(Department, Department.id == Ticket.department_id)
        .filter(Ticket.is_deleted == False, Ticket.status.in_(["Resolved", "Closed"]))
    )
    if dept_filter:
        res_q = res_q.filter(Ticket.department_id == dept_filter)
    for i, r in enumerate(res_q.group_by(Department.name).order_by(Department.name).all(), start=2):
        avg_h = round(r.avg_seconds / 3600, 1) if r.avg_seconds else "N/A"
        ws2.append([r.dept_name, r.resolved_count, avg_h])
        style_data_row(ws2, i, len(headers2), alternate=(i % 2 == 0))
    auto_width(ws2)

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 3: Tickets per Agent
    # ────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Tickets per Agent")
    headers3 = ["Agent", "Total", "Open", "Resolved / Closed", "Avg Resolution (hours)"]
    ws3.append(headers3)
    style_header_row(ws3, 1, len(headers3))

    if is_postgres:
        _adiff = func.extract("epoch", Ticket.updated_at) - func.extract("epoch", Ticket.created_at)
    else:
        _adiff = (
            db.cast(func.strftime("%s", Ticket.updated_at), db.Integer) -
            db.cast(func.strftime("%s", Ticket.created_at), db.Integer)
        )
    agent_q = (
        db.session.query(
            User.name.label("agent_name"),
            func.count(Ticket.id).label("total"),
            func.sum(case((Ticket.status == "Open", 1), else_=0)).label("open"),
            func.sum(case((Ticket.status.in_(["Resolved", "Closed"]), 1), else_=0)).label("resolved"),
            func.avg(
                case((Ticket.status.in_(["Resolved", "Closed"]), _adiff), else_=None)
            ).label("avg_seconds"),
        )
        .join(Ticket, Ticket.assigned_to == User.id)
        .filter(Ticket.is_deleted == False, User.active == True)
    )
    if dept_filter:
        agent_q = agent_q.filter(Ticket.department_id == dept_filter)
    for i, r in enumerate(agent_q.group_by(User.name).order_by(func.count(Ticket.id).desc()).all(), start=2):
        avg_h = round(r.avg_seconds / 3600, 1) if r.avg_seconds else "N/A"
        ws3.append([r.agent_name, r.total, r.open or 0, r.resolved or 0, avg_h])
        style_data_row(ws3, i, len(headers3), alternate=(i % 2 == 0))
    auto_width(ws3)

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 4: SLA Compliance
    # ────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("SLA Compliance")
    headers4 = ["Priority", "Total Resolved", "Within SLA", "Breached", "Compliance %"]
    ws4.append(headers4)
    style_header_row(ws4, 1, len(headers4))

    sla_q = (
        db.session.query(
            Ticket.priority,
            func.count(Ticket.id).label("total_resolved"),
            func.sum(case((Ticket.sla_breached == False, 1), else_=0)).label("within_sla"),
        )
        .filter(Ticket.is_deleted == False, Ticket.status.in_(["Resolved", "Closed"]))
    )
    if dept_filter:
        sla_q = sla_q.filter(Ticket.department_id == dept_filter)
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    sla_rows = sorted(sla_q.group_by(Ticket.priority).all(),
                      key=lambda x: priority_order.get(x.priority, 99))
    for i, r in enumerate(sla_rows, start=2):
        within = r.within_sla or 0
        breached = r.total_resolved - within
        pct = round(within / r.total_resolved * 100, 1) if r.total_resolved else "N/A"
        ws4.append([r.priority, r.total_resolved, within, breached, pct])
        style_data_row(ws4, i, len(headers4), alternate=(i % 2 == 0))
        # Colour-code compliance cell
        pct_cell = ws4.cell(row=i, column=5)
        if isinstance(pct, (int, float)):
            if pct < 70:
                pct_cell.fill = PatternFill("solid", fgColor="FF9999")  # red
            elif pct < 90:
                pct_cell.fill = PatternFill("solid", fgColor="FFE599")  # yellow
            else:
                pct_cell.fill = PatternFill("solid", fgColor="A9D08E")  # green
    auto_width(ws4)

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 5: Overdue Tickets
    # ────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Overdue Tickets")
    headers5 = ["Ticket #", "Title", "Priority", "Status", "Department",
                "Assignee", "Created At", "SLA Deadline"]
    ws5.append(headers5)
    style_header_row(ws5, 1, len(headers5))

    open_statuses = ("Open", "In Progress", "Waiting for Customer", "Waiting for Vendor", "Reopened")
    overdue_q = (
        Ticket.query
        .filter(Ticket.is_deleted == False, Ticket.status.in_(open_statuses),
                Ticket.sla_breached == True)
        .options(joinedload(Ticket.department), joinedload(Ticket.assignee))
        .order_by(Ticket.created_at.asc())
    )
    if dept_filter:
        overdue_q = overdue_q.filter(Ticket.department_id == dept_filter)
    for i, tk in enumerate(overdue_q.all(), start=2):
        ws5.append([
            tk.ticket_number,
            tk.title,
            tk.priority,
            tk.status,
            tk.department.name if tk.department else "—",
            tk.assignee.name  if tk.assignee  else "Unassigned",
            utc_to_local(tk.created_at).strftime("%Y-%m-%d %H:%M"),
            utc_to_local(tk.sla_deadline).strftime("%Y-%m-%d %H:%M"),
        ])
        row_fill = PatternFill("solid", fgColor="FF9999")   # red — these are all breached
        for c in range(1, len(headers5) + 1):
            cell = ws5.cell(row=i, column=c)
            cell.fill      = row_fill
            cell.alignment = LEFT
            cell.border    = THIN_BORDER
    auto_width(ws5)

    # ────────────────────────────────────────────────────────────────────────
    # Sheet 6: All Tickets (raw dump)
    # ────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("All Tickets")
    headers6 = ["Ticket #", "Title", "Type", "Priority", "Status",
                "Department", "Created By", "Assignee",
                "Created At", "Updated At", "SLA Deadline", "SLA Breached"]
    ws6.append(headers6)
    style_header_row(ws6, 1, len(headers6))

    all_q = (
        Ticket.query
        .filter(Ticket.is_deleted == False)
        .options(
            joinedload(Ticket.department),
            joinedload(Ticket.creator),
            joinedload(Ticket.assignee),
        )
        .order_by(Ticket.created_at.desc())
    )
    if dept_filter:
        all_q = all_q.filter(Ticket.department_id == dept_filter)
    for i, tk in enumerate(all_q.all(), start=2):
        ws6.append([
            tk.ticket_number,
            tk.title,
            tk.type,
            tk.priority,
            tk.status,
            tk.department.name  if tk.department else "—",
            tk.creator.name     if tk.creator    else "—",
            tk.assignee.name    if tk.assignee   else "Unassigned",
            utc_to_local(tk.created_at).strftime("%Y-%m-%d %H:%M"),
            utc_to_local(tk.updated_at).strftime("%Y-%m-%d %H:%M") if tk.updated_at else "—",
            utc_to_local(tk.sla_deadline).strftime("%Y-%m-%d %H:%M"),
            "Yes" if tk.sla_breached else "No",
        ])
        style_data_row(ws6, i, len(headers6), alternate=(i % 2 == 0))
        # Highlight SLA breached cell
        if tk.sla_breached:
            ws6.cell(row=i, column=12).fill = PatternFill("solid", fgColor="FF9999")
    auto_width(ws6)

    # ── Stream to response ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"tickets_report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─────────────────────────────────────────────
# PHASE 4 — FULL-TEXT SEARCH
# ─────────────────────────────────────────────
@admin_bp.route("/search")
@manager_required
def search():
    """
    Full-text search: title + description + ticket_number + comment body.
    Uses PostgreSQL tsvector / plainto_tsquery when on PostgreSQL (production).
    Falls back to ILIKE for SQLite (automated testing only).
    GIN index to add after first migration:
      CREATE INDEX idx_tickets_fts ON tickets
        USING GIN(to_tsvector('english', coalesce(title,'') || ' ' || coalesce(description,'')));
    """
    from sqlalchemy import func   # required for plainto_tsquery / to_tsvector (PostgreSQL FTS)
    q    = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int)

    base_query = Ticket.query.filter(Ticket.is_deleted == False)
    if current_user.role == "manager":
        base_query = base_query.filter(Ticket.department_id == current_user.department_id)

    if q:
        is_postgres = db.engine.dialect.name == "postgresql"
        if is_postgres:
            # ── PostgreSQL full-text search via tsvector ─────────────
            tsq = func.plainto_tsquery("english", q)
            matching_comment_ids = (
                db.session.query(Comment.ticket_id)
                .filter(
                    func.to_tsvector("english",
                        db.func.coalesce(Comment.body, "")
                    ).op("@@")(tsq)
                )
                .scalar_subquery()
            )
            base_query = base_query.filter(
                db.or_(
                    func.to_tsvector(
                        "english",
                        db.func.coalesce(Ticket.title, "") + " " +
                        db.func.coalesce(Ticket.description, "")
                    ).op("@@")(tsq),
                    Ticket.ticket_number.ilike(f"%{q}%"),   # ticket numbers are alphanumeric — FTS won't match TKT-2025-0001
                    Ticket.id.in_(matching_comment_ids),
                )
            )
        else:
            # ── SQLite fallback — testing environment only ────────────
            term = f"%{q}%"
            matching_comment_ids = (
                db.session.query(Comment.ticket_id)
                .filter(Comment.body.ilike(term))
                .scalar_subquery()
            )
            base_query = base_query.filter(
                db.or_(
                    Ticket.title.ilike(term),
                    Ticket.description.ilike(term),
                    Ticket.ticket_number.ilike(term),
                    Ticket.id.in_(matching_comment_ids),
                )
            )

    results = base_query.options(
        joinedload(Ticket.creator),     # template: tk.creator.name    — without this: 1 SELECT per row
        joinedload(Ticket.department),  # template: tk.department.name — without this: 1 SELECT per row
    ).order_by(Ticket.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template_string(
        TEMPLATES["templates/search_results.html"],
        query=q,
        results=results,
    )

@admin_bp.route("/users")
@admin_required
def users():
    all_users = User.query.order_by(User.created_at.desc()).all()
    return render_template_string(TEMPLATES["templates/users.html"], users=all_users)

@admin_bp.route("/users/new", methods=["GET", "POST"])
@admin_required
def new_user():
    form = EmptyForm()
    departments = Department.query.filter_by(is_deleted=False).all()
    if request.method == "POST" and form.validate_on_submit():
        email    = request.form.get("email", "").strip().lower()
        name     = request.form.get("name", "").strip()
        password  = request.form.get("password", "")
        password2 = request.form.get("password2", "")
        username  = request.form.get("username", "").strip() or None
        # ── Server-side validation (do not rely on browser alone) ──
        errors = []
        if not name:
            errors.append("Name is required")
        if not email:
            errors.append("Email is required")
        if not password:
            errors.append("Password is required")
        else:
            errors.extend(validate_password(password))
            if password != password2:
                errors.append(t("err_pw_no_match"))
        if email and User.query.filter_by(email=email).first():
            errors.append(t("err_email_taken"))
        if username:
            uname_err = validate_username(username)
            if uname_err:
                errors.append(uname_err)
            elif User.query.filter(
                db.func.lower(User.username) == username.lower()
            ).first():
                errors.append(t("err_username_taken"))
        if errors:
            for err in errors:
                flash(err, "danger")
        else:
            u = User(
                name          = name,
                email         = email,
                username      = username,
                role          = request.form.get("role", "employee") if request.form.get("role", "employee") in ("employee", "manager", "admin") else "employee",
                department_id = request.form.get("department_id") or None,
            )
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            flash(t("flash_user_ok", name=u.name), "success")
            return redirect(url_for("admin.users"))
    # Inline simple form for now
    # _html_escape() prevents XSS: d.name comes from DB (admin-entered) but must be escaped
    # because this f-string is inserted into the template BEFORE Jinja2 sees it,
    # so auto-escaping never applies here.
    depts_opts = "".join(f'<option value="{d.id}">{_html_escape(d.name)}</option>' for d in departments)
    html = f"""
    {{% extends 'base.html' %}}
    {{% block title %}}{{{{ t('new_user') }}}}{{% endblock %}}
    {{% block content %}}
    <div class="row justify-content-center"><div class="col-md-5">
    <div class="card border-0 shadow-sm"><div class="card-header fw-semibold">{{{{ t('new_user') }}}}</div>
    <div class="card-body p-4">
    <form method="POST">
      {{{{ form.hidden_tag() }}}}
      <div class="mb-3"><label class="form-label">{{{{ t('name') }}}} <span class="text-danger">*</span></label>
        <input name="name" class="form-control" required></div>
      <div class="mb-3"><label class="form-label">{{{{ t('email') }}}} <span class="text-danger">*</span></label>
        <input type="email" name="email" class="form-control" required></div>
      <div class="mb-3">
        <label class="form-label">{{{{ t('username_lbl') }}}}</label>
        <input type="text" name="username" class="form-control"
               placeholder="{{{{ t('username_placeholder') }}}}">
        <div class="form-text">{{{{ t('username_field') }}}}</div>
      </div>
      <div class="mb-3"><label class="form-label">{{{{ t('password') }}}} <span class="text-danger">*</span></label>
        <div class="input-group">
          <input type="password" name="password" id="nu_pw" class="form-control" required>
          <button type="button" class="btn btn-outline-secondary"
                  onclick="togglePw('nu_pw','nu_pw_ico')">
            <i id="nu_pw_ico" class="bi bi-eye"></i>
          </button>
        </div>
        <div class="form-text">
          <strong>{{{{ t('pw_policy_title') }}}}:</strong>
          {{{{ t('pw_min_chars') }}}} · {{{{ t('pw_uppercase') }}}} · {{{{ t('pw_digit') }}}} · {{{{ t('pw_special') }}}}
        </div>
      </div>
      <div class="mb-3"><label class="form-label">{{{{ t('confirm_password') }}}} <span class="text-danger">*</span></label>
        <div class="input-group">
          <input type="password" name="password2" id="nu_pw2" class="form-control" required>
          <button type="button" class="btn btn-outline-secondary"
                  onclick="togglePw('nu_pw2','nu_pw2_ico')">
            <i id="nu_pw2_ico" class="bi bi-eye"></i>
          </button>
        </div>
      </div>
      <div class="mb-3"><label class="form-label">{{{{ t('role') }}}}</label>
        <select name="role" class="form-select">
          <option value="employee">Employee</option>
          <option value="manager">Manager</option>
          <option value="admin">Admin</option>
        </select></div>
      <div class="mb-3"><label class="form-label">{{{{ t('dept') }}}}</label>
        <select name="department_id" class="form-select">
          <option value="">-- No Department --</option>
          {depts_opts}
        </select></div>
      <button type="submit" class="btn btn-primary w-100">{{{{ t('new_user') }}}}</button>
    </form></div></div></div></div>
    <script>
    function togglePw(inputId, iconId) {{
      var el = document.getElementById(inputId);
      var ic = document.getElementById(iconId);
      if (el.type === 'password') {{
        el.type = 'text';
        ic.className = 'bi bi-eye-slash';
      }} else {{
        el.type = 'password';
        ic.className = 'bi bi-eye';
      }}
    }}
    </script>
    {{% endblock %}}
    """
    return render_template_string(html, form=form)

@admin_bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_user(user_id):
    u = db.get_or_404(User, user_id)
    form = EmptyForm()
    if request.method == "POST" and form.validate_on_submit():
        new_email    = request.form.get("email", "").strip().lower()
        new_username = request.form.get("username", "").strip() or None
        errors = []
        # Email uniqueness check (skip if unchanged)
        if new_email and new_email != u.email:
            if User.query.filter_by(email=new_email).first():
                errors.append(t("err_email_taken"))
        # Username format + uniqueness check (skip if unchanged)
        if new_username and new_username != u.username:
            uname_err = validate_username(new_username)
            if uname_err:
                errors.append(uname_err)
            elif User.query.filter(db.func.lower(User.username) == new_username.lower()).first():
                errors.append(t("err_username_taken"))
        if errors:
            for err in errors:
                flash(err, "danger")
            return redirect(request.url)
        u.name          = request.form.get("name", u.name).strip()
        u.email         = new_email or u.email
        u.username      = new_username
        new_role = request.form.get("role", u.role)
        if new_role in ("employee", "manager", "admin"):
            u.role = new_role
        u.active        = request.form.get("active") == "1"
        u.on_leave      = request.form.get("on_leave") == "1"
        u.department_id = request.form.get("department_id") or None
        if request.form.get("password"):
            pw_errors = validate_password(request.form["password"])
            if pw_errors:
                for err in pw_errors:
                    flash(err, "danger")
                return redirect(request.url)
            u.set_password(request.form["password"])
        db.session.commit()
        flash(t("flash_user_upd"), "success")
        return redirect(url_for("admin.users"))
    departments = Department.query.filter_by(is_deleted=False).all()
    depts_opts  = "".join(
        f'<option value="{d.id}" {"selected" if u.department_id == d.id else ""}>{_html_escape(d.name)}</option>'
        for d in departments
    )
    html = f"""
    {{% extends 'base.html' %}}
    {{% block title %}}{{{{ t('edit') }}}} — {{{{ u.name }}}}{{% endblock %}}
    {{% block content %}}
    <div class="row justify-content-center"><div class="col-md-5">
    <div class="card border-0 shadow-sm">
      <div class="card-header fw-semibold">{{{{ t('edit') }}}}: {{{{ u.name }}}}</div>
    <div class="card-body p-4">
    <form method="POST">
      {{{{ form.hidden_tag() }}}}
      <div class="mb-3"><label class="form-label">{{{{ t('name') }}}}</label>
        <input name="name" class="form-control" value="{{{{ u.name }}}}" required></div>
      <div class="mb-3"><label class="form-label">{{{{ t('email') }}}}</label>
        <input type="email" name="email" class="form-control" value="{{{{ u.email }}}}" required></div>
      <div class="mb-3">
        <label class="form-label">{{{{ t('username_lbl') }}}}</label>
        <input type="text" name="username" class="form-control"
               value="{{{{ u.username or '' }}}}"
               placeholder="{{{{ t('username_placeholder') }}}}">
        <div class="form-text">{{{{ t('username_field') }}}}</div>
      </div>
      <div class="mb-3">
        <label class="form-label">{{{{ t('password') }}}} <small class="text-muted">({{{{ t('leave_blank_pw') }}}})</small></label>
        <div class="input-group">
          <input type="password" name="password" id="eu_pw" class="form-control">
          <button type="button" class="btn btn-outline-secondary"
                  onclick="var e=document.getElementById('eu_pw');var i=document.getElementById('eu_pw_ico');if(e.type==='password'){{e.type='text';i.className='bi bi-eye-slash';}}else{{e.type='password';i.className='bi bi-eye';}}">
            <i id="eu_pw_ico" class="bi bi-eye"></i>
          </button>
        </div>
      </div>
      <div class="mb-3"><label class="form-label">{{{{ t('role') }}}}</label>
        <select name="role" class="form-select">
          <option value="employee" {'selected' if u.role == 'employee' else ''}>Employee</option>
          <option value="manager"  {'selected' if u.role == 'manager'  else ''}>Manager</option>
          <option value="admin"    {'selected' if u.role == 'admin'    else ''}>Admin</option>
        </select></div>
      <div class="mb-3"><label class="form-label">{{{{ t('dept') }}}}</label>
        <select name="department_id" class="form-select">
          <option value="">-- {{{{ t('na') }}}} --</option>
          {depts_opts}
        </select></div>
      <div class="mb-3"><label class="form-label">{{{{ t('status') }}}}</label>
        <select name="active" class="form-select">
          <option value="1" {'selected' if u.active else ''}>{{{{ t('active') }}}}</option>
          <option value="0" {'selected' if not u.active else ''}>{{{{ t('disabled') }}}}</option>
        </select></div>
      <div class="mb-3"><label class="form-label">{{{{ t('on_leave') }}}}</label>
        <select name="on_leave" class="form-select">
          <option value="0" {'selected' if not u.on_leave else ''}>{{{{ t('not_on_leave') }}}}</option>
          <option value="1" {'selected' if u.on_leave else ''}>{{{{ t('on_leave') }}}}</option>
        </select>
        <div class="form-text text-muted">{{{{ t('on_leave_note') }}}}</div>
      </div>
      <div class="d-flex gap-2">
        <button type="submit" class="btn btn-primary">{{{{ t('save_changes') }}}}</button>
        <a href="{{{{ url_for('admin.users') }}}}" class="btn btn-outline-secondary">{{{{ t('cancel') }}}}</a>
      </div>
    </form></div></div></div></div>
    {{% endblock %}}
    """
    return render_template_string(html, form=form, u=u)


# ─────────────────────────────────────────────
# DEPARTMENT MANAGEMENT (Admin only)
# ─────────────────────────────────────────────

@admin_bp.route("/departments")
@admin_required
def departments():
    """List all departments with ticket counts and manager info."""
    from sqlalchemy import func

    # Get all depts (active + deleted) ordered: active first, then by name
    all_depts = (Department.query
                 .order_by(Department.is_deleted.asc(), Department.name.asc())
                 .all())

    # Ticket counts per department
    ticket_counts = dict(
        db.session.query(
            Ticket.department_id,
            func.count(Ticket.id)
        ).filter(Ticket.is_deleted == False)
         .group_by(Ticket.department_id)
         .all()
    )
    open_counts = dict(
        db.session.query(
            Ticket.department_id,
            func.count(Ticket.id)
        ).filter(Ticket.is_deleted == False, Ticket.status == "Open")
         .group_by(Ticket.department_id)
         .all()
    )

    dept_rows = [
        (d, ticket_counts.get(d.id, 0), open_counts.get(d.id, 0))
        for d in all_depts
    ]

    # Map manager_id → User for the template
    manager_ids = [d.manager_id for d in all_depts if d.manager_id]
    managers_list = User.query.filter(User.id.in_(manager_ids)).all() if manager_ids else []
    managers_map  = {u.id: u for u in managers_list}

    # All active managers/admins for the dropdowns
    managers = (User.query
                .filter_by(active=True)
                .filter(User.role.in_(["admin", "manager"]))
                .order_by(User.name)
                .all())

    form = EmptyForm()
    return render_template_string(
        TEMPLATES["templates/departments.html"],
        dept_rows=dept_rows,
        managers_map=managers_map,
        managers=managers,
        ticket_types=TICKET_TYPES,
        ticket_type_i18n=TICKET_TYPE_I18N,
        form=form,
    )


@admin_bp.route("/departments/new", methods=["POST"])
@admin_required
def new_department():
    """Create a new department (POST from modal in departments page)."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    name       = request.form.get("name", "").strip()
    manager_id = request.form.get("manager_id") or None

    if not name:
        flash(t("err_dept_name_required"), "danger")
        return redirect(url_for("admin.departments"))

    # Uniqueness check (case-insensitive, ignore soft-deleted)
    existing = Department.query.filter(
        db.func.lower(Department.name) == name.lower(),
        Department.is_deleted == False,
    ).first()
    if existing:
        flash(t("dept_already_exists"), "danger")
        return redirect(url_for("admin.departments"))

    # Allowed ticket types — checkboxes; empty = allow all
    import json as _json
    selected_types = request.form.getlist("allowed_types")
    valid_types    = [tt for tt in selected_types if tt in TICKET_TYPES]
    allowed_types_json = _json.dumps(valid_types) if valid_types else None

    dept = Department(
        name          = name,
        manager_id    = int(manager_id) if manager_id else None,
        allowed_types = allowed_types_json,
    )
    db.session.add(dept)
    db.session.commit()
    flash(t("flash_dept_created", name=name), "success")
    return redirect(url_for("admin.departments"))


@admin_bp.route("/departments/edit", methods=["POST"])
@admin_required
def edit_department():
    """Edit an existing department's name and/or manager."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    dept_id    = request.form.get("dept_id")
    name       = request.form.get("name", "").strip()
    manager_id = request.form.get("manager_id") or None

    if not dept_id or not name:
        flash(t("err_invalid_request"), "danger")
        return redirect(url_for("admin.departments"))

    try:
        dept_id_int = int(dept_id)
    except (ValueError, TypeError):
        flash(t("err_invalid_request"), "danger")
        return redirect(url_for("admin.departments"))

    dept = db.session.get(Department, dept_id_int)
    if not dept:
        abort(404)

    # Uniqueness: skip if name unchanged
    if name.lower() != dept.name.lower():
        existing = Department.query.filter(
            db.func.lower(Department.name) == name.lower(),
            Department.is_deleted == False,
            Department.id != dept.id,
        ).first()
        if existing:
            flash(t("dept_already_exists"), "danger")
            return redirect(url_for("admin.departments"))

    dept.name       = name
    dept.manager_id = int(manager_id) if manager_id else None

    import json as _json
    selected_types = request.form.getlist("allowed_types")
    valid_types    = [tt for tt in selected_types if tt in TICKET_TYPES]
    dept.allowed_types = _json.dumps(valid_types) if valid_types else None

    db.session.commit()
    flash(t("flash_dept_updated"), "success")
    return redirect(url_for("admin.departments"))


@admin_bp.route("/departments/<int:dept_id>/delete", methods=["POST"])
@admin_required
def delete_department(dept_id):
    """
    Soft-delete a department.
    Blocked if the department has active (non-deleted) tickets —
    admin must reassign them to another department first.
    """
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    dept = Department.query.filter_by(id=dept_id, is_deleted=False).first_or_404()

    # Block deletion if active tickets exist
    active_count = Ticket.query.filter_by(
        department_id=dept_id, is_deleted=False
    ).count()
    if active_count > 0:
        flash(t("flash_dept_has_tickets"), "danger")
        return redirect(url_for("admin.departments"))

    dept.is_deleted = True
    dept.deleted_at = utc_now()
    db.session.commit()
    flash(t("flash_dept_deleted"), "success")
    return redirect(url_for("admin.departments"))


@admin_bp.route("/departments/<int:dept_id>/restore", methods=["POST"])
@admin_required
def restore_department(dept_id):
    """Restore a soft-deleted department."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    dept = Department.query.filter_by(id=dept_id, is_deleted=True).first_or_404()

    # Guard: block restore if an active department already has the same name
    conflict = Department.query.filter(
        db.func.lower(Department.name) == dept.name.lower(),
        Department.is_deleted == False,
        Department.id != dept.id,
    ).first()
    if conflict:
        flash(t("dept_already_exists"), "danger")
        return redirect(url_for("admin.departments"))

    dept.is_deleted = False
    dept.deleted_at = None
    db.session.commit()
    flash(t("flash_dept_restored", name=dept.name), "success")
    return redirect(url_for("admin.departments"))


# ─────────────────────────────────────────────
# BACKUP ROUTES
# ─────────────────────────────────────────────

@admin_bp.route("/backups")
@admin_required
def backups_list():
    """Show all backup snapshots, newest first.

    Selects only the columns the listing page actually needs — intentionally
    excludes the large ``data`` TEXT column (full JSON snapshot) so that
    browsing the backup list never pulls megabytes of JSON into memory for
    every row.  The ``data`` column is only read by restore_from_backup() and
    download_backup(), which fetch a single record by PK.
    """
    backups = (
        db.session.query(
            Backup.id,
            Backup.created_at,
            Backup.size_kb,
            Backup.source,
            Backup.email_sent,
            Backup.gdrive_id,
        )
        .order_by(Backup.created_at.desc())
        .all()
    )
    return render_template_string(
        TEMPLATES["templates/backups.html"],
        backups=backups,
        form=EmptyForm(),
    )


@admin_bp.route("/backups/create", methods=["POST"])
@admin_required
def manual_backup():
    """Trigger a manual backup immediately."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    result = create_backup(source="manual")
    if result:
        flash(t("flash_backup_created"), "success")
    else:
        flash(t("flash_backup_error"), "danger")
    return redirect(url_for("admin.backups_list"))


@admin_bp.route("/backups/<int:backup_id>/download")
@admin_required
def download_backup(backup_id):
    """Stream a backup snapshot as a JSON file download."""
    import io
    backup   = db.get_or_404(Backup, backup_id)
    filename = f"backup_{utc_to_local(backup.created_at).strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(
        io.BytesIO(backup.data.encode("utf-8")),
        mimetype="application/json",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/backups/<int:backup_id>/restore", methods=["POST"])
@admin_required
def restore_backup(backup_id):
    """
    Restore the entire database from a saved backup snapshot.

    The restore is performed inside a single transaction; if anything
    fails the database is rolled back and the existing data is untouched.
    """
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    backup = db.get_or_404(Backup, backup_id)
    try:
        restore_from_backup(backup)
        flash(t("flash_backup_restored"), "success")
    except Exception as exc:
        db.session.rollback()
        app.logger.error(f"[Restore] failed for backup id={backup_id}: {exc}")
        flash(t("flash_restore_error"), "danger")
    return redirect(url_for("admin.backups_list"))


@admin_bp.route("/backups/<int:backup_id>/delete", methods=["POST"])
@admin_required
def delete_backup(backup_id):
    """Permanently delete a single backup snapshot."""
    form = EmptyForm()
    if not form.validate_on_submit():
        abort(400)

    backup = db.get_or_404(Backup, backup_id)
    try:
        db.session.delete(backup)
        db.session.commit()
        flash(t("flash_backup_deleted"), "success")
    except Exception as exc:
        db.session.rollback()
        app.logger.error(f"[Backup] delete failed for id={backup_id}: {exc}")
        flash(t("flash_delete_error"), "danger")
    return redirect(url_for("admin.backups_list"))


# ── API (HTMX) ────────────────────────────────
api_bp = Blueprint("api", __name__, url_prefix="/api")


@api_bp.route("/departments/ticket-types")
@login_required
def dept_ticket_types():
    """
    HTMX endpoint: returns <option> elements for the ticket-type select,
    filtered to the types allowed for the chosen department.
    Called when the department dropdown changes in the new-ticket form.
    """
    dept_id = request.args.get("department_id", "").strip()
    lang    = flask_session.get("lang", "en")
    translations = TRANSLATIONS.get(lang, TRANSLATIONS["en"])

    if not dept_id:
        placeholder = translations.get("choose_dept_first_type", "-- Choose Department First --")
        return f'<option value="">{placeholder}</option>'

    try:
        dept_id_int = int(dept_id)
    except (ValueError, TypeError):
        placeholder = translations.get("choose_type", "-- Choose Type --")
        return f'<option value="">{placeholder}</option>'

    dept_obj = Department.query.filter_by(id=dept_id_int, is_deleted=False).first()
    if not dept_obj:
        placeholder = translations.get("choose_type", "-- Choose Type --")
        return f'<option value="">{placeholder}</option>'

    allowed_types = get_types_for_dept(dept_obj.name, dept_obj=dept_obj)
    placeholder   = translations.get("choose_type", "-- Choose Type --")
    options = f'<option value="">{placeholder}</option>'
    for tt in allowed_types:
        i18n_key = TICKET_TYPE_I18N.get(tt, tt)
        label    = translations.get(i18n_key, tt)
        options += f'<option value="{tt}">{label}</option>'
    return options


@api_bp.route("/departments/employees")
@login_required
def dept_employees_by_select():
    """
    HTMX endpoint: returns the <option> fragment directly without redirect.
    HTMX does not follow redirects automatically — redirecting caused empty
    options in some browsers.
    """
    lang         = flask_session.get("lang", "en")
    translations = TRANSLATIONS.get(lang, TRANSLATIONS["en"])

    dept_id = request.args.get("department_id")
    if not dept_id:
        placeholder = translations.get("choose_dept_first", "-- Choose Department First --")
        return f'<option value="">{placeholder}</option>'
    try:
        dept_id = int(dept_id)
    except (ValueError, TypeError):
        placeholder = translations.get("choose_dept_first", "-- Choose Department First --")
        return f'<option value="">{placeholder}</option>'
    users = User.query.filter_by(
        department_id=dept_id, active=True
    ).filter(User.role.in_(["admin", "manager"])).all()
    if not users:
        placeholder = translations.get("no_agents_in_dept", "-- No agents in this department --")
        return f'<option value="">{placeholder}</option>'
    placeholder = translations.get("choose_assignee", "-- Select Assignee --")
    options = f'<option value="">{placeholder}</option>'
    options += "".join(f'<option value="{u.id}">{_html_escape(u.name)}</option>' for u in users)
    return options


@api_bp.route("/filter/dept-agents")
@manager_required
def filter_dept_agents():
    """
    HTMX endpoint for the admin tickets filter panel.
    When a department is selected, returns the assignee <select> populated
    with agents from that department only.
    If no department selected, returns all active agents.
    """
    dept_id = request.args.get("dept", "").strip()
    q = User.query.filter_by(active=True).filter(
        User.role.in_(["admin", "manager"])
    )
    if dept_id:
        try:
            q = q.filter_by(department_id=int(dept_id))
        except (ValueError, TypeError):
            pass
    agents = q.order_by(User.name).all()

    # Use the session language (same source as get_lang() used everywhere in the app),
    # NOT Accept-Language header which reflects browser locale, not the user's in-app choice.
    lang    = flask_session.get("lang", "en")
    all_lbl = TRANSLATIONS.get(lang, TRANSLATIONS["en"]).get("all_assignees", "All Assignees")

    html = f'<select name="assignee" class="form-select form-select-sm"><option value="">{all_lbl}</option>'
    html += "".join(f'<option value="{u.id}">{_html_escape(u.name)}</option>' for u in agents)
    html += "</select>"
    return html


@api_bp.route("/preview-ticket-number")
@login_required
def preview_ticket_number():
    """
    HTMX endpoint: returns the *expected* next ticket number as an HTML badge.
    Used by the new-ticket form to show a live preview before submission.
    This is a READ-ONLY estimate — the real number is generated atomically
    inside the POST handler.  In high-concurrency scenarios the final number
    may be different (incremented by one), which is acceptable for a preview.
    """
    year       = utc_now().year
    base_count = Ticket.query.filter(
        extract("year", Ticket.created_at) == year
    ).count() + 1
    # Walk forward until we find a free slot (READ-ONLY estimate — does not touch ticket_counter)
    for attempt in range(10):
        candidate = f"TKT-{year}-{base_count + attempt:04d}"
        if not Ticket.query.filter_by(ticket_number=candidate).first():
            break
    else:
        candidate = "TKT-{}-???".format(year)   # extremely rare fallback
    return (
        f'<span class="badge bg-secondary fs-6 font-monospace">'
        f'<i class="bi bi-hash"></i> {candidate}'
        f'</span>'
    )


# ─────────────────────────────────────────────
# REGISTER BLUEPRINTS
# ─────────────────────────────────────────────

app.register_blueprint(setup_bp)
app.register_blueprint(auth_bp)
app.register_blueprint(main_bp)
app.register_blueprint(employee_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(api_bp)


# ─────────────────────────────────────────────
# ERROR HANDLERS
# ─────────────────────────────────────────────

@app.errorhandler(429)
def too_many_requests(e):
    return render_template_string(TEMPLATES["templates/429.html"]), 429

@app.errorhandler(403)
def forbidden(e):
    return render_template_string(TEMPLATES["templates/403.html"]), 403

@app.errorhandler(404)
def not_found(e):
    return render_template_string(TEMPLATES["templates/404.html"]), 404

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    app.logger.error(f"[500] Unhandled exception: {e}", exc_info=True)
    return render_template_string(TEMPLATES["templates/500.html"]), 500


# ─────────────────────────────────────────────
# SEED CLI COMMAND  (flask seed)
# ─────────────────────────────────────────────

import click

@app.cli.command("seed")
def seed_command():
    """
    Creates the four default departments only if they do not exist.
    User accounts are created from the first-run setup page in the browser.
    Usage: flask seed
    """
    dept_names = ["IT", "HR", "Finance", "General"]
    created = 0
    for name in dept_names:
        if not Department.query.filter_by(name=name).first():
            db.session.add(Department(name=name))
            created += 1
    db.session.commit()

    if created:
        click.echo(f"✅ Created {created} department(s): {', '.join(dept_names)}")
    else:
        click.echo("✅ Departments already exist — nothing created.")

    click.echo("\n👉 Open your browser at http://127.0.0.1:5000 and create the admin account.")
    click.echo("\n📌 Performance tip — run this SQL once on PostgreSQL for faster full-text search (TC-141):")
    click.echo("   CREATE INDEX IF NOT EXISTS idx_tickets_fts ON tickets")
    click.echo("     USING GIN(to_tsvector('english', coalesce(title,'') || ' ' || coalesce(description,'')));")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

def ensure_columns():
    """
    Auto-adds any missing columns to existing tables without requiring
    manual 'flask db migrate/upgrade' after code changes.
    Safe to run on every startup — dialect-aware:
      • PostgreSQL: uses ADD COLUMN IF NOT EXISTS (atomic, no extra query)
      • SQLite:     checks information_schema manually then adds if absent
        (SQLite does not support IF NOT EXISTS in ALTER TABLE)

    Also creates the PostgreSQL GIN full-text index on first run if it does
    not exist yet.  This is done here (not in __table_args__) because
    SQLAlchemy's postgresql_using kwarg only strips USING from the DDL —
    it does NOT skip the index on non-PostgreSQL dialects, so declaring it
    in the model crashes db.create_all() on SQLite.
    """
    # Each entry: (table, column, pg_type, sqlite_type, default_clause_or_None)
    # pg_type / sqlite_type may differ (e.g. BOOLEAN vs INTEGER for SQLite)
    REQUIRED_COLUMNS = [
        ("users",       "username",            "VARCHAR(60)",      "TEXT",             None),
        ("users",       "on_leave",            "BOOLEAN NOT NULL", "INTEGER NOT NULL", "DEFAULT 0"),
        ("users",       "is_available",        "BOOLEAN NOT NULL", "INTEGER NOT NULL", "DEFAULT 1"),
        ("users",       "password_changed_at", "TIMESTAMP",        "TEXT",             None),
        ("attachments", "file_data",           "BYTEA",            "BLOB",             None),  # Railway DB storage
        ("backups",     "email_sent",          "BOOLEAN NOT NULL", "INTEGER NOT NULL", "DEFAULT 0"),
    ]
    with app.app_context():
        is_postgres = db.engine.dialect.name == "postgresql"
        with db.engine.connect() as conn:

            # ── Add missing columns ───────────────────────────────
            for table, col, pg_type, sqlite_type, default in REQUIRED_COLUMNS:
                default_clause = f" {default}" if default else ""

                if is_postgres:
                    # PostgreSQL supports IF NOT EXISTS natively
                    sql = (
                        f"ALTER TABLE {table} "
                        f"ADD COLUMN IF NOT EXISTS {col} {pg_type}{default_clause};"
                    )
                    try:
                        conn.execute(db.text(sql))
                        conn.commit()
                    except Exception as e:
                        app.logger.warning(f"ensure_columns: could not add {table}.{col} — {e}")
                else:
                    # SQLite: check if column exists first, then add if absent
                    try:
                        rows = conn.execute(db.text(f"PRAGMA table_info({table})")).fetchall()
                        existing_cols = {row[1] for row in rows}  # index 1 = column name
                        if col not in existing_cols:
                            sql = (
                                f"ALTER TABLE {table} "
                                f"ADD COLUMN {col} {sqlite_type}{default_clause};"
                            )
                            conn.execute(db.text(sql))
                            conn.commit()
                    except Exception as e:
                        app.logger.warning(f"ensure_columns: could not add {table}.{col} — {e}")

            # ── PostgreSQL GIN full-text index (TC-141) ───────────
            # Created here instead of __table_args__ to avoid crashing
            # db.create_all() on SQLite (to_tsvector is PostgreSQL-only).
            if is_postgres:
                try:
                    conn.execute(db.text(
                        "CREATE INDEX IF NOT EXISTS idx_tickets_fts ON tickets "
                        "USING GIN(to_tsvector('english', "
                        "coalesce(title,'') || ' ' || coalesce(description,'')));"
                    ))
                    conn.commit()
                except Exception as e:
                    app.logger.warning(f"ensure_columns: could not create idx_tickets_fts — {e}")


bootstrap_files()   # write templates/static to disk on first run

# Run db.create_all() + ensure_columns() at module level so both execute under Gunicorn/WSGI.
# (The if __name__=='__main__' block is only reached when running 'python app.py' directly.)
with app.app_context():
    try:
        db.create_all()          # create any missing tables (idempotent)
        ensure_columns()         # add any missing columns to existing tables
    except Exception as _ec_err:
        import warnings
        warnings.warn(f"startup db init failed: {_ec_err}", RuntimeWarning, stacklevel=1)

# ─────────────────────────────────────────────
# CLI: ensure-cols
# ─────────────────────────────────────────────
@app.cli.command("ensure-cols")
def ensure_cols_command():
    """
    Apply any new DB columns that are missing from an existing schema.
    Run this BEFORE starting Gunicorn in production whenever you deploy a
    new version that adds columns (instead of relying on the __main__ path).

    Usage: flask ensure-cols
    """
    ensure_columns()
    click.echo("✅ ensure_columns() completed.")


# ─────────────────────────────────────────────
# CLI: init-db
# ─────────────────────────────────────────────
@app.cli.command("init-db")
def init_db_command():
    """
    First-time Railway/production setup — run this once after deploying:

        flask init-db

    What it does (in order):
      1. db.create_all()   — creates every table that doesn't exist yet
      2. ensure_columns()  — adds any new columns to existing tables
      3. seed departments  — inserts the default departments if none exist

    Safe to re-run: create_all() and ensure_columns() are both idempotent.
    """
    with app.app_context():
        click.echo("⏳ Creating tables...")
        db.create_all()
        click.echo("✅ Tables created (or already exist).")

        click.echo("⏳ Ensuring columns...")
        ensure_columns()
        click.echo("✅ Columns verified.")

        # Seed default departments if none exist
        if Department.query.count() == 0:
            default_departments = [
                "Human Resources", "Information Technology", "Finance",
                "Operations", "Customer Service", "Administration",
            ]
            for name in default_departments:
                db.session.add(Department(name=name))
            db.session.commit()
            click.echo(f"✅ Seeded {len(default_departments)} default departments.")
        else:
            click.echo("✅ Departments already exist — skipped seeding.")

    click.echo("\n🎉 Database ready. Open the app and create your admin account.")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()   # safe for first run; after this use flask db upgrade
    ensure_columns()      # add any new columns missing from an existing DB
    app.run(debug=True, host="0.0.0.0", port=5000)
